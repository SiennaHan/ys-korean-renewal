#!/usr/bin/env python3
"""교재 콘텐츠 원장(xlsx) → 앱 데이터(JSON).

원장이 정본이다. 앱 JSON 은 산출물이므로 손으로 고치지 않는다 —
고치면 다음 생성에서 지워진다. 고칠 것은 xlsx 쪽이다.

    python3 scripts/build-content.py            # 저장소 루트의 최신 v*.xlsx
    python3 scripts/build-content.py --check    # 쓰지 않고 대조, 다르면 종료코드 1

원장은 교재 파생이라 저장소에 없다(.gitignore *.xlsx). 이 스크립트를 돌리려면
원장 파일이 저장소 루트에 있어야 한다.
"""

import argparse
import collections
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = Path(__file__).resolve().parents[1] / "src" / "shared" / "data"

# 시트 이름이 곧 파일 이름이다. 지금 화면이 읽는 9종
CONTENT_SHEETS = [
    "n1_word_list",
    "n1_word_quiz",
    "n2_ai_role_play",
    "n3_listen_script",
    "n3_listen_script_line",
    "n3_listen_repeat",
    "n4_blank_question",
    "n5_read_answer_text",
    "n5_read_answer_questions",
]

# 원장에는 있는데 앱이 아직 안 쓰던 것. 파일 이름을 따로 준다
EXTRA_SHEETS = {
    "n6_flashcard": "n6_flashcard.json",
    "n6_flashcard_card": "n6_flashcard_card.json",
    "문법목록": "grammar_list.json",
    # 자모 529행. 내용은 problem.ts 를 옮긴 것이라 화면이 보는 것이 안 바뀐다 —
    # 배관만 원장 쪽으로 돌린다. BLOCKERS.md §2
    "n8_jamo": "n8_jamo.json",
    # 117행 · 검수 완료(v29). ai_gender·ai_role·user_role 은 컬럼을 늘리지 않고
    # ai_persona_prompt 본문에서 뽑는다(파생 가능한 것은 옆에 안 둔다는 원칙) —
    # DERIVE_FROM_PROMPT 참고.
    "n7_mission_chat": "n7_mission_chat.json",
    # 354행 · 미션 슬롯마다 모범 문장 하나(5개 언어). 브리핑의 힌트 버튼이 쓴다 —
    # **대화 화면에는 힌트를 여는 길이 없다**(93e869f). 슬롯 수와 정확히 맞는다:
    # 칩 3개인 과 114 + 4개인 과 3 = 354. 그 짝이 어긋나면 hint_slot_match 가 멈춘다.
    "n7_mission_hint": "n7_mission_hint.json",
}

# ai_persona_prompt 본문에 박혀 있는 것을 정규식으로 뽑아 JSON 에 얹는다.
# 컬럼을 늘리지 않되, 화면이 바로 쓸 수 있게 한다(n3 의 voice CARRY_OVER 와 같은 결).
#
# **원장 v51 에서 프롬프트를 다시 쓰며 이 형식이 영어에서 한국어로 바뀌었다.**
# 옛 정규식(`**AI Gender:** female`)은 하나도 안 맞게 됐는데 **아무 데서도 안 걸렸다** —
# 파생이 빈 문자열이 되어도 조용히 통과했고, 서버 `_normGender` 가 빈 값을 female 로
# 떨어뜨려서 **남성 배역 35개가 여성 목소리로 말할 뻔했다**(2026-09-01 에 찾음).
# 그래서 아래 derive_from_prompt 가 **하나라도 못 뽑으면 실패로 멈춘다.**
DERIVE_FROM_PROMPT = {
    "ai_gender": r"- 당신의 성별:\s*(.+)",
    "ai_role": r"- 당신의 역할:\s*(.+)",
    "user_role": r"- 사용자의 역할:\s*(.+)",
}
# 프롬프트는 한국어로 적고 앱·서버는 영어 값을 쓴다
GENDER_KO = {"여성": "female", "남성": "male", "여": "female", "남": "male"}

# 컬럼 이름을 JSON 키로 쓸 수 없는 것만 바꾼다
KEY_FIX = {"grammar_tag(대표형)": "grammar_tag", "급": "level", "분류": "category",
           "도입_과": "intro_chapter", "문항수": "item_count",
           "grammar_focus_수정": "grammar_focus_revised", "오류메모": "error_note"}

# 원장에 없고 앱에만 있던 열. id 로 이어 붙인다
CARRY_OVER = {"n3_listen_script_line": ["voice"]}


LEDGER_RE = re.compile(r"^글로벌_교재기반_콘텐츠_v(\d+)\.xlsx$")


def newest_ledger() -> Path:
    """가장 높은 v 번호를 고른다.

    한글 이름으로 glob 하지 않는다 — macOS 는 파일마다 유니코드 정규화 형태가
    달라질 수 있어서, 한 형태로 짠 패턴이 멀쩡히 있는 파일을 조용히 건너뛴다.
    실제로 v23 을 놓치고 v22 를 정본으로 삼은 적이 있다. 이름을 NFC 로 맞춘 뒤 비교한다.
    """
    found = []
    for p in ROOT.glob("*.xlsx"):
        m = LEDGER_RE.match(unicodedata.normalize("NFC", p.name))
        if m:
            found.append((int(m.group(1)), p))
    if not found:
        sys.exit(f"원장을 찾지 못했다: {ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
    return max(found)[1]


def cell(v):
    """빈 칸은 빈 문자열. 정수로 떨어지는 실수는 정수로 (엑셀이 1 을 1.0 으로 준다)"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


# ── 검수 상태 ────────────────────────────────────────────────────────────────
#
# **G1 은 "reviewed 만 학생에게 낸다" 로 정했는데 그대로는 성립하지 않는다.**
# 원장 v57 을 세어 보면 `reviewed` 는 1,354행이고 `draft` 가 9,124행이다.
# 미션 대화 117행에는 `reviewed` 가 **하나도 없다**(fixed_v29 103 · fixed_v56 14).
# 그대로 걸면 **이미 배선까지 끝난 화면이 통째로 빈다.** 실제 상태 값이 두 갈래가
# 아니라 여럿이기 때문이다 — v57 기준으로 실제 쓰이는 값은 이렇다:
#
#   auto_checked · draft · reviewed · deleted
#   fixed_v14~v18 · fixed_v22/v24 · fixed_v29 · fixed_v56
#   authored_v21/v23 · tagged_v20 · filled_v19 · added_v17
#
# `draft_v52/v53/v55` 는 v57 에서 사라졌다 — 힌트 340행이 `reviewed` 가 됐다.
# 다만 **`KNOWN_STATUS` 에서는 빼지 마라.** 옛 원장으로 되돌려 돌릴 때 경고만 늘고
# 막을 이유는 없다.
#
# (자모 529행은 v41 에서 전부 `reviewed` 가 됐다 — 아래 PORTED_AS_IS 주석 참고.)
#
# 그래서 **내보내지 않을 것만 정한다**(기획 확정 2026-08-28 — "예외 상태를 둔다").
# 나머지는 낸다. 저작이 진행 중인 것을 학생이 보는 것과, 지운 것을 학생이 보는 것은
# 무게가 다르다 — 뒤쪽만 막는다.
DROP_STATUS = {"deleted"}

# 위 목록에 없는 값이 나오면 **경고하되 포함한다.** 현재 공개 정책은 삭제 표시만
# 확실히 제외하는 것이며, 새 상태가 생긴 사실을 조용히 넘기지 않도록 경고한다.
# 기계 판독 정본은 docs/project_status.json 이다.
KNOWN_STATUS = {
    "auto_checked", "draft", "reviewed", "deleted",
    "tagged_v20", "filled_v19", "added_v17", "authored_v21", "authored_v23",
} | {f"fixed_v{n}" for n in range(10, 60)} | {f"draft_v{n}" for n in range(40, 60)}

# 전에 여기 `PORTED_AS_IS` 가 있었다 — 자모 529행이 전부 `draft` 인데 그게 "저작 전"
# 이라는 뜻이 아니라는 것을 이름으로 적어 두는 예외였다. **2026-08-28 에 없앴다** —
# 원장 v41 에서 그 529행이 `reviewed` 가 됐다(받침·겹받침 포함 검수 완료).
# 예외로 덮는 대신 **원장이 사실을 말하게 한 것**이라 여기 남길 것이 없다.


def drop_unshippable(sheet, rows):
    """내보내지 않을 행을 걸러 낸다. (남은 행, 뺀 수, 처음 보는 상태) 를 준다."""
    if not rows or "review_status" not in rows[0]:
        return rows, 0, set()
    kept, dropped, unknown = [], 0, set()
    for r in rows:
        st = str(r.get("review_status") or "").strip()
        if st and st not in KNOWN_STATUS:
            unknown.add(st)
        if st in DROP_STATUS:
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped, unknown


def read_sheet(ws):
    """원장 시트 한 장을 (헤더, 행들) 로 읽는다.

    **행을 헤더 길이까지 늘려서 준다.** openpyxl 은 read_only 로 읽을 때 행 끝의
    빈 칸을 잘라서 주므로, 그냥 `zip(head, r)` 하면 **뒤쪽 열의 열쇠가 행마다
    있다 없다 한다.** `cell()` 은 「빈 칸은 빈 문자열」이라고 약속하는데 그 약속이
    맨 오른쪽 열에서만 조용히 깨지는 것이다.

    실제로 원장 v56 에서 이것이 터졌다 — `n8_jamo` 529행이 헤더 27칸인데 행은
    26칸으로 와서 `hold_reason` 이 **JSON 에서 통째로 사라졌다.** `jamo.ts` 와
    `mission-chat.ts` 는 그 열을 `hold_reason: string` 으로 선언해 두었으므로
    선언이 거짓이 됐고, 그런데도 타입 검사는 통과했다(JSON 을 캐스팅해 읽는다).
    값을 잃지는 않았다 — 잘린 칸은 그 행에서 비어 있던 칸이다. 잃은 것은 **모양의
    일관성**이고, 그래서 원장을 어떻게 저장했느냐에 따라 산출물 diff 가 2만 줄씩
    흔들려 **진짜 바뀐 곳이 묻힌다.**
    """
    it = ws.iter_rows(values_only=True)
    head = [KEY_FIX.get(str(h).strip(), str(h).strip()) if h is not None else "" for h in next(it)]
    rows = []
    for r in it:
        if not any(x is not None and str(x).strip() for x in r):
            continue  # 원장 끝의 빈 줄
        r = tuple(r) + (None,) * (len(head) - len(r))
        rows.append({k: cell(v) for k, v in zip(head, r) if k})
    return head, rows


def int_columns(rows, prev):
    """어느 열을 숫자로 둘지. 이미 있던 JSON 이 정본이고, 새 열은 값으로 정한다"""
    cols = set()
    for key in rows[0]:
        if prev and key in prev[0]:
            if isinstance(prev[0][key], int):
                cols.add(key)
            continue
        vals = [r[key] for r in rows if r[key] != ""]
        if vals and all(isinstance(v, int) for v in vals):
            cols.add(key)
    return cols


def coerce(rows, int_cols):
    for r in rows:
        for k, v in r.items():
            if k in int_cols:
                r[k] = v if isinstance(v, int) else (int(v) if str(v).strip().lstrip("-").isdigit() else 0)
            elif not isinstance(v, str):
                r[k] = str(v)
    return rows


def derive_from_prompt(sheet, rows):
    """ai_persona_prompt 본문에서 ai_gender·ai_role·user_role 을 뽑아 얹는다.

    n7_mission_chat 전용. 컬럼을 늘리지 않기로 한 결정(v27) 때문에 이 셋은
    프롬프트 텍스트 안에만 있다 — 화면이 매번 정규식을 돌리게 두지 않고
    생성 시점에 한 번 뽑아 둔다. 117행 전량 매치 확인됨(2026-09-01 재확인 — 형식이 한국어로 바뀌어 정규식을 고쳤다).
    **하나라도 못 뽑으면 멈춘다.** 조용한 빈 값이 목소리를 바꿔 놓는다.
    """
    if sheet != "n7_mission_chat":
        return
    miss = []
    for r in rows:
        p = str(r.get("ai_persona_prompt", ""))
        for key, pat in DERIVE_FROM_PROMPT.items():
            m = re.search(pat, p)
            if not m:
                miss.append(f"{r.get('item_id')} · {key}")
                r[key] = ""
                continue
            val = m.group(1).strip()
            if key == "ai_gender":
                if val not in GENDER_KO:
                    miss.append(f"{r.get('item_id')} · 모르는 성별 값 {val!r}")
                    r[key] = ""
                    continue
                val = GENDER_KO[val]
            r[key] = val
    if miss:
        raise SystemExit(
            f"★ ai_persona_prompt 에서 파생 값을 못 뽑았다 — {len(miss)}곳\n"
            f"   {', '.join(miss[:6])}{' …' if len(miss) > 6 else ''}\n"
            "   프롬프트 형식이 바뀌면 DERIVE_FROM_PROMPT 를 같이 고쳐라.\n"
            "   **조용히 빈 값으로 두면 안 된다** — 서버가 빈 성별을 female 로 떨어뜨려\n"
            "   남성 배역이 여성 목소리로 말한다(2026-09-01 에 실제로 그랬다)."
        )


def carry(sheet, rows, prev):
    """원장에 없는 앱 전용 열을 id 로 이어 붙인다"""
    keys = CARRY_OVER.get(sheet)
    if not keys or not prev:
        return 0
    old = {r.get("id"): r for r in prev}
    derived = 0
    for r in rows:
        src = old.get(r.get("id"))
        for k in keys:
            if src and src.get(k) not in (None, ""):
                r[k] = src[k]
            else:
                # 새로 생긴 행 — 성별에서 뽑는다
                r[k] = "male" if str(r.get("gender", "")).strip() == "남" else "female"
                derived += 1
    return derived


def odd_quotes(sheet, rows):
    """따옴표가 홀수인 한글 칸을 찾는다.

    엑셀은 셀 맨 앞의 ' 를 "이건 텍스트다" 라는 서식 지시로 읽고 값에서 지운다.
    그래서 저작자가 '연필'은 … 이라고 써도 연필'은 … 으로 저장된다.
    실제로 n4_blank_question 의 해설 61개가 그 상태였다(2026-08-21 복원).

    엑셀에서 그 칸을 다시 타이핑하면 또 먹히므로 생성할 때마다 본다.
    영어 아포스트로피(I'm · o'clock · one's)는 정상이라, 한글이 절반을 넘는
    칸만 센다.
    """
    out = []
    for i, row in enumerate(rows, start=2):
        for key, val in row.items():
            if not isinstance(val, str) or val.count("'") % 2 == 0:
                continue
            hangul = sum(1 for ch in val if "가" <= ch <= "힣")
            letters = sum(1 for ch in val if ch.isalpha())
            if letters and hangul / letters > 0.5:
                out.append((i, key, val))
    return out


# 그 행이 무엇인지 말해 주는 열. **번호가 밀렸는지 보려면 내용을 봐야 한다**
IDENTITY = {
    "n1_word_list": "word", "n1_word_quiz": "prompt", "n2_ai_role_play": "ko",
    "n3_listen_script": "audio_text", "n3_listen_script_line": "text",
    "n3_listen_repeat": "question", "n4_blank_question": "question",
    "n5_read_answer_text": "text", "n5_read_answer_questions": "question",
    "n6_flashcard": "set_title", "n6_flashcard_card": "word",
    "n7_mission_chat": "scenario_title", "n8_jamo": "target_word",
    "n7_mission_hint": "hint_ko",
}

# 그 시트의 **행 하나를 가리키는 열**. 기본은 item_id 다.
#
# 힌트 시트만 다르다 — item_id 는 미션 대화(`MC-1-04-001`)를 가리켜 3~4행이 나눠 쓰고,
# 행을 가리키는 것은 `hint_id`(`MH-1-04-001-1`)다. 이것을 안 갈라 두면 renumbered() 가
# 한 과의 힌트 서넛을 같은 열쇠로 뭉개서 **밀림 검사가 조용히 헛돈다.**
KEY = {"n7_mission_hint": "hint_id"}

ID_MAX = 12          # `FCW-3-12-004` 가 12자다. 학습 기록의 card_id 도 이 폭에 맞춘다


def mission_slots(detail):
    """`mission_detail` 을 라벨 목록으로. **앱의 `parseMissionDetail` 과 같은 규칙이다.**

    앱은 `" / "` 를 **조건 없이** 자르고 첫 `:` 앞을 라벨로 본다
    (`app/src/shared/data/mission-chat.ts`). 여기서 더 똑똑하게 자르면 검사만
    통과하고 화면은 딴 것을 그린다 — 그러니 규칙을 일부러 그대로 베낀다.
    """
    out = []
    for chunk in (detail or "").split(" / "):
        i = chunk.find(":")
        label = (chunk if i < 0 else chunk[:i]).strip()
        if label:
            out.append(label)
    return out


def hint_slot_mismatch(chat_rows, hint_rows):
    """힌트가 미션 슬롯과 **자리까지** 맞나. (어긋난 것, 힌트가 아예 없는 과) 를 준다.

    **왜 자리까지 보나.** 브리핑은 `keywords` 와 `hints` 를 **같은 순서로** 짝지어
    그린다(`BriefingContent.hints` 주석). 순서가 어긋나면 「이름」 밑에 인사 문장이
    붙는데 **화면은 멀쩡해 보인다** — 2026-09-01 에 구 앱 덤프가 정확히 그 꼴이었다
    (`이름 → "Say hello."`, 117과 중 109과). 그때는 사람이 눈으로 찾았다.

    그래서 라벨을 **위치별로** 견준다. 개수만 세면 순서가 바뀐 것을 못 잡는다.
    """
    want = {r.get("item_id"): mission_slots(r.get("mission_detail"))
            for r in chat_rows if r.get("item_id")}
    got = {}
    for r in hint_rows:
        got.setdefault(r.get("item_id"), []).append(r)
    bad, missing = [], []
    for item_id, labels in sorted(want.items()):
        rows = sorted(got.get(item_id, []), key=lambda r: r.get("slot_seq") or 0)
        if not rows:
            missing.append(item_id)
            continue
        if len(rows) != len(labels):
            bad.append((item_id, f"슬롯 {len(labels)}개인데 힌트 {len(rows)}개"))
            continue
        for i, (label, r) in enumerate(zip(labels, rows), start=1):
            if str(r.get("mission_label") or "").strip() != label:
                bad.append((item_id,
                            f"{i}번째 — 미션은 「{label}」인데 힌트는 「{r.get('mission_label')}」"))
                break
            if not str(r.get("hint_ko") or "").strip():
                bad.append((item_id, f"{i}번째 「{label}」의 hint_ko 가 비었다"))
                break
    # 힌트 시트에만 있고 미션에는 없는 과 — 지운 미션의 힌트가 남은 것
    orphan = sorted(set(got) - set(want))
    return bad, missing, orphan


def renumbered(sheet, rows, prev):
    """**같은 `item_id` 가 딴 것을 가리키게 됐나** — 행을 지우면서 번호를 다시 매긴 것.

    `item_id` 는 `FCW-3-12-004` 처럼 **과 안의 자리 번호**다. 가운데 행을 지우고
    번호를 다시 매기면 그 뒤가 전부 한 칸씩 밀린다. 2026-08-31 에 실제로 났다 —
    3급 12과에서 「면허증」을 지우자 `FCW-3-12-004` 가 「자가용」에서 「중고차」가 됐다.

    **왜 위험한가.** 학습 기록(`ko_learning_record.question_id`)·복습 큐·플래시카드의
    「알아요/몰라요」가 그 번호로 문항을 가리킨다. 밀리면 **맞힌 기록이 다른 문항에
    붙는다.** 채점 시트는 지금까지 삭제가 없어서 안 터졌을 뿐이다.

    **정한 것(2026-08-31) — 행을 지울 때 번호를 다시 매기지 않는다.** 빈 번호를
    남긴다. 그러면 `FCW-3-12-003` 은 영영 「면허증」의 자리이고 아무도 그것을 쓰지 않는다.

    문항을 **다시 쓰는 것**도 내용이 바뀌는 일이라(v48 의 읽기 오답 수정) 기계가 둘을
    못 가른다. 그래서 **과의 행 수가 줄어든 자리만** 밀림으로 본다 — 빈 번호를 남기면
    행 수는 줄어도 남은 번호의 뜻은 안 바뀌므로 여기 안 걸린다.
    """
    col = IDENTITY.get(sheet)
    if not col or not prev:
        return []
    key = KEY.get(sheet, "item_id")
    was = {r.get(key): r.get(col) for r in prev if r.get(key)}
    def chapters(rs):
        out = {}
        for r in rs:
            k = (r.get("book_id"), r.get("chapter"))
            out[k] = out.get(k, 0) + 1
        return out
    before, after = chapters(prev), chapters(rows)
    shrunk = {k for k, n in before.items() if after.get(k, 0) < n}
    hits = []
    for r in rows:
        k = r.get(key)
        if k in was and str(r.get(col) or "") != str(was[k] or "") \
           and (r.get("book_id"), r.get("chapter")) in shrunk:
            hits.append((k, was[k], r.get(col)))
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=None)
    ap.add_argument("--check", action="store_true", help="쓰지 않고 차이만 본다")
    args = ap.parse_args()

    ledger = args.xlsx or newest_ledger()
    print(f"원장  {ledger.name}")
    print(f"대상  {OUT_DIR.relative_to(ROOT)}\n")
    wb = openpyxl.load_workbook(ledger, read_only=True, data_only=True)

    warnings: list[tuple[str, int, str, str]] = []
    warnings_status: list[tuple[str, list[str]]] = []
    renumber_warnings: list[tuple[str, list]] = []
    long_ids: list[tuple[str, list[str]]] = []
    dropped_total = 0
    # 시트 → (상태별 행 수, 포함, 제외). **원장 상태를 눈에 보이게 하는 유일한 자리다.**
    manifest: dict[str, tuple[collections.Counter, int, int]] = {}
    built: dict[str, list] = {}   # 시트끼리 견주는 검사용 (아래 hint_slot_mismatch)
    drifted: list[str] = []       # --check 에서 원장과 다른 산출물
    plan = [(s, f"{s}.json") for s in CONTENT_SHEETS] + list(EXTRA_SHEETS.items())
    for sheet, filename in plan:
        if sheet not in wb.sheetnames:
            print(f"  {sheet:<26} 원장에 없다 — 건너뜀")
            continue
        head, rows = read_sheet(wb[sheet])
        if not rows:
            print(f"  {sheet:<26} 내용이 비었다 — 건너뜀")
            continue

        path = OUT_DIR / filename
        prev = json.loads(path.read_text(encoding="utf-8")) if path.exists() else None
        # **거르기 전에** 센다 — 제외된 행도 매니페스트에는 나와야 한다
        if "review_status" in rows[0]:
            manifest[sheet] = (
                collections.Counter(str(r.get("review_status") or "").strip() or "(빈 칸)"
                                    for r in rows),
                0, 0)
        rows, dropped, unknown = drop_unshippable(sheet, rows)
        if sheet in manifest:
            manifest[sheet] = (manifest[sheet][0], len(rows), dropped)
        if unknown:
            warnings_status.append((sheet, sorted(unknown)))
        derive_from_prompt(sheet, rows)
        rows = coerce(rows, int_columns(rows, prev))
        derived = carry(sheet, rows, prev)

        moved = renumbered(sheet, rows, prev)
        if moved:
            renumber_warnings.append((sheet, moved))
        too_long = sorted({r["item_id"] for r in rows
                           if len(str(r.get("item_id") or "")) > ID_MAX})
        if too_long:
            long_ids.append((sheet, too_long))

        built[sheet] = rows
        text = json.dumps(rows, ensure_ascii=False, indent="\t") + "\n"
        before = len(prev) if prev else 0
        delta = f"{before} → {len(rows)}" if before != len(rows) else f"{len(rows)}"
        new_cols = [k for k in rows[0] if not prev or k not in prev[0]]
        dropped_total += dropped
        note = f" · 뺀 행 {dropped}" if dropped else ""
        note += f" · 새 열 {len(new_cols)}" if new_cols else ""
        note += f" · voice 추정 {derived}" if derived else ""
        same = path.exists() and path.read_text(encoding="utf-8") == text
        if not same:
            drifted.append(filename)
        mark = "같음" if same else ("쓸 것" if args.check else "썼음")
        if not args.check and not same:
            path.write_text(text, encoding="utf-8")
        print(f"  {sheet:<26}{delta:>14}행  {mark}{note}")

        # 엑셀이 여는 따옴표를 먹었을 수 있다 — 위 odd_quotes 주석 참조
        for line, key, val in odd_quotes(sheet, rows):
            warnings.append((sheet, line, key, val))

    # ── 검수 상태 매니페스트 ──────────────────────────────────────────
    #
    # **왜 찍나(DEV-07).** 게이트는 「상태값이 정책대로 처리되는가」만 볼 수 있다.
    # 「그 상태값이 사실인가」는 못 본다 — 사람이 검수를 하고도 도장을 안 찍으면
    # 아무것도 울지 않는다. 원장 v56 에서 실제로 그랬다: 354슬롯을 전수검수했는데
    # 도장은 고친 14행에만 찍혀서, 나머지 340행이 `draft_v52` 를 달고 나갔다.
    # **그때 이 표가 있었으면 `draft_v52 307` 이 그 자리에서 눈에 걸렸다.**
    #
    # 기계는 검수 여부를 모른다. 하지만 **「이 시트에 검수 전 상태가 몇 행 있다」**
    # 는 말할 수 있고, 그거면 사람이 알아챈다. 그래서 막지 않고 보여만 준다.
    #
    # **파일로 안 남긴다.** 원장이 저장소에 없어 CI 가 이 스크립트를 못 돌리므로,
    # 파일로 두면 아무도 다시 뽑지 않아 조용히 낡는다 — 낡은 파일은 없느니만 못하다.
    if manifest:
        print("\n검수 상태 — 원장이 무엇을 검수됐다고 말하는가")
        for sheet, (counts, kept, dropped) in manifest.items():
            parts = " · ".join(f"{st} {n}" for st, n in counts.most_common())
            tail = f" · 제외 {dropped}" if dropped else ""
            print(f"  {sheet:<26} {parts}")
            print(f"  {'':<26} → 포함 {kept}{tail}")

    if dropped_total:
        print(f"\n원장에서 지운 행 {dropped_total}개를 내보내지 않았다 (review_status = deleted)")
    if warnings_status:
        print("\n⚠️  처음 보는 review_status — 내보내기는 했지만 규칙에 없다")
        print("    scripts/build-content.py 의 KNOWN_STATUS 에 넣거나, 막을 것이면 DROP_STATUS 로.")
        for sheet, vals in warnings_status:
            print(f"    {sheet}: {', '.join(vals)}")
    if warnings:
        print(f"\n⚠️  따옴표가 홀수인 한글 칸 {len(warnings)}개 — 엑셀이 여는 \' 를 먹었을 수 있다")
        print("    셀 맨 앞의 \' 는 서식 지시로 읽혀 값에서 사라진다. 원장에서 되돌려라.")
        for sheet, line, key, val in warnings[:10]:
            print(f"    {sheet} {line}행 {key}: {val[:56]}")
        if len(warnings) > 10:
            print(f"    … 그 밖에 {len(warnings) - 10}개")

    # n7_mission_chat·n8_jamo 둘 다 배선됐다(2026-08-24) — 이 안내는 더 필요 없다.
    # 새로 원장에 시트가 생기고 아직 안 쓴다면 여기에 같은 모양으로 다시 둔다.

    # ── 번호가 밀렸나 · 번호가 길어졌나 ───────────────────────────────
    #
    # 둘 다 **막는다(종료코드 1)**. 경고로 두면 다음 사람이 지나간다 —
    # 그리고 이 둘은 지나가면 학습 기록이 조용히 어긋나는 종류다.
    if renumber_warnings:
        print("\n❌ 같은 item_id 가 딴 것을 가리킨다 — **행을 지우며 번호를 다시 매긴 것 같다**")
        print("   정한 것(2026-08-31): 행을 지울 때 번호를 다시 매기지 마라. **빈 번호를 남겨라.**")
        print("   그래야 학습 기록·복습 큐·플래시카드의 「알아요」가 딴 문항에 안 붙는다.")
        for sheet, moved in renumber_warnings:
            print(f"   {sheet} — {len(moved)}개")
            for k, was, now in moved[:5]:
                print(f"     {k}  {str(was)[:18]!r}  →  {str(now)[:18]!r}")
            if len(moved) > 5:
                print(f"     … 외 {len(moved) - 5}개")
    if long_ids:
        print(f"\n❌ item_id 가 {ID_MAX}자를 넘는다 — 학습 기록의 card_id 가 그 폭이라 잘린다")
        for sheet, ids in long_ids:
            print(f"   {sheet}: {', '.join(ids[:6])}")

    # ── 힌트가 미션 슬롯과 자리까지 맞나 ───────────────────────────────
    hint_bad, hint_missing, hint_orphan = [], [], []
    if "n7_mission_chat" in built and "n7_mission_hint" in built:
        hint_bad, hint_missing, hint_orphan = hint_slot_mismatch(
            built["n7_mission_chat"], built["n7_mission_hint"])
        total = sum(len(mission_slots(r.get("mission_detail")))
                    for r in built["n7_mission_chat"])
        print(f"\n힌트  미션 슬롯 {total}개 · 힌트 {len(built['n7_mission_hint'])}행"
              f" · 힌트 없는 과 {len(hint_missing)}")
    if hint_bad:
        print("\n❌ 힌트가 미션 슬롯과 어긋난다 — **브리핑이 라벨과 문장을 자리로 짝짓는다**")
        print("   어긋나도 화면은 멀쩡해 보인다. 「이름」 밑에 인사 문장이 붙을 뿐이다.")
        for item_id, why in hint_bad[:8]:
            print(f"   {item_id}: {why}")
        if len(hint_bad) > 8:
            print(f"   … 외 {len(hint_bad) - 8}개")
    if hint_orphan:
        print("\n❌ 미션에 없는 과의 힌트가 남아 있다 — 미션을 지우고 힌트를 안 지운 것 같다")
        print(f"   {', '.join(hint_orphan[:8])}")

    if renumber_warnings or long_ids or hint_bad or hint_orphan:
        return 1
    if args.check and drifted:
        print(
            f"\n❌ 원장과 다른 산출물 {len(drifted)}개 — "
            "`--check` 이므로 쓰지 않았고 실패로 끝낸다"
        )
        for filename in drifted:
            print(f"   {filename}")
        print("   검수 뒤 `python3 scripts/build-content.py` 로 다시 만들고 변경을 커밋해라.")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main() or 0)
