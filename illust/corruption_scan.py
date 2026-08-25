#!/usr/bin/env python3
"""엑셀 한국어 텍스트의 일괄 치환 손상을 찾는다.

증상: 한 글자를 다른 문자열로 바꾸면서 앞 글자와 공백까지 먹어 버렸다.
  이용할 수 있어요   → 이이용할 있어요      (수 → 이, 한 칸 밀림)
  좋았는데 그 여자는  → 좋았는다른데 여자는   (그 → 다른, 한 칸 밀림)
  뜻이 있는 곳에      → 뜻이 있하는에        (곳 → 하는, 한 칸 밀림)
  제가 지난번에       → 제여러난번에         (지 → 여러, 한 칸 밀림)
  파는 곳인데         → 파하는인데

내용 변경과 헷갈리면 안 된다. 판별 기준은 '신판에도 구판에도 없는 문자열'이다.
교재가 바뀐 것이면 최소한 구판에는 있어야 한다. 양쪽 다 없으면 데이터 손상이거나
자체 창작이다.

산출: verify/corruption.csv
"""
import os, csv, re, json, collections, unicodedata
import fitz
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v12.xlsx"
OUT = f"{HERE}/verify"

# 손상 흔적 — 정상 한국어에 나올 수 없는 붙음
MARKS = [
    (re.compile(r"이이용할"), "이용할 수"),
    (re.compile(r"는다른데"), "는데 그"),
    (re.compile(r"ㄴ다른데|은다른데|런다른데"), "런데 그"),
    (re.compile(r"있하는에"), "있는 곳에"),
    (re.compile(r"제여러난번"), "제가 지난번"),
    (re.compile(r"파하는인데"), "파는 곳인데"),
    (re.compile(r"많았는다른데렇게"), "많았는데 그렇게"),
]
GENERIC = re.compile(r"[가-힣](다른|하는|여러)[가-힣]{0,2}(데|에|번)")


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    blobs = {}
    for b in range(1, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        old = fitz.open(f"{BASE}/{b}_yonsei3week_main.pdf")
        blobs[b] = (sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1))),
                    sq(" ".join(old[p].get_text() for p in range(old.page_count))))

    SHEETS = {"n2_ai_role_play": "ko", "n5_read_answer_text": "text",
              "n3_listen_script_line": "text", "n1_word_list": "word",
              "n4_blank_question": "question"}
    out = []
    for name, field in SHEETS.items():
        ws = wb[name]
        hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        if field not in hdr:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdr, r))
            b = d.get("book_id")
            txt = str(d.get(field) or "")
            if not b or len(txt) < 6:
                continue
            hit = next((m for pat, m in MARKS if pat.search(txt)), None)
            if not hit and not GENERIC.search(txt):
                continue
            k = sq(txt)[:24]
            innew, inold = k in blobs[int(b)][0], k in blobs[int(b)][1]
            if innew or inold:
                continue          # 교재에 있으면 손상이 아니다
            out.append(dict(sheet=name, item_id=d.get("item_id"), book=b,
                            chapter=d.get("chapter"), guess=hit or "패턴 의심",
                            text=txt[:120]))

    with open(f"{OUT}/corruption.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["sheet", "item_id", "book", "chapter",
                                           "guess", "text"])
        wr.writeheader()
        wr.writerows(out)
    print(f"손상 의심 {len(out)}건")
    print(" 시트별:", dict(collections.Counter(o["sheet"] for o in out)))
    print(" 급별:", dict(sorted(collections.Counter(o["book"] for o in out).items())))
    for o in out[:25]:
        print(f"  {o['item_id']} ({o['book']}급{o['chapter']}과) {o['text'][:78]}")
    print(f"-> {OUT}/corruption.csv")


if __name__ == "__main__":
    main()
