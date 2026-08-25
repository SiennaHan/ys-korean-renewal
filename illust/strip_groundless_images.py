#!/usr/bin/env python3
"""캡션(vocab_label) 없는 그림에 좌표만으로 걸린 어휘 이미지 40건을 지운다.

사용자가 '안녕하다·반갑다에 왜 인사 장면이 나오냐' 고 지적해서 다시 봤다 —
그 그림들은 어휘 그림면이 아니라 **본문 대화 삽화**(캡션 없음)였고, 옛 원장이
그 단어가 나온 페이지에서 가장 가까운 그림을 좌표로만 물어 붙인 것이었다.
글로벌판으로 다시 연결(apply_images.py)하면서 그 근거 없는 연결을 그대로
이어받았다 — 좌표 재계산은 옳게 했지만, 원래 연결 자체가 틀렸다.

기준: catalog.csv 의 vocab_label 이 비어 있는(=그림 아래 캡션이 전혀 없는)
그림에 links.csv 가 '좌표만' 으로만 걸어 둔 것. 캡션이 있는데 다른 단어로
읽힌 12건(예: 교과서→'책', 시청→'시청/호텔 앞')은 캡션에 그 단어가 실제로
있으므로 여기서 지우지 않는다 — 정상이다.

산출: 글로벌_교재기반_콘텐츠_v{n+1}.xlsx (image 값을 비운다)
"""
import csv, glob, os, re, shutil, sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"


def newest_ledger():
    found = [(int(m.group(1)), p) for p in glob.glob(f"{ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
              if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(found)


def main():
    by_fn = {r["filename"]: r for r in csv.DictReader(open(f"{HERE}/catalog.csv"))}
    links = [r for r in csv.DictReader(open(f"{HERE}/links.csv"))
              if r["sheet"] == "n1_word_list" and r["how"] == "좌표만"]
    groundless = {r["item_id"]: r for r in links
                  if by_fn.get(r["filename"]) and not by_fn[r["filename"]]["vocab_label"]}
    print(f"근거 없는 연결 {len(groundless)}건")

    n, src = newest_ledger()
    dst = f"{ROOT}/글로벌_교재기반_콘텐츠_v{n+1}.xlsx"
    if os.path.exists(dst):
        sys.exit(f"중단: {dst} 가 이미 있다.")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    ws = wb["n1_word_list"]
    hdr = [c.value for c in ws[1]]
    C = {h: i + 1 for i, h in enumerate(hdr) if h}
    cleared = 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        link = groundless.get(iid)
        if not link:
            continue
        old = ws.cell(row, C["image"]).value
        ws.cell(row, C["image"]).value = ""
        note = ws.cell(row, C["change_note"]).value or ""
        add = (f"이미지 제거 — '{old}'는 어휘 그림이 아니라 캡션 없는 본문 대화 삽화였다"
               f"(좌표만으로 상속된 연결, 실제 근거 없음)")
        ws.cell(row, C["change_note"]).value = (note + " / " + add) if note else add
        cleared += 1

    lg = wb.create_sheet(f"99_변경내역_v{n+1}")
    lg.append(["구분", "대상", "내용"])
    lg.append(["이미지 제거", "n1_word_list", f"{cleared}건 — 캡션 없는 본문 삽화에 좌표만으로 "
                                              "걸려 있던 것(안녕하다·반갑다·숫자 하나~열·"
                                              "추상 형용사 등). 틀린 그림보다 없는 게 낫다"])
    wb.save(dst)
    print(f"{cleared}건 제거 -> {dst}")


if __name__ == "__main__":
    main()
