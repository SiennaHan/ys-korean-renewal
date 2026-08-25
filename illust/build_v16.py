#!/usr/bin/env python3
"""v16 — 신판에서 다시 쓰인 듣기 대본 2건 반영.

두 대본은 앞서 '신판에 없음'으로 잘못 잡혔던 것들이다. 실제로는 신판에 있고,
앵커로 삼은 문구 자체가 바뀌어서(판다곰→판다) 못 찾았을 뿐이다.

5급 8과 — 다크 서클. 음식 목록이 줄고 '연어를 눈 밑에 올려 놓으면 더 좋습니다'가
          통째로 빠졌다. 그 문장을 묻던 OX 문항이 근거를 잃으므로 함께 바꾼다.
8급 3과 — 대중문화. 일제 강점기~90년대~싸이 빌보드 서술이
          케이팝 30년사(서태지→SM→보아→소녀시대→싸이→방탄소년단)로 전면 재작성됐다.
          '일제 강점기'를 묻던 문항도 근거를 잃어 새 지문에 맞춰 다시 만든다.

산출: 글로벌_교재기반_콘텐츠_v16.xlsx
"""
import re, shutil, datetime, unicodedata
import openpyxl
from global_text import GlobalPdf

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v15.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v16.xlsx"
APX = "/Users/soohyeon/Documents/2608-yonsei_renewal/book"
TARGET = {(5, "210"), (8, "340")}

LINES = {
 "LL-5-8-001":
    "피곤하면 눈 밑이 판다처럼 어두워지시나요? 어떻게 하면 이 ‘다크 서클’을 없앨 수 있을까요? "
    "비타민 A와 C가 많이 들어 있는 음식을 먹으면 좋다고 합니다. "
    "당근, 양배추, 브로콜리, 레몬, 연어 등이 좋습니다. "
    "무엇보다도 ‘다크 서클’은 대부분 피곤해서 생기기 때문에 잠을 충분히 자고 푹 쉬어야 합니다.",
 "LL-8-3-001":
    "1990년대부터 케이팝의 시대가 본격화되었습니다. 1993년 그룹 <서태지와 아이들>을 시작으로 "
    "랩 음악과 힙합 댄스가 큰 인기를 얻었고, 1996년 SM 엔터테인먼트가 출범하여 "
    "HOT(에이치오티), SES(에스이에스), 신화를 연이어 성공시키며 아이돌 음악의 새로운 장을 열었습니다. "
    "잇따라 YG(와이지)와 JYP(제이와이피) 엔터테인먼트가 설립되면서 이른바 3대 기획사를 중심으로 "
    "한국의 아이돌 음악 시장이 폭발적으로 성장했습니다. "
    "2000년대 초에는 보아와 동방신기가 일본 시장에 진출해서 큰 인기를 얻었습니다. "
    "당시 세계적으로 위상이 높았던 일본 음악 시장에서 한국 가수가 거둔 성공은 매우 특별한 성과였습니다. "
    "보아가 활동을 시작한 2001년경부터 일본에서는 제이팝에 대응하여 케이팝이라는 용어가 "
    "사용되기 시작했습니다. 이 명칭이 점차 세계적으로 퍼져 나가 현재까지 이어지고 있는 겁니다. "
    "2000년대 후반에는 카라, 소녀시대, 빅뱅, 샤이니 등 다양한 아이돌 그룹이 일본의 10대와 "
    "20대들에게 큰 사랑을 받으면서 케이팝이 하나의 음악 장르로 확실히 자리 잡았습니다. "
    "2010년대에는 디지털 음원이 보편화되고 유튜브, 페이스북 같은 소셜 네트워크 서비스(SNS)가 "
    "전 세계를 연결하면서 한류 콘텐츠가 빠르게 확산되었습니다. "
    "짧은 시간에 소비될 수 있는 디지털 음원과 뮤직비디오는 소셜 네트워크 서비스(SNS)의 특성과 "
    "잘 맞았고, 그 덕분에 보는 음악을 지향하며 퍼포먼스를 강조한 케이팝이 아시아를 넘어 "
    "전 세계로 진출할 수 있었습니다. 그 대표적인 예가 2012년 가수 싸이의 <강남스타일>입니다. "
    "<강남스타일>은 미국을 비롯한 전 세계 음원 차트에서 상위권을 휩쓸었고 유튜브에서는 "
    "춤과 뮤직비디오를 패러디한 영상이 쏟아졌습니다. 전 세계인이 한국어로 노래를 부르게 된 것입니다. "
    "2013년에는 방탄소년단이 등장했습니다. 방탄소년단은 소셜 네트워크 서비스(SNS)를 통해 "
    "팬들과 실시간으로 소통하며 전 세계적인 팬덤을 만들어 냈습니다. "
    "방탄소년단은 힙합 음악에 전 세계 청춘들의 고민과 감정을 대변하는 메시지를 담았는데요, "
    "‘Love Yourself(너 자신을 있는 그대로 사랑하라)’는 메시지처럼 특정 언어와 지역에 국한되지 않고 "
    "진정한 예술을 추구하는 인간의 노력은 세계 무대에서 그 고유한 가치를 존중받아야 한다는 "
    "인식을 확산시켰습니다.",
}

QUESTIONS = {
 # 연어를 눈 밑에 올려놓는다는 문장이 신판에서 빠졌다. 신판이 '무엇보다도'로
 # 힘을 준 인과(피곤해서 생긴다)를 묻는 것으로 바꾼다.
 "LC-5-8-002": {"question": "‘다크 서클’은 대부분 피곤해서 생긴다고 했다."},
 "LC-8-3-001": {"question": "강의의 주제를 고르십시오.",
                "selection1": "케이팝이 세계로 퍼져 나간 과정",
                "selection2": "한국 아이돌 그룹의 활동 방식",
                "selection3": "소셜 네트워크 서비스의 발달 과정",
                "selection4": "일본 음악 시장의 변화",
                "answer_index": 0},
 # 오답은 모두 지문에서 근거를 대어 틀렸다고 말할 수 있어야 한다.
 #  1) 용어는 2001년경 일본에서 생겼다  3) 전 세계 청춘의 고민을 담았다
 #  4) 2000년대 '초'에 보아·동방신기가 진출했다
 "LC-8-3-002": {"question": "들은 내용과 같은 것을 고르십시오.",
                "selection1": "케이팝이라는 말은 <강남스타일>이 세계적으로 인기를 얻으면서 생겨났다.",
                "selection2": "케이팝이라는 말은 일본에서 제이팝에 대응하여 쓰이기 시작했다.",
                "selection3": "방탄소년단은 힙합 음악에 한국의 전통적인 가치를 담아 인기를 얻었다.",
                "selection4": "2000년대 후반에 한국 가수들이 처음으로 일본 시장에 진출했다.",
                "answer_index": 1},
}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/…]", "", t)


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    for iid, txt in LINES.items():
        b = int(iid.split("-")[1])
        gp = GlobalPdf(f"{APX}/(최종본)연세글로벌한국어_{b}급_부록-최종(26.8.10).pdf")
        blob = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))
        if sq(txt) not in blob:
            # 어디서 갈리는지 짚어 준다
            t = sq(txt); k = 0
            while k < len(t) and t[:k + 1] in blob:
                k += 1
            raise SystemExit(f"✗ {iid} 신판 부록과 불일치\n  일치까지: …{t[max(0,k-30):k]}\n  어긋난 뒤: {t[k:k+30]}")
        print(f"  ✓ {iid} 신판 부록과 일치 ({len(sq(txt))}자)")

    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    ws = wb["n3_listen_script_line"]
    c_id, c_tx = col(ws, "item_id"), col(ws, "text")
    c_rs, c_nt = col(ws, "review_status"), col(ws, "change_note")
    c_bk, c_sid, c_sp = col(ws, "book_id"), col(ws, "script_id"), col(ws, "speaker")
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, c_id).value
        if iid in LINES:
            ws.cell(r, c_tx).value = LINES[iid]
            ws.cell(r, c_rs).value = "fixed_v16"
            ws.cell(r, c_nt).value = "신판 부록 듣기 지문으로 교체(전면 개정)"
            log.append(dict(구분="듣기 대본 신판 반영", item_id=iid,
                            내용=LINES[iid][:70], 근거="신판 부록 듣기 지문"))
    print(f"\n대본 {len(LINES)}줄 교체")

    wq = wb["n3_listen_repeat"]
    q_id = col(wq, "item_id")
    for r in range(2, wq.max_row + 1):
        iid = wq.cell(r, q_id).value
        if iid not in QUESTIONS:
            continue
        for k, v in QUESTIONS[iid].items():
            wq.cell(r, col(wq, k)).value = v
        wq.cell(r, col(wq, "review_status")).value = "fixed_v16"
        wq.cell(r, col(wq, "change_note")).value = "대본 전면 개정에 맞춰 문항 재작성"
        log.append(dict(구분="듣기 문항 재작성", item_id=iid,
                        내용=str(QUESTIONS[iid].get("question", ""))[:70],
                        근거="교체한 대본과 맞춤"))
    print(f"문항 {len(QUESTIONS)}개 재작성")

    ws2 = wb["n3_listen_script"]
    s_b, s_id, s_at = col(ws2, "book_id"), col(ws2, "id"), col(ws2, "audio_text")
    live = {}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, c_rs).value) == "deleted" or not ws.cell(r, c_id).value:
            continue
        key = (int(ws.cell(r, c_bk).value), str(ws.cell(r, c_sid).value))
        if key not in TARGET:
            continue
        sp, tx = ws.cell(r, c_sp).value, ws.cell(r, c_tx).value
        live.setdefault(key, []).append(f"{sp}: {tx}" if sp else str(tx))
    k = 0
    for r in range(2, ws2.max_row + 1):
        key = (int(ws2.cell(r, s_b).value or 0), str(ws2.cell(r, s_id).value))
        if key in live:
            ws2.cell(r, s_at).value = "\n".join(live[key])
            ws2.cell(r, col(ws2, "review_status")).value = "fixed_v16"
            k += 1
    print(f"audio_text {k}개 대본 다시 생성")

    lg = wb.create_sheet("99_변경내역_v16")
    lg.append(["구분", "item_id", "내용", "근거"])
    for x in log:
        lg.append([x["구분"], x["item_id"], x["내용"], x["근거"]])
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거: 신판 부록 듣기 지문(5급 p131, 8급 p53)"])
    wb.save(DST)
    print(f"\n-> {DST}")


if __name__ == "__main__":
    main()
