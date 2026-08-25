#!/usr/bin/env python3
"""신판에서 실제로 고쳐진 듣기 대목을 짚는다.

앞의 40자 토막 방식은 오탐이 난다. 앞에서 한 글자만 어긋나도 뒤 토막이
전부 밀려 '80%가 바뀌었다'로 나온다(5급 13과는 '물어보아야→물어봐야'
한 곳 때문에 17%로 찍혔다).

difflib은 어긋난 뒤 다시 맞물리는 지점을 찾아 준다. 그래서 '몇 글자가
실제로 달라졌는가'를 잰다. 바뀐 대목만 뽑아 눈으로 볼 수 있게 낸다.

산출: verify/listen_diff.csv
"""
import os, re, csv, difflib, collections, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
APX = "/Users/soohyeon/Documents/2608-yonsei_renewal/book"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v15.xlsx"


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/…]", "", t)


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    lines = [r for r in sheet(wb, "n3_listen_script_line") if r.get("book_id")]
    by = collections.defaultdict(list)
    for r in lines:
        by[(int(r["book_id"]), str(r["script_id"]))].append(r)
    cand = {(int(r["book"]), r["script_id"])
            for r in csv.DictReader(open("verify/listen_revised.csv"))}

    out, cache = [], {}
    for (b, sid), rs in sorted(by.items()):
        if (b, sid) not in cand:
            continue
        if b not in cache:
            gp = GlobalPdf(f"{APX}/(최종본)연세글로벌한국어_{b}급_부록-최종(26.8.10).pdf")
            cache[b] = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))
        blob = cache[b]
        ex = "".join(sq(r.get("text")) for r in rs)
        # 대본이 신판 어디쯤에 있는지 — 가장 긴 앵커로 잡는다
        anchor = next((sq(r.get("text"))[:16] for r in rs
                       if sq(r.get("text"))[:16] in blob), None)
        if not anchor:
            out.append(dict(book=b, chapter=rs[0].get("chapter"), script_id=sid,
                            같음률="0%", 바뀐대목="대본 전체가 신판에 없음",
                            head=str(rs[0].get("text"))[:44]))
            continue
        i = blob.find(anchor)
        seg = blob[i:i + int(len(ex) * 1.4) + 60]
        sm = difflib.SequenceMatcher(None, ex, seg, autojunk=False)
        ratio = sm.ratio()
        diffs = []
        for tag, a1, a2, b1, b2 in sm.get_opcodes():
            # 쪽 하단 판권 줄과 다음 대본으로 넘어간 부분은 차이가 아니다
            if tag == "equal" or (a2 - a1) + (b2 - b1) < 2:
                continue
            o, n = ex[a1:a2][:26], seg[b1:b2][:26]
            if re.search(r"indb|오전|오후|듣기지문|연세글로벌", n):
                continue
            if not o and len(n) > 18:      # 다음 대본이 창에 딸려 들어온 것
                continue
            diffs.append(f"구판「{o}」→ 신판「{n}」")
        out.append(dict(book=b, chapter=rs[0].get("chapter"), script_id=sid,
                        같음률=f"{ratio:.0%}", 바뀐대목=" / ".join(diffs[:4]) or "차이 없음",
                        head=str(rs[0].get("text"))[:44]))

    with open(f"{HERE}/verify/listen_diff.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)
    for o in sorted(out, key=lambda x: int(x["같음률"].rstrip("%"))):
        print(f"  {o['book']}급 {o['chapter']:>2}과 s{o['script_id']:>4} 같음 {o['같음률']:>4s}  {o['head'][:36]}")
        print(f"       {o['바뀐대목'][:170]}")
    print(f"\n-> verify/listen_diff.csv")


if __name__ == "__main__":
    main()
