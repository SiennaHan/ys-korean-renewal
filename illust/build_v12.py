#!/usr/bin/env python3
"""폐기 어휘가 오답 보기로 남은 46문항을 정리해 v12를 만든다.

오답 보기는 그 과 어휘만 쓰는 게 아니라 급 전체에서 끌어온다. 그래서
8급 1과에서 폐기한 '남아선호'가 8급 3과 문항의 오답으로 들어가 있었다.
급 단위로 다시 세니 46문항(3급 1·7급 1·8급 44)이었다.

처리
  45문항 (meaning-to-word) — 폐기 오답을 같은 과의 살아 있는 어휘로 교체한다.
    문제에 정답의 뜻이 주어지는 유형이라, 대체 오답이 정답과 뜻이 겹치지만
    않으면 복수정답이 생기지 않는다. 그래서 후보를 고를 때 en 뜻이 정답과
    겹치는 것은 제외한다.
  1문항 (image-to-word, WQ-8-3-003 '카메오') — 문항 자체를 폐기한다.
    그림은 카페에서 잘 차려입은 사람이 지나가고 종업원이 놀라는 장면이다.
    라벨 없이 '카메오'를 떠올릴 수 없다. 그림만 보고 답을 고르는 유형에
    맞지 않는다(앞서 '선생님·나·가깝다' 등을 같은 이유로 뺀 것과 같은 기준).

산출: 글로벌_교재기반_콘텐츠_v12.xlsx
"""
import os, re, shutil, collections, unicodedata
import openpyxl

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v11.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v12.xlsx"
DROP = {"WQ-8-3-003": "그림만 보고 '카메오'를 고를 수 없음 — 유형 부적합"}
STOP = set("to be a an the of in on for with and or".split())


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    t = re.sub(r"\(.*?\)", "", t)
    return re.sub(r"[\s.,·\[\]/]", "", t).strip()


def en_tokens(s):
    return {w for w in re.findall(r"[a-z]+", str(s or "").lower())
            if w not in STOP and len(w) > 2}


def col(ws, n):
    for i, c in enumerate(ws[1], 1):
        if c.value == n:
            return i


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    ws = wb["n1_word_list"]
    Cw = {n: col(ws, n) for n in ("book_id", "chapter", "word", "en", "review_status")}
    alive, gone = collections.defaultdict(set), collections.defaultdict(set)
    live_ch = collections.defaultdict(list)   # (급,과) -> [(word, en)]
    for row in range(2, ws.max_row + 1):
        b = ws.cell(row, Cw["book_id"]).value
        w = ws.cell(row, Cw["word"]).value
        if not b or not w:
            continue
        b = int(b)
        if str(ws.cell(row, Cw["review_status"]).value) == "deleted":
            gone[b].add(sq(w))
        else:
            alive[b].add(sq(w))
            live_ch[(b, int(ws.cell(row, Cw["chapter"]).value))].append(
                (w, ws.cell(row, Cw["en"]).value))

    ws = wb["n1_word_quiz"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "chapter", "type",
                                 "answer_index", "selection1", "selection2",
                                 "selection3", "selection4", "meaning_en",
                                 "review_status", "change_note", "hold_reason")}
    log = []
    n_fix = n_drop = n_fail = 0
    used = collections.Counter()     # 대체어가 한쪽에 쏠리지 않게

    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        b = ws.cell(row, C["book_id"]).value
        if not iid or not b or str(ws.cell(row, C["review_status"]).value) == "deleted":
            continue
        b = int(b)
        ch = int(ws.cell(row, C["chapter"]).value)

        if iid in DROP:
            ws.cell(row, C["review_status"]).value = "deleted"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = DROP[iid]
            n_drop += 1
            log.append(["문항 폐기", iid, f"{b}급 {ch}과", "", DROP[iid]])
            continue

        sels = [ws.cell(row, C[f"selection{i}"]).value for i in range(1, 5)]
        bad = [i for i, s in enumerate(sels)
               if s and sq(s) in gone[b] and sq(s) not in alive[b]]
        if not bad:
            continue

        ai = int(ws.cell(row, C["answer_index"]).value)
        ans_en = en_tokens(ws.cell(row, C["meaning_en"]).value)
        taken = {sq(s) for s in sels if s}
        swaps = []
        for i in bad:
            cands = [(w, e) for w, e in live_ch[(b, ch)]
                     if sq(w) not in taken and not (en_tokens(e) & ans_en)]
            if not cands:
                n_fail += 1
                log.append(["교체 실패", iid, str(sels[i]), "", "같은 과에 쓸 후보 없음"])
                continue
            cands.sort(key=lambda t: used[sq(t[0])])       # 덜 쓰인 것부터
            new = cands[0][0]
            used[sq(new)] += 1
            taken.add(sq(new))
            ws.cell(row, C[f"selection{i + 1}"]).value = new
            swaps.append(f"{sels[i]}→{new}")
        if swaps:
            ws.cell(row, C["change_note"]).value = "폐기 어휘 오답 교체: " + ", ".join(swaps)
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["review_status"]).value = "auto_checked"
            n_fix += 1
            log.append(["오답 교체", iid, f"{b}급 {ch}과", ", ".join(swaps), ""])

    name = "99_변경내역_v12"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "위치/구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "폐기 어휘가 오답 보기로 남은 46문항", "", "",
               "오답은 급 전체에서 끌어오므로 과 단위 검사로는 39건을 놓쳤었음"])
    sh.append(["안전장치", "정답과 en 뜻이 겹치는 후보는 제외", "", "",
               "meaning-to-word는 문제에 정답 뜻이 주어져 복수정답 위험이 여기서 생김"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((14, 16, 22, 46, 46), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"오답 교체 {n_fix}문항 / 문항 폐기 {n_drop} / 교체 실패 {n_fail}")
    print(f"대체어 사용 분포: {dict(used.most_common(8))}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
