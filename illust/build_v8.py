#!/usr/bin/env python3
"""보류 6턴을 확정해 v8을 만든다.

지면을 직접 읽어 갈랐다.

  부분 수정 4턴 — 한국어를 신판 문장으로 바꾸고, 바뀐 만큼만 번역을 손봤다.
    RP-5-6-004   10년쯤 됐어요 → 10년이 넘었어요
    RP-7-5-006   술이나 한 잔합시다 → 식사나 같이 할까요   (음주 표현 순화)
    RP-7-10-009  '얼짱/몸짱이라던데.' 삭제
    RP-7-12-004  그 사건에 대해 재조사를 하라고 → 그 사건을 재조사하라고

  변경 없음 2턴 — RP-5-7-008 / RP-5-7-010.
    앞서 '신판 과제1에 보기 대화가 없다'고 보류했는데, 페이지 앞부분(안내문)만
    보고 판단한 것이었다. p84 뒤쪽에 보기 대화가 그대로 있고 화자도 v7과 같다
    (제임스/슈테판). 자동 탐지의 거짓 양성이었다. 보류를 푼다.

산출: 글로벌_교재기반_콘텐츠_v8.xlsx
"""
import os, re, shutil, unicodedata
import openpyxl
from global_text import GlobalPdf

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v7.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v8.xlsx"

EDIT = {
    "RP-5-6-004": dict(
        ko=("만들어진 지 10년쯤 됐어요.", "만들어진 지 10년이 넘었어요."),
        en=("It has been about 10 years since it was formed.",
            "It has been more than 10 years since it was formed."),
        jp=("作られてから10年くらいになりました。", "作られてから10年以上になりました。"),
        cn=("制定10年左右了，", "开通十多年了，"),
        vi=("Lập ra được khoảng 10 năm rồi ạ.", "Lập ra được hơn 10 năm rồi ạ."),
        note="신판 문장 교체 (10년쯤 됐어요 → 10년이 넘었어요)"),
    "RP-7-5-006": dict(
        ko=("언제 시간 되면 술이나 한 잔합시다.", "언제 시간 되면 식사나 같이 할까요?"),
        en=("Let's go for a drink sometime when we have time.",
            "Shall we have a meal together sometime when we have time?"),
        jp=("いつか時間があったらお酒でも一杯やりましょう。",
            "いつか時間があったら食事でも一緒にしませんか。"),
        cn=("哪天有空咱们一起喝一杯吧。", "哪天有空咱们一起吃顿饭怎么样？"),
        vi=("Khi nào rảnh thì mình đi làm vài ly nhé.",
            "Khi nào rảnh mình cùng đi ăn một bữa nhé?"),
        note="신판 문장 교체 — 음주 표현이 식사로 순화됨"),
    "RP-7-10-009": dict(
        ko=(" 얼짱/몸짱이라던데.", ""),
        en=(" They say she/he is a real looker (pretty face / great body).", ""),
        jp=(" オルチャン/モムチャン(美顔/美ボディ)らしいね。", ""),
        cn=("据说颜值/身材超棒呢。", ""),
        vi=(" Nghe bảo đẹp mặt/đẹp dáng lắm.", ""),
        note="신판에서 '얼짱/몸짱이라던데.' 삭제"),
    "RP-7-12-004": dict(
        ko=("그 사건에 대해 재조사를 하라고 했다는데요.", "그 사건을 재조사하라고 했다는데요."),
        en=("ordered a reinvestigation of the case", "ordered the case to be reinvestigated"),
        jp=("その事件について再調査をするよう指示した", "その事件を再調査するよう指示した"),
        cn=("下令对这起案件重新调查了", "下令重新调查这起案件了"),
        vi=("yêu cầu điều tra lại vụ việc đó", "yêu cầu điều tra lại vụ việc này"),
        note="신판 문장 교체 (재조사를 하라고 → 재조사하라고)"),
}
UNHOLD = {
    "RP-5-7-008": "신판 p84에 보기 대화가 그대로 있고 화자도 동일 — 자동 탐지 거짓 양성",
    "RP-5-7-010": "신판 p84에 보기 대화가 그대로 있고 화자도 동일 — 자동 탐지 거짓 양성",
}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i


def main():
    blobs = {}
    for b in (5, 7):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        blobs[b] = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))

    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n2_ai_role_play"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "ko", "en", "jp", "cn", "vi",
                                 "review_status", "change_note", "hold_reason")}
    log, ok, fail = [], 0, 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        if iid in EDIT:
            e = EDIT[iid]
            b = int(ws.cell(row, C["book_id"]).value)
            before = str(ws.cell(row, C["ko"]).value or "")
            newko = before.replace(*e["ko"]).strip()
            if sq(newko)[:26] not in blobs[b]:
                ws.cell(row, C["hold_reason"]).value = "교정안이 신판 본문과 불일치 — 사람 확인 필요"
                fail += 1
                log.append(["보류 유지", iid, before[:56], newko[:56], "본문 대조 실패"])
                continue
            ws.cell(row, C["ko"]).value = newko
            for k in ("en", "jp", "cn", "vi"):
                v = str(ws.cell(row, C[k]).value or "")
                ws.cell(row, C[k]).value = v.replace(*e[k]).strip()
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = e["note"]
            ok += 1
            log.append(["부분 수정 확정", iid, before[:56], newko[:56], "본문 대조 통과"])
        elif iid in UNHOLD:
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["change_note"]).value = "변경 없음 확인 — 보류 해제"
            log.append(["보류 해제", iid, "", "", UNHOLD[iid]])

    name = "99_변경내역_v8"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "보류 6턴 확정", "", "", "부분 수정 4 / 변경 없음 2"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((16, 16, 46, 46, 44), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"부분 수정 확정 {ok}건 / 보류 해제 {len(UNHOLD)}건 / 여전히 보류 {fail}건")
    for r in log:
        print(f"  [{r[0]}] {r[1]}")
        if r[2]:
            print(f"      구: {r[2]}")
            print(f"      신: {r[3]}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
