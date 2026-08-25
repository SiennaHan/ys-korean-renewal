#!/usr/bin/env python3
"""미션대화가 신판에서 바뀐 소재를 전제로 하는지 찾는다.

문법은 신구판이 같아서 미션대화가 잘 견딘다(앞서 117개 전건 확인). 그러나
**상황**은 다르다. 시나리오가 구판 본문의 소재를 전제로 하면 신판과 어긋난다.

그래서 이번까지 대조에서 '바뀐 것'으로 확정된 낱말·소재를 목록으로 두고
시나리오·미션 지시·첫 발화·프롬프트 네 곳을 훑는다.

산출: verify/chat_stale.csv
"""
import os, re, csv, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v27.xlsx"

# (낡은 말, 신판이 쓰는 말, 근거)
STALE = [
    ("전자사전", "태블릿 피시", "6급 4과 듣기 대본 개정"),
    ("드라마 DVD", "한국 과자", "3급 13과 듣기 대본 개정"),
    ("DVD", "—", "매체가 낡음"),
    ("심리 추리 영화", "스릴러 영화", "6급 6과 어휘 교체"),
    ("체력단련실", "헬스장", "5급 3과 어휘 교체"),
    ("귀고리", "귀걸이", "3급 14과 어휘 교체"),
    ("구세대", "기성세대", "7급 11과 어휘 교체"),
    ("입체 프린터", "—", "8급 11과에서 3D 프린터 내용 자체가 빠짐"),
    ("삼차원 프린터", "—", "위와 같음"),
    ("잔돈", "신용 카드", "6급 5과 본문 재저작"),
    ("장 보다", "인공 지능", "6급 13과 본문 재저작"),
    ("바코드", "인공 지능", "6급 13과 본문 재저작"),
    ("가상 마트", "인공 지능", "6급 13과 본문 재저작"),
    ("터치스크린", "스스로 파악하는", "6급 13과 냉장고 설명 개정"),
    ("정사각형 코드", "전기차", "8급 11과 어휘 교체"),
    ("의료용 로봇", "인공지능", "8급 11과 어휘 교체"),
    ("자동차 원격 조정", "항공택시", "8급 11과 어휘 교체"),
    ("휘는 디스플레이", "증강현실", "8급 11과 어휘 교체"),
    ("전자책", "가상공간", "6급 13과 어휘 교체"),
    ("남성적이다", "남자다움", "8급 1과 재저작으로 폐기"),
    ("여성적이다", "여자다움", "8급 1과 재저작으로 폐기"),
    ("다정다감", "—", "8급 1과 본문에서 사라짐"),
    ("무절제", "—", "8급 6과 본문에서 사라짐"),
    ("서럽", "—", "8급 6과 본문에서 사라짐"),
    ("유키", "유카", "7급에서 인물명 통일"),
]
FIELDS = ["situation_ko", "situation_en", "mission_detail", "ai_first_line",
          "ai_persona_prompt", "scenario_title", "target_grammar"]


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n7_mission_chat") if r.get("item_id")]
    out = []
    for r in rows:
        for old, new, why in STALE:
            for f in FIELDS:
                v = unicodedata.normalize("NFC", str(r.get(f) or ""))
                if old in v:
                    i = v.index(old)
                    out.append(dict(item_id=r["item_id"], book=r["book_id"],
                                    chapter=r["chapter"], 낡은말=old, 신판=new,
                                    열=f, 근거=why,
                                    맥락=re.sub(r"\s+", " ", v[max(0, i-40):i+40])))
    with open(f"{HERE}/verify/chat_stale.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["item_id", "book", "chapter", "낡은말",
                                           "신판", "열", "근거", "맥락"])
        wr.writeheader()
        wr.writerows(out)
    print(f"미션대화 {len(rows)}개 · 낡은 소재 {len(out)}곳 "
          f"({len({o['item_id'] for o in out})}개 대화)")
    for o in out:
        print(f"  {o['book']}권 {o['chapter']:>2}과 [{o['낡은말']}→{o['신판']}] {o['열']}")
        print(f"       …{o['맥락']}…")
        print(f"       근거: {o['근거']}")
    print(f"\n-> verify/chat_stale.csv")


if __name__ == "__main__":
    main()
