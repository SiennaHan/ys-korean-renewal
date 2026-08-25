#!/usr/bin/env python3
"""v18 — 색인 대조에서 딸려 나온 원장 손질.

A. 낱말을 바꾸고 번역을 안 고친 것 7건
   v5에서 구판 잔재 어휘 8건을 신판 낱말로 갈아 끼웠는데, 뜻풀이가 옛 낱말
   그대로 남은 자리가 있었다. '스릴러 영화'가 네 언어 모두 '심리 추리 영화'의
   번역을 달고 있었고, '현저히'는 부사가 됐는데 뜻풀이는 형용사 서술형이었다.

B. 같은 과에 같은 낱말이 두 줄인 것 2건
   ('부르다(노래)/부르다(배)', '낫다(회복)/낫다(우월)'는 동형이의어라 정상이다.)

C. 신판 본교재·부록 어디에도 없는 어휘 5건 폐기
   전부 8급이다. '삼차원 프린터'는 짝인 '입체 프린터'만 폐기되고 이쪽은 남아
   있었다 — 신판 11과에서 3D 프린터 내용 자체가 빠졌다.

산출: 글로벌_교재기반_콘텐츠_v18.xlsx
"""
import shutil, datetime
import openpyxl

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v17.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v18.xlsx"

# A. 번역 정정 — item_id: (en, jp, cn, vi, 사유)
RETRANS = {
 "VL-5-3-017": ("gym", "ジム", "健身房", "phòng tập gym",
                "체력단련실→헬스장 교체 시 뜻풀이가 옛 낱말 그대로였음"),
 "VL-6-6-015": ("thriller (film)", "スリラー映画", "惊悚片", "phim giật gân",
                "심리 추리 영화→스릴러 영화 교체 시 네 언어 모두 옛 낱말 번역이었음"),
 "VL-7-11-003": ("older generation; the established generation", "既成世代",
                 "老一辈；上一代", "thế hệ đi trước",
                 "구세대→기성세대 교체 시 뜻풀이가 옛 낱말 그대로였음"),
 "VL-8-15-027": ("remarkably; conspicuously", "著しく", "显著地；明显地",
                 "một cách rõ rệt", "현저하다→현저히로 품사가 부사로 바뀜"),
 "VL-7-7-003": ("extreme", "極端な", "极端的", "cực đoan", "극단적이다→극단적 품사 표기 맞춤"),
 "VL-7-7-017": ("two-faced; double-sided", "二重的な", "双重的；两面的",
                "hai mặt, mâu thuẫn", "이중적이다→이중적 품사 표기 맞춤"),
 "VL-8-1-027": ("androgynous; gender-neutral", "中性的", "中性的",
                "mang tính trung tính", "중성적이다→중성적 품사 표기 맞춤"),
}

# B. 같은 과 중복 — 폐기할 쪽
DUP = {
 "VL-5-3-023": "VL-5-3-017과 같은 낱말·같은 과 중복 — 뜻풀이가 있는 쪽을 남김",
 "VL-6-6-026": "VL-6-6-015와 같은 낱말·같은 과 중복 — 뜻풀이가 있는 쪽을 남김",
}

# C. 신판에 없어 폐기
STALE = {
 "VL-8-1-006": "남성적이다 — 신판 1과는 '남자다움'으로 다시 쓰임",
 "VL-8-1-020": "여성적이다 — 신판 1과는 '여자다움'으로 다시 쓰임",
 "VL-8-2-004": "귀가하다 — 신판 본교재·부록 어디에도 없음",
 "VL-8-3-026": "치유하다 — 신판 3과는 영화 <기생충> 감상문으로 다시 쓰임",
 "VL-8-11-011": "삼차원 프린터 — 신판 11과에서 3D 프린터 내용이 빠짐(짝인 '입체 프린터'는 이미 폐기)",
}


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n1_word_list"]
    C = {k: col(ws, k) for k in ["word", "en", "jp", "cn", "vi", "item_id",
                                 "review_status", "change_note"]}
    log, n = [], [0, 0, 0]
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        w = ws.cell(r, C["word"]).value
        if iid in RETRANS:
            en, jp, cn, vi, why = RETRANS[iid]
            for k, v in zip(["en", "jp", "cn", "vi"], [en, jp, cn, vi]):
                ws.cell(r, C[k]).value = v
            ws.cell(r, C["review_status"]).value = "fixed_v18"
            ws.cell(r, C["change_note"]).value = why
            n[0] += 1
            log.append(("번역 정정", iid, w, why))
        elif iid in DUP:
            ws.cell(r, C["review_status"]).value = "deleted"
            ws.cell(r, C["change_note"]).value = DUP[iid]
            n[1] += 1
            log.append(("중복 폐기", iid, w, DUP[iid]))
        elif iid in STALE:
            ws.cell(r, C["review_status"]).value = "deleted"
            ws.cell(r, C["change_note"]).value = STALE[iid]
            n[2] += 1
            log.append(("구판 잔재 폐기", iid, w, STALE[iid]))
    print(f"A. 번역 정정 {n[0]}건 / B. 중복 폐기 {n[1]}건 / C. 구판 잔재 폐기 {n[2]}건")

    lg = wb.create_sheet("99_변경내역_v18")
    lg.append(["구분", "item_id", "낱말", "사유"])
    for x in log:
        lg.append(list(x))
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거 CSV: verify/index_check.csv, "
               f"verify/ledger_sweep.csv, verify/index_stale.csv"])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
