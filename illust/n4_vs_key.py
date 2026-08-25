#!/usr/bin/env python3
"""n4 완성문을 부록 '모범 답안'과 맞대 본다.

모범 답안은 교재가 인정하는 문장이다. 우리 완성문이 거기 있으면 확인된 것이고,
**거의 같은데 조금 다르면** 그 차이가 결함일 수 있다(활용 오류·조사 차이).
전혀 다르면 교재에 없는 창작 문항이라 이 검사로는 판정할 수 없다.

섹션(어휘/문법1/과제2) 귀속은 쓰지 않는다. 줄이 붙어 나와 절 구분이 자주
깨지는데, 대조에는 '그 과의 답 문장 모음'만 있으면 충분하다.

산출: verify/n4_vs_key.csv
"""
import os, re, csv, collections, difflib, unicodedata
import openpyxl
from answer_key import parse

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v24.xlsx"


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/…➊➋➌➍➎○]", "", t)


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    key = collections.defaultdict(list)
    for b in range(1, 9):
        for x in parse(b):
            if x["chapter"]:
                # 한 줄에 여러 답이 붙는 경우가 있어 마디로 쪼갠다
                for piece in re.split(r"[.。]\s*", x["answer"]):
                    p = sq(piece)
                    if len(p) >= 6:
                        key[(b, int(x["chapter"]))].append((p, piece.strip()))
    print("과별 모범답안 문장 수집:",
          f"{sum(len(v) for v in key.values())}개 / {len(key)}개 과")

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n4_blank_question") if r.get("book_id")]
    out, stat = [], collections.Counter()
    for r in rows:
        b, ch = int(r["book_id"]), int(r["chapter"])
        comp = sq(str(r["completion"]).replace("<b>", "").replace("</b>", ""))
        if len(comp) < 6:
            stat["짧음(제외)"] += 1
            continue
        pool = key.get((b, ch), [])
        if not pool:
            stat["그 과 모범답안 없음"] += 1
            continue
        if any(comp in p or p in comp for p, _ in pool):
            stat["교재 답과 일치"] += 1
            continue
        best, bs = None, 0.0
        for p, raw in pool:
            s = difflib.SequenceMatcher(None, comp, p).ratio()
            if s > bs:
                best, bs = raw, s
        if bs >= 0.75:
            stat["거의 같음 — 확인 필요"] += 1
            out.append(dict(item_id=r["item_id"], book=b, chapter=ch,
                            유사도=round(bs, 2),
                            우리=str(r["completion"]).replace("<b>", "").replace("</b>", ""),
                            교재=best))
        else:
            stat["교재 답에 없음(창작)"] += 1

    with open(f"{HERE}/verify/n4_vs_key.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["item_id", "book", "chapter", "유사도", "우리", "교재"])
        wr.writeheader()
        wr.writerows(sorted(out, key=lambda x: -x["유사도"]))
    print("\n", dict(stat))
    print(f"\n■ 거의 같은데 다른 것 {len(out)}건")
    for o in sorted(out, key=lambda x: -x["유사도"]):
        print(f"  {o['book']}급 {o['chapter']:>2}과 {o['item_id']} ({o['유사도']})")
        print(f"      우리: {o['우리']}")
        print(f"      교재: {o['교재']}")
    print(f"\n-> verify/n4_vs_key.csv")


if __name__ == "__main__":
    main()
