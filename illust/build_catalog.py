#!/usr/bin/env python3
"""추출한 글로벌 삽화를 엑셀의 단어·문항(item_id)과 연결한 카탈로그를 만든다.

연결 경로 두 가지:
 1) legacy_bbox — 엑셀에 이미 들어 있는 3주완성 파일명(b1_ch4_p41_1 등)을
    구 매니페스트에서 좌표로 되짚고, 같은 좌표의 글로벌 삽화를 찾는다.
    글로벌은 삽화만 새로 그렸을 뿐 판형·좌표가 같아서 기존 460건 연결이 그대로 산다.
 2) label — 어휘 박스 그림은 바로 아래에 단어가 찍혀 있다.
    (급, 과, 단어) 키로 새로 붙인다. 1)이 없는 행을 메운다.

두 경로가 모두 나오면 서로 대조해 agree/conflict를 표시한다.

산출: catalog.csv  (이미지 1장 = 1행, 붙은 단어·item_id 포함)
      needs.csv    (그림이 필요한데 아직 못 붙인 엑셀 행)
"""
import os, csv, re, collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
LEGACY = "/Users/soohyeon/Documents/2606-yonsei3week_parse/work"

def _newest_ledger():
    """가장 높은 v 번호 원장. v3 시절에 고정해 뒀던 걸 v30(2026-08-24)에 풀었다 —
    그 사이 어휘가 2827→2890행으로 늘었고(신규 63건), image 열도 계속 정정됐다."""
    import glob, re
    root = "/Users/soohyeon/Documents/2608-yonsei_renewal"
    found = [(int(m.group(1)), p) for p in glob.glob(f"{root}/글로벌_교재기반_콘텐츠_v*.xlsx")
             if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(found)[1]

XLSX = _newest_ledger()

POS_TOL = 8.0     # 구/신 좌표 허용 오차(pt)
SIZE_TOL = 0.25   # 크기 상대 오차


LEGACY_REF = re.compile(r"^b[1-8]_ch\d+_p\d+_\d+$")


def norm(s):
    """활용형 괄호·공백 제거한 매칭용 표제어.

    라벨은 그림 아래 26pt 띠를 잘라 읽어서 '양력설(' 처럼 여는 괄호에서
    끊기는 일이 잦다. 짝 없는 괄호/대괄호는 그 앞까지만 쓴다.
    """
    s = str(s or "")
    s = re.sub(r"\(.*?\)", "", s)      # 닫힌 괄호는 통째로 제거
    s = re.split(r"[(\[]", s)[0]       # 짝 없는 괄호는 그 앞까지
    return re.sub(r"[\s.,]+", "", s).strip()


def legacy_key(v):
    """엑셀 image 값이 실제 구 파일명일 때만 키로 쓴다.

    '닭갈비[닥깔비]' 처럼 발음 정보가 잘못 들어간 칸이 있다.

    ⚠️ 이 스크립트는 "구 3주완성판 파일명 → 새 글로벌판 파일명" 최초 이관
    한 번만 쓰는 도구다. apply_images.py 로 원장의 image 칸을 이미 새
    파일명으로 바꾼 뒤에 다시 돌리면, 이 함수가 새 파일명을 구 파일명
    패턴에 못 맞춰 좌표 매칭이 전멸하고 라벨 매칭만 남는다(2026-09-05,
    600dpi 재추출 때 실제로 이렇게 792건 중 좌표 매칭이 0으로 빠졌다 —
    미해결이 8→36건으로 튀는 걸 보고서야 알았다). 단순 재추출(같은 파일명,
    다른 DPI)은 이 스크립트를 다시 돌릴 필요가 없다 — 원장 image 값을
    그대로 두고 새로 렌더링한 같은 이름의 파일로 덮어쓰면 된다.
    """
    k = str(v or "").strip().replace(".png", "")
    return k if LEGACY_REF.match(k) else None


def read_sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def load_new():
    """글로벌 삽화 매니페스트 (급 -> [row])."""
    out = collections.defaultdict(list)
    for b in range(1, 9):
        for r in csv.DictReader(open(f"{HERE}/manifest_b{b}.csv")):
            for k in ("chapter", "pdf_page", "x0", "y0", "w", "h"):
                r[k] = int(r[k])
            r["text_cover"] = float(r["text_cover"])
            out[b].append(r)
    return out


def load_legacy():
    """구 매니페스트 파일명 -> 좌표. 엑셀 image 값이 이걸 가리킨다."""
    out = {}
    for b in range(1, 9):
        for r in csv.DictReader(open(f"{LEGACY}/book{b}/images_manifest.csv")):
            out[r["filename"].replace(".png", "")] = dict(
                book=b, chapter=int(r["chapter"]), page=int(r["pdf_page"]),
                x0=float(r["x0"]), y0=float(r["y0"]),
                w=float(r["w"]), h=float(r["h"]))
    return out


def book_offset(new_rows, legacy, book):
    """구 PDF와 글로벌 PDF의 좌표 원점 차이(dx, dy)를 추정한다.

    8급은 글로벌 판이 (+29,+29) 밀려 있다. 같은 쪽·같은 크기의 그림 쌍에서
    나온 (dx,dy)의 최빈값을 쓴다 — 판형이 같으니 권 전체가 같은 값이다.
    """
    by_page = collections.defaultdict(list)
    for r in new_rows:
        by_page[r["pdf_page"]].append(r)
    votes = collections.Counter()
    for L in legacy.values():
        if L["book"] != book:
            continue
        same = [r for r in by_page.get(L["page"], [])
                if abs(r["w"] - L["w"]) <= 2 and abs(r["h"] - L["h"]) <= 2]
        if len(same) == 1:
            votes[(round(same[0]["x0"] - L["x0"]), round(same[0]["y0"] - L["y0"]))] += 1
    if not votes:
        return 0, 0
    (dx, dy), n = votes.most_common(1)[0]
    return (dx, dy) if n >= 5 else (0, 0)


def find_by_bbox(new_rows, page, x0, y0, w, h, off=(0, 0)):
    """같은 쪽에서 좌표가 겹치는 글로벌 삽화 찾기."""
    x0, y0 = x0 + off[0], y0 + off[1]
    best, best_d = None, 1e9
    for r in new_rows:
        if r["pdf_page"] != page:
            continue
        d = abs(r["x0"] - x0) + abs(r["y0"] - y0)
        if d > POS_TOL * 2:
            continue
        if w and abs(r["w"] - w) / w > SIZE_TOL:
            continue
        if h and abs(r["h"] - h) / h > SIZE_TOL:
            continue
        if d < best_d:
            best, best_d = r, d
    return best


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    wl = read_sheet(wb, "n1_word_list")
    wq = read_sheet(wb, "n1_word_quiz")
    lr = read_sheet(wb, "n3_listen_repeat")

    new = load_new()
    legacy = load_legacy()

    # (급, 과, 정규화단어) -> 그 자리의 글로벌 삽화들
    by_label = collections.defaultdict(list)
    for b, rows in new.items():
        for r in rows:
            if r["vocab_label"]:
                by_label[(b, r["chapter"], norm(r["vocab_label"]))].append(r)

    # 이미지가 어떤 item_id에 쓰이는지 (어휘원장 → 그림단어퀴즈 순으로 모은다)
    uses = collections.defaultdict(list)   # new filename -> [(item_id, sheet, word)]
    links, needs = [], []

    offsets = {b: book_offset(new[b], legacy, b) for b in range(1, 9)}
    print("권별 좌표 오프셋:", {b: o for b, o in offsets.items() if o != (0, 0)} or "없음")

    def resolve(book, chapter, word, legacy_ref, prefer_label):
        """(글로벌 파일명, 근거)

        prefer_label: 어휘 학습용(n1_*)은 '그림 아래 그 단어가 찍힌 것'이 정본이다.
        구 좌표는 신판에서 페이지가 재편집되면 엉뚱한 데를 가리킨다
        (5급 12과 '둥글다'가 그림자 레이어에 걸린 사례).
        반대로 듣기 선택지(n3)는 그 문항 자리의 그림이 정본이라 좌표를 따른다.
        """
        rows = new.get(book, [])
        hit_bbox = None
        key = legacy_key(legacy_ref)
        if key:
            L = legacy.get(key)
            if L:
                hit_bbox = find_by_bbox(rows, L["page"], L["x0"], L["y0"],
                                        L["w"], L["h"], offsets[book])
        cands = by_label.get((book, chapter, norm(word)), [])
        hit_label = cands[0] if cands else None
        many = "(후보多)" if len(cands) > 1 else ""

        if hit_bbox and hit_label:
            if hit_bbox["filename"] == hit_label["filename"]:
                return hit_bbox["filename"], "좌표+라벨 일치" + many
            if prefer_label:
                return hit_label["filename"], "라벨 우선(좌표와 불일치)" + many
            return hit_bbox["filename"], "좌표 우선(라벨과 불일치)" + many
        if hit_label:
            return hit_label["filename"], "라벨만" + many
        if hit_bbox:
            return hit_bbox["filename"], "좌표만"
        return "", "미해결"

    for r in wl:
        b, ch, word = r.get("book_id"), r.get("chapter"), r.get("word")
        if not b or not word:
            continue
        if not r.get("image"):
            continue
        fn, how = resolve(int(b), int(ch), word, r.get("image"), True)
        links.append(dict(item_id=r.get("item_id"), sheet="n1_word_list", book=b, chapter=ch,
                          word=word, legacy=r.get("image"), filename=fn, how=how))
        if fn:
            uses[fn].append((r.get("item_id"), "n1_word_list", word))
        else:
            needs.append(dict(item_id=r.get("item_id"), sheet="n1_word_list", book=b,
                              chapter=ch, word=word, legacy=r.get("image"),
                              why="구 파일명 좌표로도, 아래라벨로도 못 찾음"))

    for r in wq:
        if not r.get("image"):
            continue
        b, ch = int(r["book_id"]), int(r["chapter"])
        ai = r.get("answer_index")
        word = r.get(f"selection{(ai or 0) + 1}")
        fn, how = resolve(b, ch, word, r.get("image"), True)
        links.append(dict(item_id=r.get("item_id"), sheet="n1_word_quiz", book=b, chapter=ch,
                          word=word, legacy=r.get("image"), filename=fn, how=how))
        if fn:
            uses[fn].append((r.get("item_id"), "n1_word_quiz", word))
        else:
            needs.append(dict(item_id=r.get("item_id"), sheet="n1_word_quiz", book=b,
                              chapter=ch, word=word, legacy=r.get("image"),
                              why="그림보고 단어고르기인데 그림을 못 찾음"))

    for r in lr:
        b, ch = int(r["book_id"]), int(r["chapter"])
        for i in range(1, 5):
            ref = r.get(f"selection{i}_image")
            if not ref:
                continue
            word = r.get(f"selection{i}") or ""
            fn, how = resolve(b, ch, word, ref, False)
            links.append(dict(item_id=f"{r.get('item_id')}#s{i}", sheet="n3_listen_repeat",
                              book=b, chapter=ch, word=word, legacy=ref,
                              filename=fn, how=how))
            if fn:
                uses[fn].append((f"{r.get('item_id')}#s{i}", "n3_listen_repeat", word))
            else:
                needs.append(dict(item_id=f"{r.get('item_id')}#s{i}", sheet="n3_listen_repeat",
                                  book=b, chapter=ch, word=word, legacy=ref,
                                  why="듣기 선택지 그림을 못 찾음"))

    # ── 카탈로그: 삽화 1장 = 1행
    with open(f"{HERE}/catalog.csv", "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow(["filename", "book", "chapter", "ch_kind", "ch_title", "pdf_page",
                     "w", "h", "vocab_label", "band_text", "text_cover",
                     "linked_word", "item_ids", "use_count", "kind"])
        for b in range(1, 9):
            for r in new[b]:
                u = uses.get(r["filename"], [])
                words = sorted({w for _, _, w in u if w})
                kind = ("어휘그림" if r["vocab_label"] else
                        "텍스트패널의심" if r["text_cover"] > 0.25 else
                        "장면/기타")
                wr.writerow([r["filename"], b, r["chapter"], r["ch_kind"], r["ch_title"],
                             r["pdf_page"], r["w"], r["h"], r["vocab_label"],
                             r["band_text"], r["text_cover"],
                             " / ".join(words), " ".join(i for i, _, _ in u), len(u), kind])

    with open(f"{HERE}/links.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["item_id", "sheet", "book", "chapter",
                                           "word", "legacy", "filename", "how"])
        wr.writeheader()
        wr.writerows(links)

    with open(f"{HERE}/needs.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["item_id", "sheet", "book", "chapter",
                                           "word", "legacy", "why"])
        wr.writeheader()
        wr.writerows(needs)

    how = collections.Counter(l["how"] for l in links)
    print(f"엑셀에서 그림 쓰는 행 {len(links)}건")
    for k, v in how.most_common():
        print(f"   {k:28s} {v}")
    print(f"미해결 {len(needs)}건 -> needs.csv")
    linked_imgs = sum(1 for v in uses.values() if v)
    print(f"삽화 {sum(len(v) for v in new.values())}장 중 문항에 연결된 것 {linked_imgs}장")


if __name__ == "__main__":
    main()
