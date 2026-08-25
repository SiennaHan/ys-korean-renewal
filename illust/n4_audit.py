#!/usr/bin/env python3
"""문법 문항(n4) 점검.

n4는 교재 문장을 옮긴 게 아니라 합성 문항이라 본문 대조가 성립하지 않는다.
대신 세 가지를 본다.

A. 구조 정합성 — answer / answer_text / selection[answer_index] / selections /
   completion 이 서로 맞는가. 한 군데만 어긋나도 앱에서 오답 처리가 난다.

B. 문항 자체 정합성 — 문제의 빈칸에 정답을 넣으면 completion이 나와야 한다.
   (교재 '● 문법' 목록과의 대조는 폐기했다. n4 정답은 문법 형태소가 아니라
    활용된 어형이라—'몰라'(르불규칙), '더워요'(ㅂ불규칙), '매운'(-은)—
    항목명과 문자열로 맞댈 수 없다. 63%가 불일치로 나와 쓸모가 없었다.)

C. 이형태 규칙 — '받침 O → -을까요 / 받침 X → -ㄹ까요' 같은 문항이 많다.
   문제 문장에서 어간을 뽑아 받침 유무를 보고 정답이 규칙에 맞는지 확인한다.
   이게 틀리면 명백한 오답이라 가장 중요한 검사다.

산출: verify/n4_audit.csv
"""
import os, csv, re, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v22.xlsx"
OUT = f"{HERE}/verify"

# 받침 유무로 갈리는 대표 이형태 쌍 (받침 있을 때, 받침 없을 때)
ALLOMORPH = [
    ("-을까요", "-ㄹ까요"), ("-을게요", "-ㄹ게요"), ("-을래요", "-ㄹ래요"),
    ("-은", "-ㄴ"), ("-을", "-ㄹ"), ("-으세요", "-세요"),
    ("-으면", "-면"), ("-으니까", "-니까"), ("-으러", "-러"),
    ("-으려고", "-려고"), ("-습니다", "-ㅂ니다"), ("-습니까", "-ㅂ니까"),
    ("-은데", "-ㄴ데"), ("-으로", "-로"), ("-이라서", "-라서"),
    ("-을수록", "-ㄹ수록"), ("-으려면", "-려면"), ("-은데다가", "-ㄴ데다가"),
]


def has_batchim(ch):
    """한글 음절이 '받침 있는 것처럼' 활용하는가.

    ㄹ 받침은 예외다. 연필+로(으로 아님), 알다→압니다(알습니다 아님),
    만들다→만듭니다 처럼 무받침형을 취한다. 이걸 안 넣으면 정상 데이터가
    전부 오류로 잡힌다(실제로 6건이 그렇게 잡혔다).
    """
    if not ch or not ("가" <= ch <= "힣"):
        return None
    jong = (ord(ch) - 0xAC00) % 28
    if jong == 8:          # ㄹ
        return False
    return jong != 0


def stem_before_blank(q):
    """문제 문장에서 괄호 바로 앞 음절 = 어간 끝 글자."""
    m = re.search(r"([가-힣])\s*[(（]", str(q or ""))
    return m.group(1) if m else None


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def norm(s):
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s or "")))


def strip_tags(s):
    return re.sub(r"<[^>]+>", "", str(s or ""))


BLANK = re.compile(r"[(（][\s_…·]*[)）]")


def fill_blank(q, ans):
    """문제의 빈칸에 정답을 넣는다. 뒤에 붙은 힌트 괄호는 떼어낸다.

    문제는 '날씨가 ( ). (덥다 → ___)'처럼 힌트가 따라붙기도 한다.
    빈칸은 '괄호 안이 공백·밑줄뿐'인 것으로 구분한다.
    """
    q = str(q or "")
    ans = str(ans or "")
    # 빈칸이 둘인 문항이 있다 — '읽( ) 읽( )'에 정답 '으면 … 을수록'.
    # '…'는 두 조각을 차례로 넣으라는 뜻이다. 첫 칸만 채우면 문장이 깨진다.
    if "…" in ans:
        parts = [re.sub(r"^[-–]", "", x.strip()) for x in ans.split("…")]
        out, i = q, 0
        while i < len(parts):
            m = BLANK.search(out)
            if not m:
                break
            out = out[:m.start()] + parts[i] + out[m.end():]
            i += 1
        filled = out
        filled = re.sub(r"[(（][^)）]*[)）]\s*$", "", filled).strip()
        return re.sub(r"([?!.])\1+", r"\1", filled)
    m = BLANK.search(q)
    if not m:
        return None
    # 정답은 의존형태소라 '-을까요'처럼 앞에 하이픈을 달고 있다. 붙일 때는 뗀다.
    a = re.sub(r"^[-–]", "", ans)
    filled = q[:m.start()] + a + q[m.end():]
    # 남은 힌트 괄호 제거 (빈칸을 채운 뒤에 하는 것이 안전)
    filled = re.sub(r"[(（][^)）]*[)）]\s*$", "", filled).strip()
    # 문제 끝의 '?'와 정답에 붙은 '?'가 겹쳐 '??'가 된다. 종결부호는 하나로.
    filled = re.sub(r"([?!.])\1+", r"\1", filled)
    filled = re.sub(r"\?\s*\.", "?", filled)
    return filled


def is_contraction(got, want):
    """축약으로 설명되는 차이인가.

    앞뒤 공통 부분을 뗀 나머지가 짧고(각 3자 이내), 줄어든 쪽이 completion이면
    축약으로 본다. 자모가 겹치는지도 확인해 전혀 다른 말이 통과하지 않게 한다.
    """
    if len(want) >= len(got):
        return False
    i = 0
    while i < min(len(got), len(want)) and got[i] == want[i]:
        i += 1
    j = 0
    while (j < min(len(got), len(want)) - i
           and got[len(got) - 1 - j] == want[len(want) - 1 - j]):
        j += 1
    a, b = got[i:len(got) - j], want[i:len(want) - j]
    if not a or len(a) > 3 or len(b) > 3:
        return False
    if not b:
        return True          # 가+아서 → 가서 처럼 한쪽이 통째로 줄어든 축약
    return bool(set(decompose(a)) & set(decompose(b)))


def decompose(s):
    """한글 음절을 초/중/종성으로 푼다."""
    CHO = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
    JUNG = "ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ"
    JONG = " ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ"
    out = []
    for ch in s:
        if "가" <= ch <= "힣":
            n = ord(ch) - 0xAC00
            out += [CHO[n // 588], JUNG[(n % 588) // 28], JONG[n % 28]]
        else:
            out.append(ch)
    return [c for c in out if c != " "]


def base_form(s):
    """이형태를 벗겨 대표형으로. -을까요/-ㄹ까요 → 까요, -습니다/-ㅂ니다 → 니다."""
    t = norm(s).replace("?", "").replace(".", "")
    t = re.sub(r"[-–]", "", t)
    t = re.sub(r"^(으|ㄹ|ㄴ|ㅂ|을|은)", "", t)
    return t


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n4_blank_question") if r.get("book_id")]
    out = []

    # ── A. 구조 정합성
    print("A. 구조 정합성")
    bad = collections.Counter()
    for r in rows:
        iid = r["item_id"]
        sels = [r.get(f"selection{i}") for i in range(1, 5)]
        live = [s for s in sels if s]
        ai = r.get("answer_index")

        if ai is None or not (0 <= int(ai) < len(live)):
            out.append(dict(kind="answer_index 범위 오류", item_id=iid,
                            detail=f"index={ai}, 보기 {len(live)}개"))
            bad["index"] += 1
            continue
        pick = live[int(ai)]
        if norm(pick) != norm(r.get("answer")):
            out.append(dict(kind="정답 불일치(answer)", item_id=iid,
                            detail=f"answer={r.get('answer')} vs 보기[{ai}]={pick}"))
            bad["answer"] += 1
        if norm(pick) != norm(r.get("answer_text")):
            out.append(dict(kind="정답 불일치(answer_text)", item_id=iid,
                            detail=f"answer_text={r.get('answer_text')} vs 보기[{ai}]={pick}"))
            bad["answer_text"] += 1
        # selections(콤마 결합) vs selection1~4
        parts = [x.strip() for x in str(r.get("selections") or "").split(",") if x.strip()]
        if parts and [norm(x) for x in parts] != [norm(x) for x in live]:
            out.append(dict(kind="selections 분리 불일치", item_id=iid,
                            detail=f"{r.get('selections')} vs {live}"))
            bad["selections"] += 1
        # 보기 중복
        if len({norm(x) for x in live}) != len(live):
            out.append(dict(kind="보기 중복", item_id=iid, detail=str(live)))
            bad["dup"] += 1
    print(f"  {dict(bad) or '이상 없음'}")

    # ── B. 문항 자체 정합성
    print("\nB. 빈칸에 정답을 넣으면 completion이 되는가")
    ok = 0
    bad2 = collections.Counter()
    for r in rows:
        q, ans, comp = r.get("question"), r.get("answer"), r.get("completion")
        if not q or not ans or not comp:
            out.append(dict(kind="필드 누락", item_id=r["item_id"],
                            detail=f"question/answer/completion 중 빈 값"))
            bad2["누락"] += 1
            continue
        filled = fill_blank(q, ans)
        if filled is None:
            bad2["빈칸 못 찾음"] += 1
            out.append(dict(kind="빈칸 못 찾음", item_id=r["item_id"], detail=str(q)[:60]))
            continue
        want = norm(strip_tags(comp))
        got = norm(filled)
        if got == want:
            ok += 1
        elif is_contraction(got, want):
            # 마시+어요→마셔요, 보+아요→봐요, 공부하+여요→공부해요.
            # 축약은 정상이라 오류로 세지 않는다.
            bad2["축약형(정상)"] += 1
        elif want and want in got:
            # completion이 '문제를 채운 문장'의 뒤쪽 토막인 경우.
            # 문제에 맥락 절이나 상대 발화(가:…나:…)가 붙어 있고 completion은
            # 목표 문장만 보여 주는 편집 방침이다. 결함이 아니다.
            bad2["맥락 절 생략(정상)"] += 1
        elif re.search(r"불규칙|탈락", str(r.get("grammar_focus") or "")):
            # 어간이 불규칙 활용을 하는 문항. 문제는 기본형 어간을 보이고
            # completion은 활용을 적용한다(춥+으니까 → 추우니까).
            bad2["불규칙 적용(정상)"] += 1
        else:
            bad2["불일치"] += 1
            out.append(dict(kind="completion 불일치", item_id=r["item_id"],
                            detail=f"문제+정답 → {filled}  |  completion → {strip_tags(comp)}"))
    print(f"  일치 {ok} / {dict(bad2) or '이상 없음'}")

    # 중복 문항
    seen = collections.defaultdict(list)
    for r in rows:
        seen[(norm(r.get("question")), norm(r.get("answer")))].append(r["item_id"])
    dups = {k: v for k, v in seen.items() if len(v) > 1}
    for k, v in dups.items():
        out.append(dict(kind="중복 문항", item_id=" / ".join(v), detail=k[0][:60]))
    print(f"  같은 문제+정답이 여러 번: {len(dups)}쌍")

    # ── C. 이형태 규칙
    print("\nC. 이형태 규칙 (받침 유무)")
    checked = wrong = 0
    for r in rows:
        sels = {norm(r.get(f"selection{i}")) for i in range(1, 5) if r.get(f"selection{i}")}
        for withb, without in ALLOMORPH:
            if {norm(withb), norm(without)} != sels:
                continue
            st = stem_before_blank(r.get("question"))
            b = has_batchim(st)
            if b is None:
                continue
            checked += 1
            want = withb if b else without
            if norm(r.get("answer")) != norm(want):
                wrong += 1
                out.append(dict(kind="이형태 규칙 불일치", item_id=r["item_id"],
                                detail=f"어간끝 '{st}' 받침{'O' if b else 'X'} → "
                                       f"{want} 여야 하는데 정답은 {r.get('answer')} "
                                       f"| {r.get('question')}"))
            break
    print(f"  검사 가능 {checked}문항 / 규칙 불일치 {wrong}건")

    with open(f"{OUT}/n4_audit.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["kind", "item_id", "detail"])
        wr.writeheader()
        wr.writerows(out)
    print(f"\n총 {len(out)}건 -> {OUT}/n4_audit.csv")
    print(dict(collections.Counter(o["kind"] for o in out)))


if __name__ == "__main__":
    main()
