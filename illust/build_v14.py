#!/usr/bin/env python3
"""v14 — 듣기(n3) 반영.

1) 빠뜨린 말 17건 되살리기 (n3_listen_script_line + n3_listen_script.audio_text)
   구판 부록 원문에 있는데 엑셀로 옮기면서 떨어져 나간 말들이다.
   '다크 서클은 대부분 [피곤해서] 생기기 때문에'처럼 빠지면 문장이 깨진다.

2) 7급 '유키' → '유카' 25칸
   신판 7급 본교재에는 유키가 없고 유카로 통일됐다. 6급은 신판에도 유키가
   그대로 있으므로 건드리지 않는다.

산출: 글로벌_교재기반_콘텐츠_v14.xlsx (+ 99_변경내역_v14 시트)
"""
import json, re, shutil, datetime
import openpyxl

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v13.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v14.xlsx"
WIN = 22          # audio_text에서 갈아끼울 창의 앞뒤 길이


def col(ws, name):
    hdr = [c.value for c in ws[1]]
    return hdr.index(name) + 1


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    fixes = json.load(open("verify/listen_fixes.json"))
    log = []

    # ── 1) 빠뜨린 말 되살리기
    ws = wb["n3_listen_script_line"]
    c_id, c_text = col(ws, "item_id"), col(ws, "text")
    c_rs, c_note = col(ws, "review_status"), col(ws, "change_note")
    byid = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, c_id).value
        if v:
            byid.setdefault(v, r)

    windows = []          # audio_text에도 같은 자리를 갈아끼우기 위해
    # 한 줄에 고칠 자리가 여럿인 대본이 있다(6급 6과는 네 군데).
    # 각 수정문은 원문 기준으로 만들어졌으므로 그대로 덮어쓰면 두 번째부터
    # 원문이 안 맞는다. 창(window) 단위로 하나씩 갈아끼워 누적한다.
    by_item = {}
    for f in fixes:
        by_item.setdefault(f["item_id"], []).append(f)

    for iid, fs in by_item.items():
        r = byid.get(iid)
        if not r:
            print(f"  ! {iid} 행 없음")
            continue
        cur = ws.cell(r, c_text).value
        if cur != fs[0]["before"]:
            print(f"  ! {iid} 원문 불일치 — 건너뜀")
            continue
        words = []
        for f in fs:
            i = 0
            while (i < min(len(f["before"]), len(f["after"]))
                   and f["before"][i] == f["after"][i]):
                i += 1
            a, z = max(0, i - WIN), i + WIN
            ow = f["before"][a:z]
            nw = f["after"][a:z + len(f["빠진말"]) + 1]
            if ow not in cur:
                print(f"  ! {iid} '{f['빠진말']}' 자리 못 찾음 — 건너뜀")
                continue
            cur = cur.replace(ow, nw, 1)
            words.append(f["빠진말"])
            windows.append((f["book"], f["script_id"], ow, nw))
            log.append(dict(구분="듣기 대본 누락 복원", 급=f["book"], 과=f["chapter"],
                            item_id=iid, 내용=f"'{f['빠진말']}' 복원",
                            근거="구판 부록 듣기 지문"))
        if not words:
            continue
        ws.cell(r, c_text).value = cur
        ws.cell(r, c_rs).value = "fixed_v14"
        ws.cell(r, c_note).value = "구판 부록에 있는 " + ", ".join(
            f"'{w}'" for w in words) + " 누락 복원"

    print(f"1) 빠뜨린 말 복원 {len(windows)}건")

    ws2 = wb["n3_listen_script"]
    c_b, c_sid = col(ws2, "book_id"), col(ws2, "id")
    c_at, c_rs2 = col(ws2, "audio_text"), col(ws2, "review_status")
    n_at = 0
    for r in range(2, ws2.max_row + 1):
        b, sid = ws2.cell(r, c_b).value, ws2.cell(r, c_sid).value
        at = ws2.cell(r, c_at).value
        if not at:
            continue
        changed = False
        for fb, fs, old, new in windows:
            if int(b or 0) == fb and str(sid) == str(fs) and old in at:
                at = at.replace(old, new)
                changed = True
        if changed:
            ws2.cell(r, c_at).value = at
            ws2.cell(r, c_rs2).value = "fixed_v14"
            n_at += 1
    print(f"   audio_text 동기화 {n_at}개 대본")

    # ── 2) 7급 유키 → 유카
    n_yu = 0
    for name in ["n3_listen_script", "n3_listen_script_line", "n3_listen_repeat"]:
        w = wb[name]
        cb = col(w, "book_id")
        rs = col(w, "review_status")
        for r in range(2, w.max_row + 1):
            if int(w.cell(r, cb).value or 0) != 7:
                continue
            touched = False
            for c in range(1, w.max_column + 1):
                v = w.cell(r, c).value
                if isinstance(v, str) and "유키" in v:
                    w.cell(r, c).value = v.replace("유키", "유카")
                    touched = True
                    n_yu += 1
            if touched:
                w.cell(r, rs).value = "fixed_v14"
        log.append(dict(구분="인물명 교체", 급=7, 과="", item_id="",
                        내용=f"{name}: 유키 → 유카",
                        근거="신판 7급 본교재는 유카로 통일(유키 없음)"))
    print(f"2) 7급 유키 → 유카 {n_yu}칸")

    # ── 변경내역
    lg = wb.create_sheet("99_변경내역_v14")
    lg.append(["구분", "급", "과", "item_id", "내용", "근거"])
    for x in log:
        lg.append([x["구분"], x["급"], x["과"], x["item_id"], x["내용"], x["근거"]])
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거 CSV: verify/listen_gaps.csv, "
               f"verify/listen_provenance.csv, verify/listen_names.csv"])

    wb.save(DST)
    print(f"\n-> {DST}")


if __name__ == "__main__":
    main()
