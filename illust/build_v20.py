#!/usr/bin/env python3
"""v20 — n4 문법 태그·보기유형 부여.

태그 값은 **교재가 쓰는 한국어 대표형** 그대로다(사용자 결정). 1급 70건도
기존 의미 기반 태그(G-TOPIC…)에서 대표형(은/는…)으로 바꿔 8개 급을 한 체계로 맞춘다.
1급 대표형은 문법태그_1급_확정_v3.xlsx 문법목록 시트의 값을 그대로 쓴다.

배정 근거는 교재 과 첫 쪽의 ● 문법 목록이다. 문항의 grammar_focus·정답을
그 목록과 맞대어 붙인다. 붙지 않는 자리는 손으로 정한다(FALLBACK).

distractor_type은 보기에서 자동 판정한다 — 1급 70건 기록값과 70/70 일치를 확인했다.

산출: 글로벌_교재기반_콘텐츠_v20.xlsx (+ 문법목록 시트)
"""
import re, shutil, datetime, collections
import openpyxl
from n4_assign import load_syllabus, assign, sheet
from n4_inventory import distractor_type

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v19.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v20.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")

TAGDEF = "/Users/soohyeon/Documents/2608-yonsei_renewal/문법태그_1급_확정_v3.xlsx"

# 교재 목록과 문자열로 안 붙는 자리 — 근거를 적어 손으로 정한다.
FALLBACK = {
    (2, 2): lambda r: "에서 ~까지",          # 부터/를 대비 — 구간 표현 단원
    (2, 6): lambda r: "-으시-",              # -으세요/-세요는 -으시-의 실현형
    (2, 9): lambda r: "-었-",                # 과거 뒤 -어요 결합
    (2, 10): lambda r: "-은",                # 관형형 -은/-ㄴ
    (2, 11): lambda r: ("-는" if "는" in str(r["answer"]) else
                        "-을" if ("ㄹ" in str(r["answer"]) or "을" in str(r["answer"]))
                        else "-은"),
    (3, 11): lambda r: "-어",                # 반말
    (4, 5): lambda r: "-나요?",              # -은가요/ㄴ가요는 -나요?의 형용사형
    (5, 5): lambda r: "-는다고 하다",         # 명사 간접인용 이라고/라고
    (5, 7): lambda r: "-는 줄 알다",
    (5, 13): lambda r: "얼마나 -는지 모르다",
    (7, 6): lambda r: "-었던 것 같아요",
    (8, 5): lambda r: "-을래야 -을 수가 없다",  # 교재 표기, focus는 -(으)려야
}
# 급 범위를 벗어난 문항 — 태그는 실제 문법으로 달되 표시를 남긴다
OUT_OF_SCOPE = {
    "GF-3-14-002": ("-어 보이다",
                    "3급 14과 문법은 -는데2·-어 보다2인데 이 문항은 -어 보이다(4급 9과)를 묻는다"),
}


def classify(t):
    """대표형 꼴로 분류를 매긴다(자동). 1급은 확정 파일 값을 그대로 쓴다."""
    import re
    if (re.fullmatch(r"[ㄱ-ㅎ르으] ?(동사|불규칙)", t) or "불규칙" in t
            or "반말" in t or t in ("피동", "사동", "존댓말")):
        return "활용"
    if "화법" in t or "인용" in t:
        return "표현"
    if not t.startswith("-") and " " not in t and len(t) <= 6:
        return "조사"
    if t.rstrip("?").endswith(("다", "되다", "있다")):
        return "표현"
    if t.rstrip("?").endswith(("요", "까", "죠", "다니", "군요", "네요")):
        return "종결"
    return "연결"


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def load_g1_map():
    wb = openpyxl.load_workbook(TAGDEF, read_only=True)
    ws = wb["문법목록"]
    out, cls, ch = {}, {}, {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r[0]:
            continue
        out[r[0]] = r[1]
        cls[r[1]] = r[2]
        ch[r[1]] = r[3]
    return out, cls, ch


def main():
    syl = load_syllabus()
    g1map, g1cls, g1ch = load_g1_map()

    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n4_blank_question"]
    C = {k: col(ws, k) for k in ["book_id", "chapter", "answer", "grammar_focus",
                                 "completion", "selections", "selection1", "selection2",
                                 "selection3", "selection4", "item_id", "grammar_tag",
                                 "distractor_type", "review_status", "change_note",
                                 "hold_reason"]}
    stat = collections.Counter()
    tags = collections.defaultdict(lambda: {"급": None, "과": None, "n": 0})
    unresolved = []

    for r in range(2, ws.max_row + 1):
        b = ws.cell(r, C["book_id"]).value
        if not b:
            continue
        b, ch = int(b), int(ws.cell(r, C["chapter"]).value)
        iid = ws.cell(r, C["item_id"]).value
        row = {k: ws.cell(r, C[k]).value for k in C}

        if b == 1:
            tag = g1map.get(ws.cell(r, C["grammar_tag"]).value)
            stat["1급 대표형 전환" if tag else "1급 전환 실패"] += 1
        elif iid in OUT_OF_SCOPE:
            tag, why = OUT_OF_SCOPE[iid]
            ws.cell(r, C["hold_reason"]).value = why
            stat["급 범위 밖(표시)"] += 1
        else:
            it, s = assign(row, syl.get((b, ch), []))
            if s < 4 and (b, ch) in FALLBACK:
                it = FALLBACK[(b, ch)](row)
                stat["손으로 정함"] += 1
            elif s >= 4:
                stat["교재 목록 배정"] += 1
            tag = it
            if not tag:
                unresolved.append((b, ch, iid))
                stat["미배정"] += 1

        if tag:
            ws.cell(r, C["grammar_tag"]).value = tag
            k = (b, tag)
            t = tags[k]
            t["급"] = b
            t["과"] = ch if t["과"] is None else min(t["과"], ch)
            t["n"] += 1
        d = distractor_type(row)
        if d:
            ws.cell(r, C["distractor_type"]).value = d
        if b > 1:
            ws.cell(r, C["review_status"]).value = "tagged_v20"
            ws.cell(r, C["change_note"]).value = "교재 과 문법 목록 기준 태그 부여"

    print("부여 결과:", dict(stat))
    print("미배정:", len(unresolved), unresolved[:6])

    # ── 문법목록 시트
    if "문법목록" in wb.sheetnames:
        del wb["문법목록"]
    lg = wb.create_sheet("문법목록")
    lg.append(["급", "grammar_tag(대표형)", "분류", "도입_과", "문항수"])
    for (b, tag), v in sorted(tags.items(), key=lambda x: (x[0][0], x[1]["과"], x[0][1])):
        lg.append([b, tag, g1cls.get(tag) or classify(tag), v["과"], v["n"]])
    print(f"문법목록 {len(tags)}개")

    ch = wb.create_sheet("99_변경내역_v20")
    ch.append(["구분", "내용"])
    ch.append(["태그 체계", "grammar_tag 값을 교재 한국어 대표형으로 통일 — 1급 70건도 전환"])
    ch.append(["배정 근거", "교재 과 첫 쪽 ● 문법 목록 (파싱 누락 2급 4·5·8·12과는 교재에서 보강)"])
    ch.append(["distractor_type", "보기에서 자동 판정 — 1급 기록값과 70/70 일치 확인"])
    ch.append(["급 범위 밖", "GF-3-14-002 — 3급 14과인데 -어 보이다(4급 9과)를 묻는다. hold_reason 표시"])
    ch.append([f"작성 {datetime.date.today()}", "근거 CSV: verify/n4_inventory.csv"])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
