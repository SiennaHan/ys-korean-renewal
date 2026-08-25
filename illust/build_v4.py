#!/usr/bin/env python3
"""2~8급 글로벌 대조 결과를 v3에 반영해 v4를 만든다.

반영하는 것
  A. 전면 재저작 대화 5개(37턴) — 신판 한국어 + 새로 쓴 en/jp/cn/vi
  B. 화자 교체 5개 시나리오 — speaker/gender만 교정(한국어·번역 그대로)
  C. 부분 수정 3턴 — 단어 치환. 해당 어구만 번역에도 반영
  D. 어휘 구판 잔재 48건 — hold_reason 표기(삭제/교체는 사람 판단 필요)
  E. 어휘 과 이동 4건 — chapter 교정
  F. 교재 오류 교정 2건 기록

보류하는 것
  - 부분 수정 4턴: 신판 대응 턴 정렬이 어긋나 확정 못 함
  - n3 듣기 전체: 신판 부록(듣기 지문) 부재로 대조 불가
"""
import os, csv, shutil, collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v3.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v4.xlsx"
V = f"{HERE}/verify"

# ── C. 부분 수정 (구 어구 → 신 어구). 번역에도 같은 치환을 적용한다.
PARTIAL = {
    "RP-7-6-001": [("꿈속에서", "꿈에서")],
    "RP-7-8-003": [("쇼핑을 너무 많이 하는", "물건을 너무 많이 사는"),
                   ("shopped way too much", "bought way too many things"),
                   ("ショッピングしすぎて", "物を買いすぎて"),
                   ("购物买太多", "买太多东西"),
                   ("mua sắm quá tay", "mua quá nhiều đồ")],
    "RP-7-13-001": [("인터넷 사교 사이트에", "SNS에"),
                    ("on a social networking site", "on social media"),
                    ("ネットの交流サイトに", "SNSに"),
                    ("在网络社交网站上", "在社交媒体上"),
                    ("lên trang mạng xã hội", "lên mạng xã hội")],
}

# ── B. 화자 교체 (item_id 접두 → {구 화자: (신 화자, 성별)})
#
# 자동 탐지(speaker_check.py)는 '뭔가 어긋났다'는 것까지만 믿을 수 있다.
# 두 사람이 번갈아 말하는 대화에서는 발화 덩어리가 한 칸 밀려도 같은 치환이
# 반복돼 보여서, 도구가 알려주는 '새 화자 이름'이 틀린다. 실제로 4건 중 3건이
# 틀렸다(7급 2과는 마이클이 아니라 유카, 7급 14과는 유리, 5급 15과는 변경 없음).
# 그래서 아래는 전부 지면을 직접 읽어 확인한 것만 넣는다.
SPEAKER_FIX = {
    ("RP-7-10", "bonmun-10"): {"민철 엄마": ("옆집 아줌마", "여"),   # p104 확인
                               "옆집 아줌마": ("민철 엄마", "여")},
    ("RP-7-2", "bonmun-2"): {"민철": ("유카", "여")},               # p22 확인
    ("RP-7-14", "bonmun-14"): {"샤오밍": ("유리", "여")},           # p144 확인
    # 5급 15과 gwaje1-15 — 지면(p180)이 v3과 동일. 자동 탐지의 거짓 양성.
    # 5급 7과 gwaje1-7 — 신판 과제1(p84)에 보기 대화가 없어 확정 불가 → 보류.
}

# ── F. 교재 오류·데이터 오류 교정
TEXT_FIX = {
    "RP-4-13-009": ("전통 물건을 파하는인데", "전통 물건을 파는 곳인데",
                    "엑셀 오타 — 교재 원문은 '파는 곳인데'"),
    "RP-4-13-011": ("서울타워", "N서울타워", "신판에서 명칭 변경"),
}

HOLD = {
    "RP-5-7-008": "신판 과제1(p84)에 보기 대화가 없음 — 화자 교체 여부 확정 불가",
    "RP-5-7-010": "신판 과제1(p84)에 보기 대화가 없음 — 화자 교체 여부 확정 불가",
    "RP-5-6-004": "부분 수정 추정 — 신판 대응 턴 정렬 미확정",
    "RP-7-5-006": "부분 수정 추정 — 신판 대응 턴 정렬 미확정",
    "RP-7-10-009": "부분 수정 추정 — 신판 대응 턴 정렬 미확정",
    "RP-7-12-004": "부분 수정 추정 — 신판 대응 턴 정렬 미확정",
}


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    # ── A. 재저작 대화
    authored = {r["item_id"]: r for r in csv.DictReader(open(f"{V}/authored_dialogue.csv"))}
    ws = wb["n2_ai_role_play"]
    C = {n: col(ws, n) for n in ("item_id", "speaker", "gender", "ko", "en", "jp",
                                 "cn", "vi", "scenario_id", "review_status",
                                 "change_note", "hold_reason", "source_page")}
    done_a = done_b = done_c = done_f = done_h = 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        if not iid:
            continue
        sid = str(ws.cell(row, C["scenario_id"]).value)

        if iid in authored:
            a = authored[iid]
            for k in ("speaker", "ko", "en", "jp", "cn", "vi"):
                ws.cell(row, C[k]).value = a[k]
            ws.cell(row, C["gender"]).value = a["gender"] or None
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["change_note"]).value = (
                "신판 대화 전면 재저작 — 한국어는 본교재 원문, 번역 신규 저작")
            done_a += 1
            log.append(["n2 재저작", iid, "", a["ko"][:60], "신판 본문 대조 완료"])
            continue

        pre = "-".join(str(iid).split("-")[:3])   # RP-7-10-009 → RP-7-10
        fix = SPEAKER_FIX.get((pre, sid))
        if fix:
            cur = ws.cell(row, C["speaker"]).value
            if cur in fix:
                new, g = fix[cur]
                ws.cell(row, C["speaker"]).value = new
                ws.cell(row, C["gender"]).value = g
                ws.cell(row, C["change_note"]).value = f"신판 화자 교체 {cur}→{new}"
                done_b += 1
                log.append(["n2 화자교체", iid, cur, new, "대사·번역은 그대로"])

        if iid in PARTIAL:
            for k in ("ko", "en", "jp", "cn", "vi"):
                v = ws.cell(row, C[k]).value
                if not v:
                    continue
                for old, new in PARTIAL[iid]:
                    if old in str(v):
                        v = str(v).replace(old, new)
                ws.cell(row, C[k]).value = v
            ws.cell(row, C["change_note"]).value = "신판 어구 교체 반영"
            ws.cell(row, C["review_status"]).value = "reviewed"
            done_c += 1
            log.append(["n2 부분수정", iid, PARTIAL[iid][0][0], PARTIAL[iid][0][1], ""])

        if iid in TEXT_FIX:
            old, new, why = TEXT_FIX[iid]
            v = str(ws.cell(row, C["ko"]).value or "")
            if old in v:
                ws.cell(row, C["ko"]).value = v.replace(old, new)
                ws.cell(row, C["change_note"]).value = why
                done_f += 1
                log.append(["n2 교정", iid, old, new, why])

        if iid in HOLD:
            ws.cell(row, C["hold_reason"]).value = HOLD[iid]
            ws.cell(row, C["review_status"]).value = "draft"
            done_h += 1
            log.append(["n2 보류", iid, "", "", HOLD[iid]])

    # ── D·E. 어휘
    stale, moved = {}, {}
    for b in range(2, 9):
        for r in csv.DictReader(open(f"{V}/n1_b{b}.csv")):
            if r["verdict"] == "구판 잔재":
                stale[r["item_id"]] = r["word"]
            elif r["moved_to"]:
                moved[r["item_id"]] = (r["word"], r["moved_to"])
    ws = wb["n1_word_list"]
    Cn = {n: col(ws, n) for n in ("item_id", "chapter", "word", "review_status",
                                  "change_note", "hold_reason")}
    done_d = done_e = 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, Cn["item_id"]).value
        if iid in stale:
            ws.cell(row, Cn["hold_reason"]).value = (
                "신판 본교재에 없음 — 구판 잔재. 삭제/교체 판단 필요")
            ws.cell(row, Cn["review_status"]).value = "draft"
            done_d += 1
            log.append(["n1 구판잔재", iid, stale[iid], "", "삭제/교체 판단 필요"])
        elif iid in moved:
            w, to = moved[iid]
            ws.cell(row, Cn["change_note"]).value = f"신판에서 {to}로 이동 — 확인 필요"
            done_e += 1
            log.append(["n1 과이동", iid, w, to, ""])

    # ── 변경내역 시트
    name = "99_변경내역_v4"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구판/구값", "신판/신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "2~8급 글로벌 신판 전수 대조", "", "", "1급은 v1에서 완료"])
    sh.append(["대조 방법", "엑셀 문자열을 신판·구판 본문에서 각각 탐색 + 시나리오 묶음 유사도 + 화자 대조",
               "", "", "공백 제거 후 비교(신판은 자간을 좌표로 잡아 공백 글리프가 없음)"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    sh.append(["", "", "", "", ""])
    sh.append(["보류", "n3 듣기 전체", "", "",
               "듣기 대본은 부록 수록. 신판 PDF에 부록이 없어 대조 불가(1급과 동일)"])
    sh.append(["참고", "교재 자체 오류", "과 첫 쪽 어휘 목록", "본문 글로스",
               "8급 1과·6급 5과·13과: 본문만 새로 쓰고 첫 쪽 목록은 구판 그대로. 연세대 제보 대상"])
    for i, w in enumerate((14, 16, 46, 46, 60), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"A 재저작 {done_a}턴 / B 화자교체 {done_b}턴 / C 부분수정 {done_c}턴 / "
          f"F 교정 {done_f}건 / 보류 {done_h}턴")
    print(f"D 어휘 구판잔재 {done_d}건 / E 과이동 {done_e}건")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
