#!/usr/bin/env python3
"""캡션은 있지만 실제로는 그 단어를 그린 그림이 아닌 케이스를 지운다.

strip_groundless_images.py(v32→v33)는 "캡션 자체가 없는" 40건을 걸러냈다.
이번 건은 다르다 — 캡션은 있는데("선생님"), 그 캡션이 붙은 그림이 실은
빈칸을 채우는 말하기 연습 템플릿(밑줄 두 줄 + "씨," + "네")이었다. 같은
쪽(43쪽)에 "이현준 선생님"·"와타나베 유리 씨" 라벨이 붙은 채워진 예시
그림들도 있는데, 그 예시들도 "선생님/씨" 라는 존칭 표현 연습이 목적이고
실제로 교사를 그린 그림은 아니다 — 그래서 대안으로 바꿔 넣지 않고 그냥
지운다.

이 스크립트는 파이프라인이 아니라 사람이 육안으로 골라낸 개별 건을
기록하는 용도다 — item_id 를 코드에 나열한다. 자동 판정 규칙이 아니다
(text_cover 등 수치로는 이런 케이스를 못 가른다 — 빈 밑줄은 OCR에
안 걸려서 오히려 text_cover 가 낮게 나올 수도, 말풍선 테두리 때문에
높게 나올 수도 있다. 확인해보니 실제로 사람이 봐야 한다).

산출: 글로벌_교재기반_콘텐츠_v{n+1}.xlsx (image 값을 비운다)
"""
import glob
import os
import re
import shutil
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"

# (item_id, 지우는 이유)
MISLABELED = [
    (
        "VL-1-4-009",
        "'선생님' 캡션이 붙어 있었지만 실제 그림은 밑줄 두 줄을 채우는 "
        "존칭 호칭(이름+씨) 말하기 연습 템플릿(빈칸+'네')이었다 — 교사를 "
        "그린 그림이 아니다. 같은 43쪽의 채워진 예시들(이현준 선생님·"
        "와타나베 유리 씨)도 같은 연습 목적이라 대안으로 쓸 수 없다",
    ),
]


def newest_ledger():
    found = [(int(m.group(1)), p) for p in glob.glob(f"{ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
              if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(found)


def main():
    n, src = newest_ledger()
    dst = f"{ROOT}/글로벌_교재기반_콘텐츠_v{n + 1}.xlsx"
    if os.path.exists(dst):
        sys.exit(f"중단: {dst} 가 이미 있다.")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    ws = wb["n1_word_list"]
    hdr = [c.value for c in ws[1]]
    C = {h: i + 1 for i, h in enumerate(hdr) if h}
    by_id = dict(MISLABELED)
    cleared = 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        reason = by_id.get(iid)
        if not reason:
            continue
        old = ws.cell(row, C["image"]).value
        ws.cell(row, C["image"]).value = ""
        note = ws.cell(row, C["change_note"]).value or ""
        add = f"이미지 제거 — '{old}': {reason}"
        ws.cell(row, C["change_note"]).value = (note + " / " + add) if note else add
        cleared += 1

    lg = wb.create_sheet(f"99_변경내역_v{n + 1}")
    lg.append(["구분", "대상", "내용"])
    lg.append(["이미지 제거", "n1_word_list",
                f"{cleared}건 — 캡션은 있으나 실제로는 그 단어를 그린 그림이 아니었던 "
                "개별 오매칭(사람이 육안 확인). 선생님(존칭 호칭 연습 템플릿)"])
    wb.save(dst)
    print(f"{cleared}건 제거 -> {dst}")


if __name__ == "__main__":
    main()
