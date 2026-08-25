#!/usr/bin/env python3
"""v19 — 어휘 번역 빈칸 167건 채우기.

출처가 둘이다.
  99건 — 본교재 어휘 면에 영·일·중 대역이 실려 있다. 그쪽 값을 그대로 쓴다.
         (원래 파싱은 본문 오른쪽 뜻풀이 열만 봐서 이 면을 놓쳤다.)
  68건 — 요일·그림 낱말·동작 표현구처럼 교재에 대역이 없는 말들. 새로 저작한다.
         잘려 들어간 영어도 여기서 바로잡는다.

베트남어는 교재에 아예 없으므로 167건 모두 저작한다.

일본어·중국어를 읽으려면 MS-PMincho·KaiTi 글리프 브리지가 있어야 한다.
없으면 57~72%가 깨져 나온다. glyph_bridge.py main 을 먼저 돌릴 것.

산출: 글로벌_교재기반_콘텐츠_v19.xlsx
"""
import csv, re, shutil, datetime, unicodedata, collections
import openpyxl
from v19_data import AUTHORED, VI, OVERRIDE

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v18.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v19.xlsx"


def norm(w):
    w = unicodedata.normalize("NFC", str(w or ""))
    w = re.sub(r"[(（][^)）]*[)）]", "", w)
    return re.sub(r"[\s·.?!]+", "", w).strip("-").strip()


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    tr = collections.defaultdict(list)
    for r in csv.DictReader(open("verify/vocab_trans.csv")):
        tr[(int(r["book"]), norm(r["word"]))].append(r)

    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n1_word_list"]
    C = {k: col(ws, k) for k in ["book_id", "chapter", "word", "en", "jp", "cn", "vi",
                                 "item_id", "review_status", "change_note"]}
    log = []
    n_book = n_auth = 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        if iid in VI:
            b = int(ws.cell(r, C["book_id"]).value)
            ch = ws.cell(r, C["chapter"]).value
            w = ws.cell(r, C["word"]).value
            k = (b, norm(w))
            # 동형이의어가 있으므로 같은 과의 것을 먼저 고른다
            cand = [x for x in tr[k] if str(x["chapter"]) == str(ch)] or tr[k]
            t = cand[0]
            jp, cn = OVERRIDE.get(iid, (t["jp"], t["cn"]))
            ws.cell(r, C["en"]).value = t["en"]
            ws.cell(r, C["jp"]).value = jp
            ws.cell(r, C["cn"]).value = cn
            ws.cell(r, C["vi"]).value = VI[iid]
            ws.cell(r, C["review_status"]).value = "filled_v19"
            ws.cell(r, C["change_note"]).value = f"본교재 어휘 면(p{t['page']}) 대역 반영 · 베트남어 저작"
            n_book += 1
            log.append(("교재 대역", iid, w, t["en"], jp, cn, VI[iid]))
        elif iid in AUTHORED:
            en, jp, cn, vi = AUTHORED[iid]
            w = ws.cell(r, C["word"]).value
            if en:
                ws.cell(r, C["en"]).value = en
            ws.cell(r, C["jp"]).value = jp
            ws.cell(r, C["cn"]).value = cn
            ws.cell(r, C["vi"]).value = vi
            ws.cell(r, C["review_status"]).value = "filled_v19"
            ws.cell(r, C["change_note"]).value = "교재에 대역이 없어 저작"
            n_auth += 1
            log.append(("저작", iid, w, ws.cell(r, C["en"]).value, jp, cn, vi))
    print(f"교재 대역 {n_book}건 / 저작 {n_auth}건 = {n_book + n_auth}건")

    # 남은 빈칸 확인
    left = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, C["review_status"]).value) == "deleted":
            continue
        if not ws.cell(r, C["book_id"]).value:
            continue
        empt = [k for k in ["en", "jp", "cn", "vi"] if not ws.cell(r, C[k]).value]
        if empt:
            left.append((ws.cell(r, C["item_id"]).value, ws.cell(r, C["word"]).value, empt))
    print(f"남은 번역 빈칸: {len(left)}건" + ("" if not left else f" {left[:5]}"))

    lg = wb.create_sheet("99_변경내역_v19")
    lg.append(["출처", "item_id", "낱말", "en", "jp", "cn", "vi"])
    for x in log:
        lg.append(list(x))
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거 CSV: verify/vocab_trans.csv"])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
