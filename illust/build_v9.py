#!/usr/bin/env python3
"""정렬이 안 잡혀 남아 있던 구판 문장 7턴을 확정해 v9를 만든다.

지면을 직접 읽어 갈랐다.

  4급 10과 과제1 — 6턴이 5턴으로 재구성됐다. 인사말 턴('그래? 긴장되겠구나.')이
    빠지고, 마지막 턴이 '단 음식을 좋아하실지 모르겠어 / 건강 차를 선물해'로
    바뀌었다. 5턴을 다시 쓰고 6번째 턴은 닫는다.

  부분 수정 4턴
    RP-5-9-003   신분증이나 의료보험증을 → 신분증을      (어휘 '의료보험증' 폐기와 짝)
    RP-5-11-018  마이클 씨 스마트폰에 댄스 음악이 → 요즘 유행하는 노래
    RP-7-13-004  사교 사이트 → SNS, 그리고 '주고받하는이라서' 손상까지 교정
    RP-7-13-006  나도 내가 → 저도 제가                   (화계 통일)

  RP-7-13-004는 손상 스캔이 놓친 건이다. '주고받을 수 있는 곳이라서'가
  '주고받하는이라서'로 깨져 있었는데 내 패턴에 안 걸렸다. 내용 변경과 손상이
  한 문장에 겹쳐 있던 경우다.

산출: 글로벌_교재기반_콘텐츠_v9.xlsx
"""
import os, re, shutil, unicodedata
import openpyxl
from global_text import GlobalPdf

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v8.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v9.xlsx"

# 통째로 다시 쓰는 시나리오: item_id -> (화자, ko, en, jp, cn, vi)
REWRITE = {
    "RP-4-10-007": ("친구",
        "지나야, 무슨 생각을 그렇게 하고 있니?",
        "Jina, what are you thinking about so hard?",
        "ジナ、何をそんなに考えてるの？",
        "智娜，你在想什么想得那么入神？",
        "Jina à, cậu đang nghĩ gì mà chăm chú thế?"),
    "RP-4-10-008": ("지나",
        "사실은 다음 주에 남자 친구 부모님께 처음 인사를 드리러 가. 뭘 사 가지고 가면 좋아하실까?",
        "Actually, next week I'm going to meet my boyfriend's parents for the first time. "
        "What should I buy to take along that they'd like?",
        "実は来週、彼氏のご両親に初めて挨拶しに行くの。何を買って持っていったら喜んでくださるかな。",
        "其实下周我要第一次去见男朋友的父母。买什么带过去他们会喜欢呢？",
        "Thật ra tuần sau mình sẽ đến chào bố mẹ bạn trai lần đầu tiên. "
        "Mua gì mang đi thì hai bác sẽ thích nhỉ?"),
    "RP-4-10-009": ("친구",
        "음, 그럼 과일이나 맛있는 케이크를 사 가지고 가는 게 어때? 값도 별로 비싸지 않고 가족들과 다 같이 먹을 수도 있잖아.",
        "Hmm, then how about buying some fruit or a tasty cake to take along? "
        "It's not that expensive, and everyone in the family can eat it together.",
        "うーん、じゃあ果物かおいしいケーキを買って持っていくのはどう？"
        "値段もそんなに高くないし、家族みんなで一緒に食べられるじゃない。",
        "嗯，那买些水果或者好吃的蛋糕带过去怎么样？价格也不太贵，还能跟家人一起吃啊。",
        "Ừm, vậy thì mua trái cây hoặc bánh kem ngon mang đi thì sao? "
        "Giá cũng không đắt lắm, mà cả nhà còn có thể cùng nhau ăn nữa mà."),
    "RP-4-10-010": ("지나",
        "그런데 남자 친구 부모님께서 단 음식을 좋아하실지 잘 모르겠어.",
        "But I'm not sure whether my boyfriend's parents like sweet things.",
        "でも、彼氏のご両親が甘いものをお好きかどうか分からなくて。",
        "可是我不确定男朋友的父母喜不喜欢吃甜的。",
        "Nhưng mình không chắc bố mẹ bạn trai có thích đồ ngọt không."),
    "RP-4-10-011": ("친구",
        "그럼 건강에 좋은 건강 차를 선물해.",
        "Then give them some healthy tea as a gift.",
        "じゃあ、体にいい健康茶をプレゼントしなよ。",
        "那就送对身体好的养生茶吧。",
        "Vậy thì tặng trà thảo mộc tốt cho sức khỏe đi."),
}
CLOSE = {"RP-4-10-012": "신판에서 6턴 → 5턴으로 줄면서 이 턴이 없어짐"}

# 어구 치환: item_id -> {필드: (구, 신)}
EDIT = {
    "RP-5-9-003": dict(
        ko=("신분증이나 의료보험증을", "신분증을"),
        en=("your ID card or health insurance card", "your ID card"),
        jp=("身分証か医療保険証を", "身分証を"),
        cn=("你的身份证或者医疗保险证", "您的身份证"),
        vi=("chứng minh thư hoặc thẻ bảo hiểm y tế", "chứng minh thư"),
        note="신판 문장 교체 — 의료보험증 삭제(어휘 폐기와 짝)"),
    "RP-5-11-018": dict(
        ko=("마이클 씨 스마트폰에 댄스 음악이 많이 있잖아요.", "요즘 유행하는 노래 있잖아요."),
        en=("Michael, you have a lot of dance music on your smartphone, right?",
            "You know those songs that are popular these days?"),
        jp=("マイケルさんのスマホにダンス音楽がたくさんあるじゃないですか。",
            "最近流行っている歌があるじゃないですか。"),
        cn=("迈克你手机里不是有很多舞曲嘛。", "最近不是有很流行的歌嘛。"),
        vi=("Anh Michael có nhiều nhạc dance trong điện thoại mà.",
            "Có mấy bài hát đang thịnh hành dạo này mà."),
        note="신판 문장 교체"),
    "RP-7-13-004": dict(
        ko=("사교 사이트는 친구들을 사귀고 서로 연락을 주고받하는이라서",
            "SNS는 친구들을 새로 사귀고 서로 연락을 주고받을 수 있는 곳이라서"),
        en=("Since a social networking site is a place to make friends and stay in touch with each other",
            "Since social media is a place where you can make new friends and stay in touch with each other"),
        jp=("交流サイトって友達を作ったりお互いに連絡を取り合ったりする場所だから",
            "SNSって友達を新しく作ったりお互いに連絡を取り合ったりできる場所だから"),
        cn=("社交网站是交朋友、彼此联络的地方",
            "社交媒体是可以结交新朋友、彼此联络的地方"),
        vi=("Trang mạng xã hội là nơi kết bạn và liên lạc qua lại với nhau",
            "Mạng xã hội là nơi có thể kết bạn mới và liên lạc qua lại với nhau"),
        note="신판 문장 교체(사교 사이트→SNS) + 일괄 치환 손상 교정(주고받하는이라서)"),
    "RP-7-13-006": dict(
        ko=("나도 내가 올린", "저도 제가 올린"),
        en=("", ""), jp=("", ""), cn=("", ""), vi=("", ""),
        note="신판 화계 통일 (나도 내가 → 저도 제가). 번역은 영향 없음"),
}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i


def main():
    blobs = {}
    for b in (4, 5, 7):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        blobs[b] = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))

    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n2_ai_role_play"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "speaker", "ko", "en", "jp",
                                 "cn", "vi", "review_status", "change_note",
                                 "hold_reason")}
    log, ok, fail = [], 0, 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        b = ws.cell(row, C["book_id"]).value

        if iid in REWRITE:
            sp, ko, en, jp, cn, vi = REWRITE[iid]
            if sq(ko)[:22] not in blobs[int(b)]:
                fail += 1
                log.append(["재저작 보류", iid, "", ko[:52], "본문 대조 실패"])
                continue
            ws.cell(row, C["speaker"]).value = sp
            for k, v in (("ko", ko), ("en", en), ("jp", jp), ("cn", cn), ("vi", vi)):
                ws.cell(row, C[k]).value = v
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = "신판 과제1 대화 재구성(6턴→5턴)에 따라 재저작"
            ok += 1
            log.append(["시나리오 재구성", iid, "", ko[:52], "본문 대조 통과"])

        elif iid in CLOSE:
            ws.cell(row, C["review_status"]).value = "deleted"
            ws.cell(row, C["change_note"]).value = CLOSE[iid]
            ws.cell(row, C["hold_reason"]).value = None
            log.append(["턴 폐기", iid, str(ws.cell(row, C["ko"]).value)[:52], "", CLOSE[iid]])

        elif iid in EDIT:
            e = EDIT[iid]
            before = str(ws.cell(row, C["ko"]).value or "")
            newko = before.replace(*e["ko"]).strip()
            if sq(newko)[:24] not in blobs[int(b)]:
                fail += 1
                log.append(["수정 보류", iid, before[:52], newko[:52], "본문 대조 실패"])
                continue
            ws.cell(row, C["ko"]).value = newko
            for k in ("en", "jp", "cn", "vi"):
                old, new = e[k]
                if old:
                    v = str(ws.cell(row, C[k]).value or "")
                    ws.cell(row, C[k]).value = v.replace(old, new)
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = e["note"]
            ok += 1
            log.append(["부분 수정", iid, before[:52], newko[:52], "본문 대조 통과"])

    name = "99_변경내역_v9"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "정렬 미확정으로 남아 있던 구판 문장 7턴", "", "",
               "시나리오 재구성 1개(5턴) + 턴 폐기 1 + 부분 수정 4"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((16, 16, 46, 46, 44), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"반영 {ok}건 / 폐기 {len(CLOSE)}건 / 실패 {fail}건")
    for r in log:
        print(f"  [{r[0]}] {r[1]}  {r[3][:56] or r[2][:56]}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
