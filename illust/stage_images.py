#!/usr/bin/env python3
"""links.csv 가 가리키는 글로벌 삽화를 앱의 public/textbook/ 에 내려놓는다.

**480px 상한을 720px로 올렸다(2026-09-05).** 480은 구판 자산(평균 40KB짜리
jpg, 365~472px) 크기에 맞춘 값이었는데, 실제로 필요한 건 화면 표시 칸과
레티나 배율(2~3배)을 곱한 값이었다 — 단어 퀴즈 화면(228px 정사각)을 3배
폰에서 채우려면 684px, 낱말 학습(120px)은 360px가 필요하다. 480 상한에선
퀴즈 화면 575장 중 574장이 확대(업스케일)돼 흐리게 보였다. 720이면 그 684px
요구를 여유 있게 덮는다 — DPI 도 300→600 으로 같이 올려서(extract_global_
images.py) 이 상한이 실제로 의미가 있게 만들었다.

원장의 image 값은 .png 인데 여기서는 .jpg 로 저장한다 — apply_images.py 로
만든 원장을 이 스크립트 뒤에 한 번 더 손봐서 확장자를 맞춘다(v31 -> v32).

산출: app/public/textbook/{book}/{filename}.jpg
      글로벌_교재기반_콘텐츠_v{n+1}.xlsx (확장자 .png -> .jpg 로 바꾼 사본)
"""
import csv, glob, os, re, shutil, sys

import openpyxl
from PIL import Image

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
PUBLIC = f"{ROOT}/app/public/textbook"
MAX_SIDE = 720
QUALITY = 85


def newest_ledger():
    found = [(int(m.group(1)), p) for p in glob.glob(f"{ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
              if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(found)


def stage():
    rows = [r for r in csv.DictReader(open(f"{HERE}/links.csv")) if r["filename"]]
    need = {(int(r["book"]), r["filename"]) for r in rows}
    written = 0
    for book, fn in sorted(need):
        src = f"{HERE}/images/b{book}/{fn}"
        dst_dir = f"{PUBLIC}/{book}"
        os.makedirs(dst_dir, exist_ok=True)
        dst = f"{dst_dir}/{fn.replace('.png', '.jpg')}"
        im = Image.open(src).convert("RGB")
        im.thumbnail((MAX_SIDE, MAX_SIDE))
        im.save(dst, "JPEG", quality=QUALITY)
        written += 1
    total = sum(os.path.getsize(f"{PUBLIC}/{b}/{fn.replace('.png', '.jpg')}")
                for b, fn in need)
    print(f"{written}개 저장 -> {PUBLIC}/{{book}}/  (총 {total/1e6:.1f}MB)")


def swap_extension():
    """방금 만든 원장(apply_images.py 산출)의 .png 값을 .jpg 로 바꾼 새 버전."""
    n, src = newest_ledger()
    dst = f"{ROOT}/글로벌_교재기반_콘텐츠_v{n+1}.xlsx"
    if os.path.exists(dst):
        sys.exit(f"중단: {dst} 가 이미 있다.")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    changed = 0
    for sheet_name, cols in (("n1_word_list", ["image"]),
                              ("n1_word_quiz", ["image"]),
                              ("n3_listen_repeat",
                               [f"selection{i}_image" for i in range(1, 5)])):
        ws = wb[sheet_name]
        hdr = [c.value for c in ws[1]]
        C = {h: i + 1 for i, h in enumerate(hdr) if h}
        for col in cols:
            if col not in C:
                continue
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, C[col])
                v = cell.value
                if v and str(v).endswith(".png"):
                    cell.value = str(v)[:-4] + ".jpg"
                    changed += 1
    wb.save(dst)
    print(f"확장자 .png -> .jpg : {changed}건")
    print(f"-> {dst}")


if __name__ == "__main__":
    stage()
    swap_extension()
