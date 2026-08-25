#!/usr/bin/env python3
"""links.csv(글로벌 삽화 재연결 결과)를 원장에 실제로 써넣는다.

build_catalog.py 가 계산까지는 8/14에 끝내 놨는데, 그 결과(links.csv)를
원장에 쓰는 스크립트가 없어서 지금까지 한 번도 반영되지 않았다 —
n1_word_list.image 가 지금도 구판(3주완성) 파일명을 그대로 물고 있다.

값은 맨 파일명만 쓴다(디렉터리 접두어 없음) — 프론트가
`/textbook/{book_id}/{image}` 로 붙이는 지금 관례와 같다(word-learning.tsx·
word-quiz-card.tsx). 구판 값도 이미 그 관례였으니 자리만 바꾼다.

⚠️ 이 스크립트는 원장의 image 값을 "최신 글로벌판 파일명"으로 바꿀 뿐이다.
그 파일명이 가리키는 실제 PNG(`illust/images/b*/`, 중복 없는 사본은
`illust/by_item/`)를 앱이 읽는 자산 서버(`/textbook/` 또는
PUBLIC_RES_URL_ROOT)에 올리는 것은 별개의 배포 작업이다 — 올리기 전에는
화면에서 404 난다.

산출: 글로벌_교재기반_콘텐츠_v{n+1}.xlsx
"""
import csv, glob, os, re, shutil, sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"


def newest_ledger():
    found = [(int(m.group(1)), p) for p in glob.glob(f"{ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
              if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(found)[0], max(found)[1]


def main():
    n, src = newest_ledger()
    dst = f"{ROOT}/글로벌_교재기반_콘텐츠_v{n+1}.xlsx"
    if os.path.exists(dst):
        sys.exit(f"중단: {dst} 가 이미 있다.")
    shutil.copy(src, dst)
    print(f"원장  {os.path.basename(src)} -> {os.path.basename(dst)}")

    rows = [r for r in csv.DictReader(open(f"{HERE}/links.csv")) if r["filename"]]
    by_sheet = {}
    for r in rows:
        by_sheet.setdefault(r["sheet"], []).append(r)

    LOW_CONF = {"라벨 우선(좌표와 불일치)", "좌표 우선(라벨과 불일치)",
                "좌표+라벨 일치(후보多)", "좌표 우선(라벨과 불일치)(후보多)"}

    wb = openpyxl.load_workbook(dst)
    changed = {"n1_word_list": 0, "n1_word_quiz": 0, "n3_listen_repeat": 0}
    flagged = 0

    for sheet_name in ("n1_word_list", "n1_word_quiz"):
        ws = wb[sheet_name]
        hdr = [c.value for c in ws[1]]
        C = {h: i + 1 for i, h in enumerate(hdr) if h}
        by_item = {r["item_id"]: r for r in by_sheet.get(sheet_name, [])}
        for row in range(2, ws.max_row + 1):
            iid = ws.cell(row, C["item_id"]).value
            link = by_item.get(iid)
            if not link:
                continue
            old = ws.cell(row, C["image"]).value
            ws.cell(row, C["image"]).value = link["filename"]
            note = ws.cell(row, C["change_note"]).value or ""
            add = f"글로벌판 삽화로 재배선({old} → {link['filename']}, {link['how']})"
            ws.cell(row, C["change_note"]).value = (note + " / " + add) if note else add
            if link["how"] in LOW_CONF:
                hold = ws.cell(row, C["hold_reason"]).value or ""
                flag = "이미지 재배선 확인 필요 — 좌표·라벨 판정 불일치"
                ws.cell(row, C["hold_reason"]).value = (hold + " / " + flag) if hold else flag
                flagged += 1
            changed[sheet_name] += 1

    ws = wb["n3_listen_repeat"]
    hdr = [c.value for c in ws[1]]
    C = {h: i + 1 for i, h in enumerate(hdr) if h}
    by_item_lr = {}
    for r in by_sheet.get("n3_listen_repeat", []):
        base, _, seq = r["item_id"].partition("#s")
        by_item_lr.setdefault(base, {})[int(seq)] = r
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        slots = by_item_lr.get(iid)
        if not slots:
            continue
        touched = []
        for seq, link in slots.items():
            col = f"selection{seq}_image"
            if col not in C:
                continue
            old = ws.cell(row, C[col]).value
            ws.cell(row, C[col]).value = link["filename"]
            touched.append(f"{col}: {old} → {link['filename']}({link['how']})")
            if link["how"] in LOW_CONF:
                hold = ws.cell(row, C["hold_reason"]).value or ""
                flag = f"{col} 재배선 확인 필요 — 좌표·라벨 판정 불일치"
                ws.cell(row, C["hold_reason"]).value = (hold + " / " + flag) if hold else flag
                flagged += 1
        if touched:
            note = ws.cell(row, C["change_note"]).value or ""
            add = "글로벌판 삽화로 재배선 — " + " / ".join(touched)
            ws.cell(row, C["change_note"]).value = (note + " / " + add) if note else add
            changed["n3_listen_repeat"] += 1

    lg = wb.create_sheet(f"99_변경내역_v{n+1}")
    lg.append(["구분", "대상", "내용"])
    lg.append(["삽화 재배선", "n1_word_list", f"{changed['n1_word_list']}행"])
    lg.append(["삽화 재배선", "n1_word_quiz", f"{changed['n1_word_quiz']}행"])
    lg.append(["삽화 재배선", "n3_listen_repeat", f"{changed['n3_listen_repeat']}행(선택지 단위)"])
    lg.append(["확인 필요 표시", "", f"{flagged}건 — 좌표·라벨 판정이 불일치했던 자리"])
    lg.append(["미해결", "", "8건 — needs.csv. 그중 3건은 image 칸에 파일명이 아니라 "
                              "발음표기(예: 닭갈비[닥깔비])가 들어 있던 원장 자체 오류"])
    lg.append(["⚠️ 배포 전 필요", "", "illust/by_item/(825개) 를 앱 자산 서버(/textbook/ 또는 "
                                     "PUBLIC_RES_URL_ROOT)에 올려야 실제로 뜬다 — "
                                     "지금은 로컬에만 있다"])

    wb.save(dst)
    print(f"어휘 {changed['n1_word_list']}건 · 단어퀴즈 {changed['n1_word_quiz']}건 · "
          f"듣기선택지 {changed['n3_listen_repeat']}건 재배선")
    print(f"확인 필요 표시 {flagged}건")
    print(f"-> {dst}")


if __name__ == "__main__":
    main()
