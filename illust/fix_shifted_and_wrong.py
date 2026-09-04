#!/usr/bin/env python3
"""600dpi 재추출 후 파일명이 안 맞은 2건을 손으로 확인해 정리한다.

재추출은 페이지 안 그림 순번(idx)이 이전 실행과 달라질 수 있다 — 감지 순서가
안정적이지 않다. 603개 참조 중 601개는 이름이 그대로 맞았고, 2개만 어긋났다.
육안으로 확인했다:

  b3_ch6_p74_21.jpg (샌드위치) → 실제로는 건물·교차로 지도 삽화였다.
    이미 근거 없는 연결이었다(재추출과 무관, 전부터 그랬다) — 대안을 찾지 않고
    지운다. strip_groundless_images.py 의 원칙과 같다: 틀린 그림보다 없는 게 낫다.

  b4_ch4_p49_5.jpg (열람실·개방 시간) → 새 추출에서 같은 장면이 b4_ch4_p49_8.png
    로 다시 잡혔다(옆의 9는 "반납" 쪽만 잘린 다른 조각). 8번이 "열람실"
    간판과 "개방 시간" 안내문을 둘 다 담고 있어 정확히 대응한다 — 파일명만 바꾼다.
"""
import glob
import os
import re
import shutil
import sys

import openpyxl

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"

RENAME = {"b4_ch4_p49_5.jpg": "b4_ch4_p49_8.jpg"}
DROP = {"b3_ch6_p74_21.jpg": "재추출 후 확인해 보니 실제로는 건물·교차로 지도 삽화였다 "
        "(원래부터 근거 없는 연결) — 대안 없이 지운다"}


def newest():
    f = [(int(m.group(1)), p) for p in glob.glob(f"{ROOT}/글로벌_교재기반_콘텐츠_v*.xlsx")
         if (m := re.match(r".*_v(\d+)\.xlsx$", p))]
    return max(f)


def main():
    n, src = newest()
    dst = f"{ROOT}/글로벌_교재기반_콘텐츠_v{n + 1}.xlsx"
    if os.path.exists(dst):
        sys.exit(f"중단: {dst} 가 이미 있다.")
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["n1_word_list"]
    C = {h: i + 1 for i, h in enumerate([c.value for c in ws[1]]) if h}

    renamed = dropped = 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, C["item_id"]).value:
            continue
        img = ws.cell(r, C["image"]).value
        if img in RENAME:
            ws.cell(r, C["image"]).value = RENAME[img]
            note = ws.cell(r, C["change_note"]).value or ""
            add = f"600dpi 재추출로 파일명 변경: {img} -> {RENAME[img]}(같은 장면, idx만 밀림)"
            ws.cell(r, C["change_note"]).value = (note + " / " + add) if note else add
            renamed += 1
        elif img in DROP:
            ws.cell(r, C["image"]).value = ""
            note = ws.cell(r, C["change_note"]).value or ""
            add = f"이미지 제거 — '{img}': {DROP[img]}"
            ws.cell(r, C["change_note"]).value = (note + " / " + add) if note else add
            dropped += 1

    log = wb["99_변경내역"]
    from openpyxl.styles import Font, PatternFill
    rows = [[f"v{n + 1}   (2건)", "", "", "", "", "", "", ""],
            [f"v{n + 1}", "정리", "n1_word_list", "", "", "",
             f"600dpi 재추출 후 idx 밀림 정리 — 이름 변경 {renamed}건, 근거 없던 연결 제거 {dropped}건", ""]]
    log.insert_rows(2, amount=len(rows))
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            log.cell(2 + i, j).value = v
    for c in log[2]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEDFF")

    wb.save(dst)
    print(f"이름 변경 {renamed}건 · 제거 {dropped}건 -> {dst}")


if __name__ == "__main__":
    main()
