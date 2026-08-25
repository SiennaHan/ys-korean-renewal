#!/usr/bin/env python3
"""살아 있는 어휘 중 신판 본교재 어디에도 없는 것을 찾는다.

색인 대조를 하다 8급 11과 '삼차원 프린터'가 걸렸다. 짝인 '입체 프린터'는 이미
폐기됐는데 이쪽만 살아 있었다 — 신판에서 3D 프린터 내용 자체가 빠졌기 때문이다.
같은 잔재가 더 있는지 급 전체로 훑는다(과 안에서만 보면 놓친다).

산출: verify/ledger_sweep.csv
"""
import os, re, csv, collections, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v18.xlsx"
APX = "/Users/soohyeon/Documents/2608-yonsei_renewal/book"


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/…]", "", t)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n1_word_list"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)
            if r[hdr.index("book_id")]]

    out = []
    for b in range(1, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        main_b = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))
        ap = GlobalPdf(f"{APX}/(최종본)연세글로벌한국어_{b}급_부록-최종(26.8.10).pdf")
        apx_b = sq(" ".join(ap.text(p) for p in range(1, len(ap) + 1)))
        n = 0
        for r in rows:
            if int(r["book_id"]) != b:
                continue
            if str(r.get("review_status")) == "deleted":
                continue
            w = sq(r.get("word"))
            if len(w) < 2 or w in main_b:
                continue
            n += 1
            out.append(dict(book=b, chapter=r.get("chapter"), item_id=r.get("item_id"),
                            word=r.get("word"), category=r.get("category"),
                            부록="O" if w in apx_b else "X",
                            review_status=r.get("review_status")))
        print(f"  {b}급 신판 본교재에 없는 어휘 {n}건")

    with open(f"{HERE}/verify/ledger_sweep.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]) if out else ["book"])
        wr.writeheader()
        wr.writerows(out)
    hard = [o for o in out if o["부록"] == "X"]
    print(f"\n■ 본교재·부록 어디에도 없음 {len(hard)}건")
    for o in hard:
        print(f"   {o['book']}급 {o['chapter']:>2}과 {o['item_id']:14s} {o['word']}  ({o['category']})")
    print(f"\n(부록에는 있음 {len(out)-len(hard)}건 — 부록 문법연습·모범답안에 쓰인 말)")
    print(f"-> verify/ledger_sweep.csv")


if __name__ == "__main__":
    main()
