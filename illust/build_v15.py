#!/usr/bin/env python3
"""v15 — 신판에서 고쳐진 듣기 대본 7건을 신판 쪽으로 맞춘다.

구판 내용이 틀린 것은 아니지만 물건이 낡았거나(전자사전·드라마 DVD)
지금 기준으로 걸리는 장면(직원이 선생님 전화번호를 알려 줌)이라
신판이 일부러 고친 자리들이다.

대본만 바꾸면 문항이 어긋난다. 그래서 문항 보기·발문도 같이 손본다.
특히 6급 13과는 신판에서 '전자 잡지'가 빠지고 '스마트 렌즈'가 들어와
'소개하지 않은 것' 정답이 둘이 되므로 보기를 바꿔 유일성을 되살린다.

띄어쓰기는 신판 부록에서 뽑을 수 없다(자간을 좌표로 주어 공백 글리프가 없다).
그래서 문장을 직접 적고, 공백을 턴 문자열이 신판 부록과 정확히 같은지 대조한다.

산출: 글로벌_교재기반_콘텐츠_v15.xlsx
"""
import re, shutil, datetime, unicodedata
import openpyxl
from global_text import GlobalPdf

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v14.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v15.xlsx"
APX = "/Users/soohyeon/Documents/2608-yonsei_renewal/book"

# ── 대본 교체 (item_id → 새 text; speaker 바꿀 것은 (text, speaker))
LINES = {
 # 6급 4과 — 전자사전 → 태블릿 피시
 "LL-6-4-001": "태블릿 피시 좀 빌려 줄래?",
 "LL-6-4-006": "태블릿 피시를 그렇게 관리하면 안 되지. 나는 태블릿 피시가 떨어지지 않게 조심하고 "
               "닦을 때도 젖은 수건으로 닦지 않고 마르고 부드러운 천으로 닦아. "
               "그리고 화면을 누를 때도 고장이 나지 않게 너무 세게 누르지 말고.",
 "LL-6-4-007": "야, 네 태블릿 피시는 새것 같다. 언제 산 거야?",
 "LL-6-4-008": "5년 전에 산 건데 아직 한 번도 고장 난 적이 없어.",
 "LL-6-4-009": "너는 태블릿 피시를 정말 잘 관리하고 있구나. 조심해야 할 점이 있으면 더 알려 줘.",
 "LL-6-4-010": "지난번에 보니까 태블릿 피시를 그냥 가방에 바로 넣던데 그렇게 하면 안 돼. "
               "화면 부분이 깨질 수 있거든. 꼭 케이스에 넣어서 보관해야 해.",
 # 3급 13과 — 드라마 DVD → 한국 과자
 "LL-3-13-012": "저는 한국 과자를 선물할 거예요. 대만에서 한국 과자가 인기가 많기 때문에 "
                "친구들이 좋아할 것 같아요.",
 # 6급 13과 — 터치스크린 냉장고·전자 잡지 → 인공지능 냉장고·홀로그램·스마트 렌즈
 "LL-6-13-001": "제3회 “스마트한 세상” 전시회에 오신 것을 진심으로 환영합니다. "
                "여러분들께서는 이번 전시회를 통해 과학기술이 우리 생활을 어떻게 변화시킬지 "
                "먼저 체험해 보실 수 있습니다. 먼저 앞쪽의 전자 제품 부스를 한번 볼까요? "
                "이 냉장고는 스스로 안에 들어 있는 식재료의 양과 종류를 파악합니다. "
                "그리고 언제 샀는지와 유통 기한을 자동으로 관리한답니다. "
                "그래서 사용자가 다 기억하지 않아도 재료가 곧 상한다거나 식재료가 떨어졌다는 "
                "알림을 보내 줍니다. 똑똑한 냉장고지요? 그 옆을 보십시오. "
                "이제 태블릿 피시를 더 이상 손에 들고 화면을 볼 필요가 없습니다. "
                "홀로그램으로 화면을 띄워서 볼 수 있고 가상 화면에서 글을 쓰고 그림을 그릴 수 있습니다. "
                "다음으로 이 스마트 렌즈를 볼까요? 이렇게 눈동자에 직접 착용하면 정보를 바로 "
                "볼 수 있답니다. 길 안내 화살표가 도로 위에 바로 보이고 눈을 깜빡여서 사진을 찍거나 "
                "인터넷 검색을 해서 정보를 바로 확인할 수도 있습니다. "
                "이 밖에도 이 전시회장 곳곳에서는 30여 가지의 첨단 기술을 더 보실 수 있습니다. "
                "그럼 여러분 천천히 구경하시면서 미래의 편리한 생활을 미리 만나 보시길 바랍니다. 감사합니다.",
 # 2급 14과 — 학당 직원이 선생님 번호를 알려 주는 장면 → 유리에게 직접 물어봄
 "LL-2-14-006": ("여보세요, 유리 씨지요? 저는 샤오밍이에요.", "샤오밍", "남"),
 "LL-2-14-007": ("안녕하세요, 샤오밍 씨. 무슨 일이에요?", "유리", "여"),
 "LL-2-14-008": ("내일 부모님이 한국에 오셔서 학교에 갈 수 없어요. "
                 "유리 씨, 김 선생님 전화번호를 알지요? 김 선생님 전화번호 좀 가르쳐 주세요.",
                 "샤오밍", "남"),
 "LL-2-14-009": ("네, 잠깐만요. 010-2049-5867이에요.", "유리", "여"),
 "LL-2-14-010": ("고마워요, 유리 씨.", "샤오밍", "남"),
 "LL-2-14-011": None,                    # 신판은 다섯 턴 — 남는 줄은 폐기
 # 5급 8과 — 불면증
 "LL-5-8-002": "‘불면증’은 잠을 잘 자지 못하는 증세입니다. 불면증이 있으면 늘 피곤하고 "
               "공부나 일도 잘 할 수 없지요. 불면증을 해결하기 위해서 약을 먹기도 하고 "
               "차를 마시기도 합니다. 그럼 지금부터 불면증에 좋은 몇 가지 차를 소개해 드리겠습니다. "
               "대추차는 긴장과 스트레스를 풀어주기 때문에 불면증에 좋습니다. "
               "솔잎차는 불면증에 도움이 되는 데다가 피도 깨끗하게 해 주고 위도 튼튼하게 해 줍니다. "
               "카모마일차와 꿀차도 불면증에 좋습니다.",
 # 5급 13과 — 코트 벗는 예절 → 식사 인사말
 "LL-5-13-001": "다른 사람의 집을 방문할 때는 방문해도 되는지 먼저 물어봐야 합니다. "
                "특별히 식사에 초대받은 것이 아니면, 식사 시간과 아침 일찍, 그리고 밤늦게 "
                "방문하는 것은 좋지 않습니다. 보통 약속한 시간보다 너무 일찍 도착하지 않는 것이 좋고 "
                "꽃이나 과자 같은 간단한 선물을 가지고 가는 것이 좋습니다. "
                "음식을 먹을 때는 ‘잘 먹겠습니다’, ‘잘 먹었습니다’ 같은 말을 해야 합니다. "
                "방문을 끝내고 나올 때도 ‘고맙습니다’, ‘폐를 끼쳤습니다’처럼 정중하게 인사를 합니다. "
                "하지만 너무 길게 인사를 해서 주인이 오래 밖에 서 있게 하는 것은 좋지 않습니다.",
 # 4급 3과 — 사전 → 책
 "LL-4-3-003": "그럼 그 책 좀 잠깐 봐도 돼요? 제가 책을 안 가지고 와서요.",
 "LL-4-3-004": "네, 저는 지금 안 보니까 마이클 씨 보세요. 여기 있어요.",
}

TARGET = {(6, "249"), (3, "127"), (6, "279"), (2, "84"),
          (5, "211"), (5, "228"), (4, "142")}

# ── 문항 손보기 (item_id → {필드: 값})
QUESTIONS = {
 "LC-6-4-001": {"question": "태블릿 PC를 잘 사용하고 있는 것을 고르십시오.",
                "selection4": "그냥 가방에 바로 넣는다."},
 "LC-3-13-002": {"selection1": "한국 과자"},
 "LC-3-13-004": {"question": "리핑 씨는 친구들에게 한국 과자를 선물할 겁니다."},
 # 신판에서 전자 잡지가 빠지고 스마트 렌즈가 들어왔다. 그대로 두면
 # '첨단 안경'과 '전자 잡지'가 둘 다 정답이 된다.
 "LC-6-13-002": {"selection2": "태블릿 PC", "selection3": "스마트 렌즈",
                 "selection4": "전자 잡지", "answer_index": 3},
 "LC-4-3-001": {"selection1": "마이클 씨는 책을 잃어버렸습니다.",
                "selection4": "유카 씨는 지금 책을 안 봐서 마이클 씨에게 빌려줬습니다."},
 "LC-4-3-002": {"question": "들은 내용과 같으면 O, 다르면 X 하십시오: 마이클 씨는 책을 오래 사용할 것입니다."},
 "LC-4-3-003": {"question": "들은 내용과 같으면 O, 다르면 X 하십시오: 마이클 씨는 책을 본 후에 유카 씨 책상 위에 둘 겁니다."},
}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/…]", "", t)


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    # ── 적어 넣은 문장이 신판 부록과 글자까지 같은지 먼저 대조
    blobs = {}
    bad = 0
    for iid, v in LINES.items():
        if v is None:
            continue
        b = int(iid.split("-")[1])
        if b not in blobs:
            gp = GlobalPdf(f"{APX}/(최종본)연세글로벌한국어_{b}급_부록-최종(26.8.10).pdf")
            blobs[b] = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))
        txt = v[0] if isinstance(v, tuple) else v
        if sq(txt) not in blobs[b]:
            print(f"  ✗ {iid} 신판 부록과 불일치")
            bad += 1
    if bad:
        raise SystemExit(f"{bad}건 불일치 — 중단")
    print(f"신판 부록 대조 {len(LINES)-1}건 전부 일치\n")

    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    ws = wb["n3_listen_script_line"]
    c_id, c_sp, c_g = col(ws, "item_id"), col(ws, "speaker"), col(ws, "gender")
    c_tx, c_rs, c_nt = col(ws, "text"), col(ws, "review_status"), col(ws, "change_note")
    n = 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, c_id).value
        if iid not in LINES:
            continue
        v = LINES[iid]
        if v is None:
            ws.cell(r, c_rs).value = "deleted"
            ws.cell(r, c_nt).value = "신판은 다섯 턴 구성 — 남는 줄 폐기"
        else:
            if isinstance(v, tuple):
                ws.cell(r, c_tx).value, ws.cell(r, c_sp).value = v[0], v[1]
                ws.cell(r, c_g).value = v[2]
            else:
                ws.cell(r, c_tx).value = v
            ws.cell(r, c_rs).value = "fixed_v15"
            ws.cell(r, c_nt).value = "신판 부록 듣기 지문으로 교체"
        n += 1
        log.append(dict(구분="듣기 대본 신판 반영", item_id=iid,
                        내용="폐기" if v is None else str(v[0] if isinstance(v, tuple) else v)[:60],
                        근거="신판 부록 듣기 지문"))
    print(f"대본 {n}줄 교체")

    wq = wb["n3_listen_repeat"]
    q_id = col(wq, "item_id")
    m = 0
    for r in range(2, wq.max_row + 1):
        iid = wq.cell(r, q_id).value
        if iid not in QUESTIONS:
            continue
        for k, val in QUESTIONS[iid].items():
            wq.cell(r, col(wq, k)).value = val
        wq.cell(r, col(wq, "review_status")).value = "fixed_v15"
        wq.cell(r, col(wq, "change_note")).value = "대본 교체에 맞춰 문항 조정"
        m += 1
        log.append(dict(구분="듣기 문항 조정", item_id=iid,
                        내용=", ".join(f"{k}={v}" for k, v in QUESTIONS[iid].items())[:90],
                        근거="교체한 대본과 맞춤"))
    print(f"문항 {m}개 조정")

    # ── audio_text 다시 만들기
    # n3_listen_script.audio_text는 줄들을 '화자: 말' 로 이어 붙인 것이다.
    # 줄만 고치고 두면 대본 원문이 옛것으로 남는다.
    # 고친 줄이 실제로 속한 대본만 다시 만든다. 과 단위로 잡으면 손대지 않은
    # 대본까지 형식이 바뀌어 '(잠시 후)' 같은 지시문과 문항 번호가 사라진다.
    ws2 = wb["n3_listen_script"]
    s_b, s_id = col(ws2, "book_id"), col(ws2, "id")
    s_at, s_rs = col(ws2, "audio_text"), col(ws2, "review_status")
    c_bk, c_ch, c_sid = col(ws, "book_id"), col(ws, "chapter"), col(ws, "script_id")
    live = {}
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, c_id).value
        if not iid or str(ws.cell(r, c_rs).value) == "deleted":
            continue
        key = (int(ws.cell(r, c_bk).value), str(ws.cell(r, c_sid).value))
        if key not in TARGET:
            continue
        sp = ws.cell(r, c_sp).value
        tx = ws.cell(r, c_tx).value
        live.setdefault(key, []).append(f"{sp}: {tx}" if sp else str(tx))
    k = 0
    for r in range(2, ws2.max_row + 1):
        key = (int(ws2.cell(r, s_b).value or 0), str(ws2.cell(r, s_id).value))
        if key not in live:
            continue
        ws2.cell(r, s_at).value = "\n".join(live[key])
        ws2.cell(r, s_rs).value = "fixed_v15"
        k += 1
    print(f"audio_text {k}개 대본 다시 생성")

    lg = wb.create_sheet("99_변경내역_v15")
    lg.append(["구분", "item_id", "내용", "근거"])
    for x in log:
        lg.append([x["구분"], x["item_id"], x["내용"], x["근거"]])
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거 CSV: verify/listen_diff.csv"])
    wb.save(DST)
    print(f"\n-> {DST}")


if __name__ == "__main__":
    main()
