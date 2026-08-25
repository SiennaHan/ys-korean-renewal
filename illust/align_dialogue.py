#!/usr/bin/env python3
"""엑셀 대화 행과 신판 추출 대화를 짝지어, 바뀐 턴의 신판 원문을 붙인다.

같은 과 안에서 turn_seq 순서는 유지되므로 difflib으로 순서 정렬한다.
화자가 같고 문장만 바뀐 경우가 대부분이라 정렬이 잘 붙는다.

산출: verify/dialogue_align.csv
      (item_id, 화자, 구판 ko, 신판 ko, 상태)
"""
import os, csv, re, difflib, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v3.xlsx"
OUT = f"{HERE}/verify"

# 문항 지시문·연습 예문은 대화가 아니다
INSTRUCTION = re.compile(r"하십시오|하세요|완성|고르|쓰십시오|다음을|보기|빈칸|알맞은")


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def read_sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    n2 = read_sheet(wb, "n2_ai_role_play")

    new = collections.defaultdict(list)
    for r in csv.DictReader(open(f"{OUT}/new_dialogue.csv")):
        if r["kind"] == "대화":
            new[(int(r["book"]), int(r["chapter"]))].append(r)

    stale = set()
    for b in range(2, 9):
        for r in csv.DictReader(open(f"{OUT}/n2_b{b}.csv")):
            if r["verdict"] == "구판 잔재":
                stale.add(r["item_id"])

    out = []
    for (b, ch), cand in sorted(new.items()):
        rows = [r for r in n2 if r.get("book_id") == b and r.get("chapter") == ch]
        rows.sort(key=lambda r: (str(r.get("scenario_id")), r.get("turn_seq") or 0))
        if not rows:
            continue
        # 문항 지시문·문법 연습 예문이 대화로 잡힌 것은 뺀다
        cand = [r for r in cand if not INSTRUCTION.search(r["ko"])]
        a = [sq(r.get("ko")) for r in rows]
        c = [sq(r["ko"]) for r in cand]
        sm = difflib.SequenceMatcher(a=a, b=c, autojunk=False)
        pair = {}
        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == "equal":
                for k in range(i2 - i1):
                    pair[i1 + k] = j1 + k
            elif tag == "replace":
                for k in range(min(i2 - i1, j2 - j1)):
                    pair[i1 + k] = j1 + k
        # 한 턴만 바뀐 경우는 opcode가 짝을 안 만들어 준다.
        # 앞뒤로 짝이 잡힌 턴 사이를 위치로 메운다.
        known = sorted(pair)
        for i in range(len(rows)):
            if i in pair:
                continue
            prev = max((k for k in known if k < i), default=None)
            nxt = min((k for k in known if k > i), default=None)
            if prev is not None and nxt is not None and (nxt - prev) == (pair[nxt] - pair[prev]):
                pair[i] = pair[prev] + (i - prev)
            elif prev is not None and pair[prev] + 1 < len(cand):
                pair[i] = pair[prev] + 1
        for i, r in enumerate(rows):
            iid = r.get("item_id")
            if iid not in stale:
                continue
            j = pair.get(i)
            nk = cand[j]["ko"] if j is not None else ""
            nsp = cand[j]["speaker"] if j is not None else ""
            out.append(dict(item_id=iid, book=b, chapter=ch,
                            scenario=r.get("scenario_id"), turn=r.get("turn_seq"),
                            speaker_old=r.get("speaker"), speaker_new=nsp,
                            ko_old=r.get("ko"), ko_new=nk,
                            status="신판 원문 확보" if nk else "짝 못 찾음"))

    with open(f"{OUT}/dialogue_align.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)
    c = collections.Counter(o["status"] for o in out)
    print(f"바뀐 턴 {len(out)}건: {dict(c)}")
    sp = sum(1 for o in out if o["speaker_new"] and o["speaker_old"] != o["speaker_new"])
    print(f"  화자까지 바뀐 것 {sp}건")
    print(f"-> {OUT}/dialogue_align.csv")


if __name__ == "__main__":
    main()
