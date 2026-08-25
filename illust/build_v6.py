#!/usr/bin/env python3
"""일괄 치환 사고로 깨진 한국어 11건을 교정해 v6을 만든다.

증상은 한 가지다. 한 글자를 다른 문자열로 바꾸면서 앞 글자와 공백을 먹었다.
  이용할 수 있어요  → 이이용할 있어요
  좋았는데 그 여자는 → 좋았는다른데 여자는
  뜻이 있는 곳에     → 뜻이 있하는에
  제가 지난번에      → 제여러난번에
  이용하는 곳이에요  → 이용하하는이에요

교재 개정과는 무관하다. 신판·구판 어디에도 없는 문자열이고, 고친 뒤에는
신판 본문에서 그대로 찾아진다. 그 확인을 통과한 것만 반영한다.

산출: 글로벌_교재기반_콘텐츠_v6.xlsx
"""
import os, csv, re, shutil, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v5.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v6.xlsx"

FIX = {
    "RP-4-4-005": [("이이용할 있", "이용할 수 있")],
    "RP-4-14-006": [("제여러난번에", "제가 지난번에")],
    "RP-4-15-010": [("많았는다른데렇게", "많았는데 그렇게")],
    "RP-5-3-001": [("이이용할 있", "이용할 수 있")],
    "RP-5-3-009": [("이이용할 있", "이용할 수 있")],
    "RP-5-3-010": [("이이용할 있", "이용할 수 있")],
    "RP-7-1-007": [("이용하하는이에요", "이용하는 곳이에요")],
    "RP-7-2-002": [("좋았는다른데 여자는", "좋았는데 그 여자는")],
    "RP-7-13-002": [("무슨 사진인다른데래요", "무슨 사진인데 그래요")],
    "RP-8-2-003": [("그런다른데 집에", "그런데 그 집에")],
    "RP-8-9-006": [("있하는에 길이", "있는 곳에 길이")],
}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i


def main():
    blobs = {}
    for b in (4, 5, 7, 8):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        blobs[b] = sq(" ".join(gp.text(p) for p in range(1, len(gp) + 1)))

    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n2_ai_role_play"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "chapter", "ko",
                                 "review_status", "change_note", "hold_reason")}

    log, ok, fail = [], 0, 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        if iid not in FIX:
            continue
        b = int(ws.cell(row, C["book_id"]).value)
        before = str(ws.cell(row, C["ko"]).value or "")
        after = before
        for old, new in FIX[iid]:
            after = after.replace(old, new)
        # 교정 결과가 신판 본문에 실제로 있는지 확인한 것만 반영
        if sq(after)[:26] in blobs[b]:
            ws.cell(row, C["ko"]).value = after
            ws.cell(row, C["change_note"]).value = (
                "일괄 치환 손상 교정 — 신판 본문 대조 확인")
            ws.cell(row, C["review_status"]).value = "reviewed"
            ok += 1
            log.append(["텍스트 손상 교정", iid, before[:60], after[:60], "본문 대조 통과"])
        else:
            ws.cell(row, C["hold_reason"]).value = (
                "일괄 치환 손상으로 보이나 교정안이 본문과 불일치 — 사람 확인 필요")
            fail += 1
            log.append(["텍스트 손상 보류", iid, before[:60], after[:60], "본문 대조 실패"])

    name = "99_변경내역_v6"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "일괄 치환 손상 교정", "", "",
               "교재 개정과 무관. 엑셀 데이터 자체가 깨져 있던 것"])
    sh.append(["증상", "한 글자 치환 시 앞 글자와 공백을 먹음", "이용할 수 있어요",
               "이이용할 있어요", "v71(3주완성) 단계부터 있던 손상으로 보임"])
    sh.append(["판별", "신판·구판 어디에도 없는 문자열", "", "",
               "교재가 바뀐 것이면 최소한 구판에는 있어야 한다"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((18, 16, 46, 46, 34), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"교정 {ok}건 / 보류 {fail}건")
    for r in log:
        print(f"  [{r[0][-2:]}] {r[1]}  {r[2][:44]}")
        print(f"        → {r[3][:44]}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
