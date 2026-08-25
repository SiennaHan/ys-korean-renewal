#!/usr/bin/env python3
"""글로벌_교재기반_콘텐츠 v3 을 만든다. v2를 읽어 결정 사항만 반영한다.

반영하는 것 (사용자 결정):
 1) 그림이 단어와 다른 4건 → image 비움 (그림 없이 간다)
 2) 신판에서 어휘가 바뀐 6건 → 신판 단어로 교정
 3) 그림만으로 못 맞히는 단어 → 단어 퀴즈(image-to-word) 문항 제거
 4) 검토를 통과한 어휘 177건 → n1_word_list 에 추가

추가 행의 번역: 교재에 인쇄된 영어 뜻만 넣는다(band_text에서 추출).
일본어·중국어는 교재에 있으나 해당 폰트가 cmap 없는 서브셋이라 추출 불가,
베트남어는 교재에 아예 없다. 지어내지 않고 빈 칸 + hold_reason 으로 남긴다.

산출: 글로벌_교재기반_콘텐츠_v3.xlsx  (+ 99_변경내역 시트)
"""
import os, csv, re, collections
import openpyxl
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v2_스키마완성.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v3.xlsx"

# ① 그림이 단어와 다름 → image 비움
DROP_IMAGE = [(2, 2, "강남"), (2, 2, "기숙사"), (2, 2, "신촌"), (3, 10, "덕수궁")]

# ② 신판에서 갈린 어휘 → 교정 (구 단어, 신 단어)
RENAME = [(6, 13, "전자책", "가상공간"),
          (8, 11, "의료용 로봇", "인공지능"),
          (8, 11, "정사각형 코드", "전기차"),
          (8, 11, "자동차 원격 조정", "항공택시"),
          (8, 11, "컴퓨터 원격 제어", "지능형 공장"),
          (8, 11, "휘는 디스플레이", "증강현실")]

# ③ 그림만으로 단어를 못 맞히는 것 → 단어 퀴즈에서 제거
QUIZ_DROP = ["WQ-3-1-001", "WQ-3-1-002", "WQ-5-14-001"]


def norm(s):
    return re.sub(r"\s+", "", str(s or "")).strip()


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def main():
    wb = openpyxl.load_workbook(SRC)
    log = []

    wl, wq = wb["n1_word_list"], wb["n1_word_quiz"]
    c_book, c_ch, c_word = col(wl, "book_id"), col(wl, "chapter"), col(wl, "word")
    c_img, c_note = col(wl, "image"), col(wl, "change_note")
    c_rs, c_hold = col(wl, "review_status"), col(wl, "hold_reason")
    c_item, c_page = col(wl, "item_id"), col(wl, "source_page")

    # ① image 비움
    for b, ch, w in DROP_IMAGE:
        for r in range(2, wl.max_row + 1):
            if (wl.cell(r, c_book).value == b and wl.cell(r, c_ch).value == ch
                    and norm(wl.cell(r, c_word).value) == norm(w)):
                old = wl.cell(r, c_img).value
                wl.cell(r, c_img).value = None
                wl.cell(r, c_note).value = "삽화 검수: 그림이 단어와 불일치 → 그림 없이 사용"
                log.append(["① 그림 삭제", f"{b}급 {ch}과", w, old, "(비움)",
                            "인물 초상화·놀이공원 등 단어와 무관한 그림"])

    # ② 신판 어휘로 교정
    for b, ch, old, new in RENAME:
        for ws, wcol in ((wl, c_word), (wq, None)):
            if ws is wl:
                for r in range(2, ws.max_row + 1):
                    if (ws.cell(r, c_book).value == b and ws.cell(r, c_ch).value == ch
                            and norm(ws.cell(r, wcol).value) == norm(old)):
                        ws.cell(r, wcol).value = new
                        ws.cell(r, c_note).value = f"신판 어휘 교체: {old} → {new}"
                        log.append(["② 어휘 교정", f"{b}급 {ch}과", old, old, new,
                                    "글로벌 신판에서 해당 자리 학습 어휘가 교체됨"])
            else:
                qb, qch = col(wq, "book_id"), col(wq, "chapter")
                for r in range(2, wq.max_row + 1):
                    if wq.cell(r, qb).value != b or wq.cell(r, qch).value != ch:
                        continue
                    for i in range(1, 5):
                        cc = col(wq, f"selection{i}")
                        if cc and norm(wq.cell(r, cc).value) == norm(old):
                            wq.cell(r, cc).value = new

    # ③ 단어 퀴즈 문항 제거
    qitem = col(wq, "item_id")
    kill = [r for r in range(2, wq.max_row + 1)
            if wq.cell(r, qitem).value in QUIZ_DROP]
    for r in sorted(kill, reverse=True):
        iid = wq.cell(r, qitem).value
        wq.delete_rows(r)
        log.append(["③ 퀴즈 제거", "", iid, "image-to-word 문항", "(삭제)",
                    "그림만으로 정답 단어를 특정할 수 없음 — 뜻 보고 고르기로 대체"])

    # ④ 검토 통과 어휘 추가
    adds = [r for r in csv.DictReader(open(f"{HERE}/vocab_final.csv"))
            if r["final"] == "추가"]
    exist = {(wl.cell(r, c_book).value, wl.cell(r, c_ch).value,
              norm(wl.cell(r, c_word).value)) for r in range(2, wl.max_row + 1)}
    c_en = col(wl, "en")
    seq = collections.Counter()
    for r in range(2, wl.max_row + 1):
        iid = str(wl.cell(r, c_item).value or "")
        m = re.match(r"VL-(\d+)-(\d+)-(\d+)$", iid)
        if m:
            seq[(int(m.group(1)), int(m.group(2)))] = max(
                seq[(int(m.group(1)), int(m.group(2)))], int(m.group(3)))

    added = 0
    for a in adds:
        b, ch, w = int(a["book"]), int(a["chapter"]), a["word"].strip()
        if (b, ch, norm(w)) in exist:
            continue
        seq[(b, ch)] += 1
        row = wl.max_row + 1
        wl.cell(row, c_book).value = b
        wl.cell(row, c_ch).value = ch
        wl.cell(row, c_word).value = w
        if a["en_in_book"]:
            wl.cell(row, c_en).value = a["en_in_book"]
        wl.cell(row, c_img).value = a["asset"]
        wl.cell(row, c_item).value = f"VL-{b}-{ch}-{seq[(b, ch)]:03d}"
        wl.cell(row, c_rs).value = "draft"
        wl.cell(row, c_page).value = a["pdf_page"]
        wl.cell(row, c_note).value = "삽화 라벨 대조로 신규 확인(글로벌 신판)"
        miss = [x for x, present in
                (("jp", 0), ("cn", 0), ("vi", 0)) if not present]
        wl.cell(row, c_hold).value = (
            "번역 미입력(jp·cn은 교재 폰트 추출 불가, vi는 교재에 없음)"
            + ("" if a["en_in_book"] else " · en도 미확인"))
        exist.add((b, ch, norm(w)))
        added += 1
        log.append(["④ 어휘 추가", f"{b}급 {ch}과", w,
                    "(없음)", w, f"교재 {a['pdf_page']}쪽 그림 라벨"])

    # 변경내역 시트
    if "99_변경내역" in wb.sheetnames:
        del wb["99_변경내역"]
    ws = wb.create_sheet("99_변경내역")
    ws.append(["구분", "위치", "대상", "이전", "이후", "근거"])
    for row in log:
        ws.append(row)
    for i, wid in enumerate([14, 12, 26, 26, 20, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = wid

    wb.save(DST)
    c = collections.Counter(r[0] for r in log)
    print("v3 반영:", dict(c))
    print(f"  어휘 추가 {added}건 (중복 제외)")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
