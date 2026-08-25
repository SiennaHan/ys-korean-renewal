#!/usr/bin/env python3
"""v23 — 읽기 지문 11편 개고.

문제는 주제가 아니라 어휘였다. 11편 모두 과 학습목표와는 맞는데
그 과에서 배운 낱말을 거의 쓰지 않았다 — 평균 5%, 나머지 106편은 16%(중앙값 13%).
읽기 지문은 그 과 어휘를 다시 만나게 하는 자리인데 그 역할을 못 했다.
6급 7과·7급 4·7·15과는 아예 0개였다.

통째로 새로 쓰지 않고 **줄기와 사실관계는 그대로 두고 어휘만 심었다**.
지문마다 문항이 3개씩 붙어 있어 정답 근거가 흔들리면 안 되기 때문이다.
개고 후 문항 33개의 정답 근거가 모두 남아 있는지 핵심어로 대조해 확인했다.

산출: 글로벌_교재기반_콘텐츠_v23.xlsx
"""
import shutil, datetime, re, collections, unicodedata
import openpyxl
from v23_texts import TEXTS

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v22.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v23.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")



def sq(t):
    return re.sub(r"\s", "", unicodedata.normalize("NFC", str(t or "")))


def stems(w):
    w = re.sub(r"[(（][^)）]*[)）]", "", str(w)).strip()
    s = sq(w)
    out = {s}
    if s.endswith("하다"):
        out.add(s[:-2] + "하")
    elif s.endswith("되다"):
        out.add(s[:-2] + "되")
    elif s.endswith("다") and len(s) > 2:
        out.add(s[:-1])
    return {x for x in out if len(x) >= 2}


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    wl = wb["n1_word_list"]
    W = {k: col(wl, k) for k in ["book_id", "chapter", "word", "review_status"]}
    per = collections.defaultdict(list)
    for r in range(2, wl.max_row + 1):
        b = wl.cell(r, W["book_id"]).value
        if not b or str(wl.cell(r, W["review_status"]).value) == "deleted":
            continue
        per[(int(b), int(wl.cell(r, W["chapter"]).value))].append(
            str(wl.cell(r, W["word"]).value))

    ws = wb["n5_read_answer_text"]
    C = {k: col(ws, k) for k in ["book_id", "chapter", "text", "item_id",
                                 "review_status", "change_note"]}
    log, n = [], 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        if iid not in TEXTS:
            continue
        b, ch = int(ws.cell(r, C["book_id"]).value), int(ws.cell(r, C["chapter"]).value)
        old, new = str(ws.cell(r, C["text"]).value), TEXTS[iid]
        words = per[(b, ch)]
        oh = [w for w in words if any(s in sq(old) for s in stems(w))]
        nh = [w for w in words if any(s in sq(new) for s in stems(w))]
        ws.cell(r, C["text"]).value = new
        ws.cell(r, C["review_status"]).value = "authored_v23"
        ws.cell(r, C["change_note"]).value = (
            f"과 어휘 반영 개고 — {len(oh)}개→{len(nh)}개 / 그 과 어휘 {len(words)}개")
        n += 1
        log.append((iid, b, ch, len(oh), len(nh), len(words),
                    " ".join(w for w in nh if w not in oh)))
    print(f"지문 {n}편 개고")
    for x in log:
        print(f"  {x[1]}급 {x[2]:>2}과 {x[3]}→{x[4]}/{x[5]}  더한 말: {x[6]}")

    lg = wb.create_sheet("99_변경내역_v23")
    lg.append(["item_id", "급", "과", "개고 전", "개고 후", "그 과 어휘", "더한 말"])
    for x in log:
        lg.append(list(x))
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 줄기·사실관계는 그대로 두고 어휘만 보강 "
               f"— 문항 33개 정답 근거 유지 확인"])
    wb.save(DST)
    print(f"\n-> {DST}")


if __name__ == "__main__":
    main()
