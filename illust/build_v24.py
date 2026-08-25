#!/usr/bin/env python3
"""v24 — 부록 '모범 답안'과 대조해 나온 n4 결함 2건 정정.

n4 완성문 836개를 교재 모범 답안(부록 4절, 8권 2,660문장)과 맞댔다.
  109건 교재 답과 일치 · 72건 거의 같음 · 586건 교재에 없는 창작
'거의 같음' 72건을 하나씩 보니 70건은 문체·조사·부사 차이거나 같은 문법의
다른 문제였다. 실질 결함은 아래 2건이다.

GF-4-11-007 — '커피를 많이 마시지 말고 물을 자주 드세요'로 높임이 섞였다.
   같은 상대에게 하는 말인데 앞은 안 높이고 뒤만 높였다. 게다가 마시다·드시다는
   같은 동작의 짝이라 더 어색하다. 교재 모범 답안은 '드시지 말고 … 드세요'다.
GF-8-11-004 — '보조키'는 우리 어휘 원장에도 없고 교재도 '보조 열쇠'라 쓴다.

산출: 글로벌_교재기반_콘텐츠_v24.xlsx
"""
import shutil, datetime
import openpyxl

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v23.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v24.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")


FIX = {
 "GF-4-11-007": dict(
    question="커피를 많이 ( ) 물을 자주 드세요.",
    selections="드시지 말고, 드셔서",
    selection1="드시지 말고", selection2="드셔서",
    answer="드시지 말고", answer_text="드시지 말고", answer_index=0,
    completion="커피를 많이 <b>드시지 말고</b> 물을 자주 드세요.",
    grammar_focus=("-지 말고: 앞 행동을 하지 말고 다른 행동을 권유할 때 쓴다. "
                   "'마시다'의 높임말 '드시다'에 붙여 뒤의 '드세요'와 높임을 맞춘다."),
    change_note="높임 섞임 정정 — 교재 모범 답안 '드시지 말고 … 드세요'에 맞춤"),
 "GF-8-11-004": dict(
    question="보조 열쇠를 (    ) 없었으면 차 문을 못 열 뻔했어요.",
    completion="보조 열쇠를 <b>갖고 있었기에 망정이지</b> 없었으면 차 문을 못 열 뻔했어요.",
    change_note="'보조키' → '보조 열쇠' — 교재 표기이며 '보조키'는 어휘 원장에 없음"),
}


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n4_blank_question"]
    hdr = [c.value for c in ws[1]]
    C = {k: hdr.index(k) + 1 for k in hdr if k}
    n = 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        if iid not in FIX:
            continue
        for k, v in FIX[iid].items():
            ws.cell(r, C[k]).value = v
        ws.cell(r, C["review_status"]).value = "fixed_v24"
        n += 1
    print(f"정정 {n}건")
    lg = wb.create_sheet("99_변경내역_v24")
    lg.append(["item_id", "내용", "근거"])
    for iid, v in FIX.items():
        lg.append([iid, v["change_note"], "부록 4. 모범 답안 대조"])
    lg.append(["검증 범위",
               "n4 완성문 836개 vs 교재 모범 답안 2,660문장 — 일치 109 / 거의같음 72 / 창작 586",
               ""])
    lg.append([f"작성 {datetime.date.today()}", "", ""])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
