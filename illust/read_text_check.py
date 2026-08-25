#!/usr/bin/env python3
"""읽기 지문(n5)이 신판에서도 통하는지 본다.

n5는 교재 원문을 그대로 옮긴 게 아니라 교재를 바탕으로 새로 쓴 지문이다
(1급 changelog에 '자체 창작 지문이라 교재 원문과 연동되지 않음'이라 적혀 있다).
그래서 문자열 대조로는 8급 15과 전부가 '교재 밖'으로 나와 아무 정보가 없다.

대신 주제 어휘가 겹치는지 본다. 지문에서 특징적인 낱말을 뽑아, 그 과의
신판 본문과 구판 본문에 각각 몇 개나 있는지 센다.
  구판에 많고 신판에 없다 → 원천 내용이 갈렸다. 지문을 다시 써야 한다.
  양쪽에 고루 있다        → 그대로 쓸 수 있다.

8급 3과가 실제 사례다. 구판 지문은 영화 '아름다운 비행' 소개였고
'늪·부화시키다·서먹하다'가 그 지문 어휘인데, 신판에는 하나도 없다.

산출: verify/read_text.csv
"""
import os, csv, re, json, collections, unicodedata
import fitz
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v23.xlsx"
OUT = f"{HERE}/verify"

# 아무 지문에나 나오는 말은 주제어가 아니다
STOP = set("""그리고 그러나 하지만 그래서 때문 이것 그것 저것 사람 우리 자신 경우 정도
때문에 통해 대해 위해 있다 없다 하다 되다 같다 많다 크다 좋다 이다 아니다 무엇 어떤
가지 생각 사실 문제 방법 시간 이런 저런 그런 여러 모든 다른 하나 둘 처음 마지막""".split())


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def keywords(text, n=28):
    """지문에서 특징적인 낱말 — 2자 이상 한국어 명사/용언 어간 후보."""
    words = re.findall(r"[가-힣]{2,}", str(text or ""))
    cnt = collections.Counter(w for w in words if w not in STOP and len(w) >= 2)
    # 자주 나오는 것보다 '이 지문에만 있을 법한' 긴 낱말을 우선
    return [w for w, _ in sorted(cnt.items(), key=lambda kv: (-len(kv[0]), -kv[1]))[:n]]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n5_read_answer_text"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    n5 = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    out = []
    for b in range(1, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        old = fitz.open(f"{BASE}/{b}_yonsei3week_main.pdf")
        toc = json.load(open(f"{BASE}/work/book{b}/toc.json"))
        ranges = {int(k): v for k, v in toc["ranges"].items()}
        for r in [x for x in n5 if x.get("book_id") == b]:
            ch = int(r["chapter"])
            if ch not in ranges:
                continue
            a, z = ranges[ch]
            nb = sq(" ".join(gp.text(p) for p in range(a, min(z, len(gp)) + 1)))
            ob = sq(" ".join(old[p].get_text()
                             for p in range(a - 1, min(z, old.page_count))))
            kw = keywords(r.get("text"))
            if len(kw) < 6:
                continue
            innew = sum(1 for w in kw if sq(w) in nb)
            inold = sum(1 for w in kw if sq(w) in ob)
            rn, ro = innew / len(kw), inold / len(kw)
            if ro >= 0.35 and rn < ro * 0.5:
                v = "원천 교체 — 지문 재저작 필요"
            elif rn < 0.2 and ro < 0.2:
                v = "양쪽과 무관 — 순수 창작"
            else:
                v = "유지"
            out.append(dict(item_id=r.get("item_id"), book=b, chapter=ch,
                            keywords=len(kw), in_new=innew, in_old=inold,
                            new_ratio=round(rn, 2), old_ratio=round(ro, 2),
                            verdict=v, head=str(r.get("text"))[:60]))

    with open(f"{OUT}/read_text.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)
    c = collections.Counter(o["verdict"] for o in out)
    print(f"읽기 지문 {len(out)}건: {dict(c)}")
    for o in out:
        if o["verdict"].startswith("원천 교체"):
            print(f"  {o['book']}급 {o['chapter']:>2}과 {o['item_id']}  "
                  f"신판 {o['in_new']}/{o['keywords']} vs 구판 {o['in_old']}/{o['keywords']}")
            print(f"      {o['head']}")
    print(f"-> {OUT}/read_text.csv")


if __name__ == "__main__":
    main()
