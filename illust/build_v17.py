#!/usr/bin/env python3
"""v17 — 부록 색인 대조 반영.

1) 8급 '추세' 2과 → 1과
   색인도 본문 뜻풀이도 1과라고 한다. 원장만 2과였다.

2) 어휘 63개 원장에 추가
   부록 색인과 본문 뜻풀이가 **같은 과로** 가리키는데 원장에 없던 낱말들이다.
   8급 1·2·3과에 50개가 몰려 있다 — 신판에서 본문을 다시 쓴 바로 그 과들이다.
   (8급 3과는 읽기 지문이 영화 <기생충> 감상문으로 바뀌면서 계층·정규직·흉기
    같은 어휘가 통째로 들어왔다.)

   6~8급 부록에는 번역 절이 없어 공식 번역이 없다. 그래서 en/jp/cn/vi는
   새로 저작한다(재저작 대화 번역과 같은 방침).

산출: 글로벌_교재기반_콘텐츠_v17.xlsx
"""
import shutil, datetime, collections
import openpyxl

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v16.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v17.xlsx"

# (급, 과, 낱말, 분류, en, jp, cn, vi)
NEW = [
 (6, 5, "신용 카드", "주제어휘", "credit card", "クレジットカード", "信用卡", "thẻ tín dụng"),
 (6, 13, "감정", "과제어휘", "emotion; feeling", "感情", "感情", "cảm xúc"),
 (6, 13, "뉘앙스", "과제어휘", "nuance", "ニュアンス", "语感；微妙的差别", "sắc thái"),
 (6, 13, "작곡", "주제어휘", "musical composition", "作曲", "作曲", "sự sáng tác nhạc"),
 (6, 13, "촬영하다", "주제어휘", "to film; to shoot", "撮影する", "拍摄", "quay phim, chụp ảnh"),
 (7, 7, "이변", "과제어휘", "an unexpected turn; upset", "異変", "意外事件；爆冷", "sự biến bất ngờ"),
 (7, 9, "뒤떨어지다", "새어휘", "to fall behind; to lag", "遅れる、劣る", "落后", "tụt hậu, kém hơn"),
 (7, 10, "한바탕", "과제어휘", "a bout of; a round of", "ひとしきり", "一场；一阵", "một trận, một hồi"),
 (7, 12, "거북해하다", "과제어휘", "to feel uncomfortable", "気まずく思う", "感到不自在", "cảm thấy khó chịu"),
 # 8급 1과 — 성역할
 (8, 1, "강조하다", "주제어휘", "to emphasize", "強調する", "强调", "nhấn mạnh"),
 (8, 1, "거부하다", "주제어휘", "to refuse; to reject", "拒否する", "拒绝", "từ chối, khước từ"),
 (8, 1, "경계", "주제어휘", "boundary; border", "境界", "界限；分界", "ranh giới"),
 (8, 1, "구애받다", "과제어휘", "to be bound by; to be constrained", "とらわれる", "受拘束；拘泥于", "bị ràng buộc bởi"),
 (8, 1, "기준", "주제어휘", "standard; criterion", "基準", "标准", "tiêu chuẩn"),
 (8, 1, "넘나들다", "주제어휘", "to cross back and forth", "行き来する、越境する", "来回穿越；跨越", "đi lại qua lại, vượt qua"),
 (8, 1, "배제하다", "주제어휘", "to exclude", "排除する", "排除", "loại trừ"),
 (8, 1, "생계", "새어휘", "livelihood", "生計", "生计", "kế sinh nhai"),
 (8, 1, "선보이다", "주제어휘", "to unveil; to present", "披露する", "展示；亮相", "ra mắt, giới thiệu"),
 (8, 1, "섬세하다", "새어휘", "delicate; sensitive", "繊細だ", "细腻；细致", "tinh tế, tỉ mỉ"),
 (8, 1, "성별화", "주제어휘", "gendering; gender division", "性別化", "性别化", "sự phân hóa theo giới tính"),
 (8, 1, "얽매이다", "새어휘", "to be tied down; to be bound", "縛られる", "被束缚", "bị trói buộc"),
 (8, 1, "온화하다", "새어휘", "gentle; mild", "温和だ", "温和", "ôn hòa, hiền hòa"),
 (8, 1, "전담하다", "새어휘", "to take exclusive charge of", "専ら担当する", "专门负责", "chuyên trách, đảm nhiệm riêng"),
 (8, 1, "특성", "주제어휘", "characteristic; trait", "特性", "特性", "đặc tính"),
 (8, 1, "판단", "주제어휘", "judgment", "判断", "判断", "sự phán đoán"),
 (8, 1, "활약하다", "새어휘", "to be active; to play an active role", "活躍する", "活跃；大显身手", "hoạt động tích cực"),
 # 8급 2과 — 가족
 (8, 2, "밀집", "과제어휘", "dense concentration", "密集", "密集", "sự tập trung dày đặc"),
 (8, 2, "수준", "과제어휘", "level; standard", "水準", "水平", "trình độ, mức độ"),
 (8, 2, "아낌없이", "과제어휘", "without sparing; generously", "惜しみなく", "毫不吝惜地", "một cách không tiếc"),
 (8, 2, "이방인", "과제어휘", "stranger; outsider", "異邦人", "异乡人；外人", "người xa lạ"),
 (8, 2, "일원", "과제어휘", "a member (of a group)", "一員", "一员", "một thành viên"),
 (8, 2, "정착하다", "과제어휘", "to settle down", "定着する", "定居", "định cư"),
 (8, 2, "특색 있다", "과제어휘", "to be distinctive", "特色がある", "有特色", "có nét đặc sắc"),
 (8, 2, "해결하다", "과제어휘", "to solve; to resolve", "解決する", "解决", "giải quyết"),
 # 8급 3과 — 대중문화 (읽기 지문: 영화 <기생충> 감상문)
 (8, 3, "계층", "주제어휘", "social class; stratum", "階層", "阶层", "tầng lớp"),
 (8, 3, "구조", "주제어휘", "structure", "構造", "结构", "cấu trúc"),
 (8, 3, "국한되다", "주제어휘", "to be limited to", "限定される", "局限于", "bị giới hạn trong"),
 (8, 3, "난동", "주제어휘", "violent disturbance; rampage", "乱暴騒ぎ", "骚乱；行凶", "vụ náo loạn"),
 (8, 3, "대변하다", "주제어휘", "to speak for; to represent", "代弁する", "代言；代表", "phát ngôn thay, đại diện cho"),
 (8, 3, "대응하다", "주제어휘", "to respond; to counter", "対応する", "应对", "ứng phó, đối phó"),
 (8, 3, "벌이다", "주제어휘", "to carry out; to stir up", "繰り広げる、起こす", "展开；闹出", "gây ra, tiến hành"),
 (8, 3, "보편화", "주제어휘", "becoming widespread", "普遍化", "普及化", "sự phổ biến hóa"),
 (8, 3, "본격화", "주제어휘", "full-scale start", "本格化", "全面展开", "sự bước vào giai đoạn chính thức"),
 (8, 3, "불평등하다", "과제어휘", "to be unequal", "不平等だ", "不平等", "bất bình đẳng"),
 (8, 3, "양상", "주제어휘", "aspect; pattern", "様相", "面貌；情形", "diện mạo, hình thái"),
 (8, 3, "연이어", "주제어휘", "one after another", "相次いで", "接连", "liên tiếp"),
 (8, 3, "위상", "주제어휘", "status; standing", "地位、位相", "地位", "vị thế"),
 (8, 3, "잇따라", "주제어휘", "successively; in succession", "相次いで", "接连不断", "nối tiếp nhau"),
 (8, 3, "정규직", "주제어휘", "regular (permanent) position", "正規職", "正式员工", "vị trí chính thức"),
 (8, 3, "제시하다", "주제어휘", "to present; to put forward", "提示する", "提出", "đưa ra, trình bày"),
 (8, 3, "지향하다", "주제어휘", "to aim for; to pursue", "志向する", "追求；以…为方向", "hướng tới"),
 (8, 3, "진출하다", "주제어휘", "to advance into; to make inroads", "進出する", "进军", "tiến vào, thâm nhập"),
 (8, 3, "처지", "주제어휘", "situation; circumstances", "境遇、立場", "处境", "hoàn cảnh"),
 (8, 3, "탈출하다", "주제어휘", "to escape", "脱出する", "逃脱", "trốn thoát"),
 (8, 3, "폭발적", "주제어휘", "explosive", "爆発的", "爆发性的", "mang tính bùng nổ"),
 (8, 3, "풍자적", "주제어휘", "satirical", "風刺的", "讽刺的", "mang tính châm biếm"),
 (8, 3, "협박하다", "주제어휘", "to threaten; to blackmail", "脅迫する", "威胁；恐吓", "đe dọa"),
 (8, 3, "흉기", "주제어휘", "deadly weapon", "凶器", "凶器", "hung khí"),
 (8, 3, "흥행하다", "주제어휘", "to be a box-office hit", "ヒットする", "卖座；票房成功", "ăn khách"),
 (8, 5, "십 년이면 강산도 변한다", "새어휘",
  "Ten years can change even the mountains and rivers — much changes over time",
  "十年経てば山河も変わる", "十年江山都会变；十年一大变",
  "Mười năm thì núi sông cũng đổi thay"),
 (8, 6, "물려주다", "새어휘", "to hand down; to pass on", "譲り渡す、受け継がせる", "传给；留给", "truyền lại, để lại"),
 (8, 6, "열정", "새어휘", "passion", "情熱", "热情", "nhiệt huyết, đam mê"),
 (8, 7, "배보다 배꼽이 더 크다", "새어휘",
  "The navel is bigger than the belly — the side cost outweighs the main one",
  "本末転倒だ", "本末倒置；喧宾夺主", "Rốn to hơn bụng — cái phụ lớn hơn cái chính"),
]


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n1_word_list"]
    C = {k: col(ws, k) for k in ["id", "book_id", "chapter", "word", "en", "jp", "cn",
                                 "vi", "category", "item_id", "review_status",
                                 "change_note"]}

    # ── 1) 8급 추세 2과 → 1과
    moved = 0
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, C["book_id"]).value == 8 and ws.cell(r, C["word"]).value == "추세"
                and ws.cell(r, C["chapter"]).value == 2):
            ws.cell(r, C["chapter"]).value = 1
            ws.cell(r, C["review_status"]).value = "fixed_v17"
            ws.cell(r, C["change_note"]).value = "부록 색인·본문 뜻풀이 모두 1과 — 2과에서 옮김"
            moved += 1
    print(f"1) 추세 과 이동 {moved}건")

    # ── 2) 어휘 추가
    max_id = 0
    seq = collections.Counter()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, C["id"]).value
        if isinstance(v, (int, float)):
            max_id = max(max_id, int(v))
        iid = ws.cell(r, C["item_id"]).value
        if iid and str(iid).startswith("VL-"):
            p = str(iid).split("-")
            if len(p) == 4 and p[3].isdigit():
                k = (int(p[1]), int(p[2]))
                seq[k] = max(seq[k], int(p[3]))

    row = ws.max_row + 1
    for b, ch, w, cat, en, jp, cn, vi in NEW:
        max_id += 1
        seq[(b, ch)] += 1
        vals = {"id": max_id, "book_id": b, "chapter": ch, "word": w, "en": en,
                "jp": jp, "cn": cn, "vi": vi, "category": cat,
                "item_id": f"VL-{b}-{ch}-{seq[(b, ch)]:03d}",
                "review_status": "added_v17",
                "change_note": "부록 색인·본문 뜻풀이에 있으나 원장에 없어 추가"}
        for k, v in vals.items():
            ws.cell(row, C[k]).value = v
        row += 1
    print(f"2) 어휘 추가 {len(NEW)}건")

    lg = wb.create_sheet("99_변경내역_v17")
    lg.append(["구분", "급", "과", "낱말", "분류", "내용", "근거"])
    lg.append(["과 배정 수정", 8, "2→1", "추세", "", "2과에서 1과로",
               "부록 색인·본문 뜻풀이 모두 1과"])
    for b, ch, w, cat, en, *_ in NEW:
        lg.append(["어휘 추가", b, ch, w, cat, en, "부록 색인 + 본문 뜻풀이"])
    lg.append([])
    lg.append([f"작성 {datetime.date.today()} · 근거 CSV: verify/index_vocab.csv, "
               f"verify/index_check.csv"])
    wb.save(DST)
    print(f"\n-> {DST}")


if __name__ == "__main__":
    main()
