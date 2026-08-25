#!/usr/bin/env python3
"""n4 문항에서 '문법 항목' 목록을 뽑는다.

신호가 급마다 다르다.
  2~3급 — grammar_focus가 '많다: 받침O(ㄶ) → -습니다'처럼 설명문이라 못 쓴다.
          대신 보기·정답이 깨끗한 형태소다(-습니다, -ㅂ니다).
  4~8급 — grammar_focus 앞머리가 문법형 그대로다(-어지다/아지다/여지다).

그래서 focus가 문법형처럼 보이면 그걸 쓰고, 아니면 정답을 이형태 통일해 쓴다.
이형태는 한 항목으로 묶고(태그 원칙 1), 표면형이 같아도 교재가 다른 과에서
다른 기능으로 가르치면 나눈다(원칙 2) — 그래서 과를 열쇠에 함께 넣되,
같은 급 안에서 같은 형태가 이어지는 과에 나오면 처음 과로 합친다.

산출: verify/n4_inventory.csv
"""
import re, csv, os, collections, unicodedata
import openpyxl
from n4_group import norm_answer

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v19.xlsx"

# 표면형이 같아도 기능이 갈리는 것들 — 과별로 나눈다(태그 원칙 2)
HOMOGRAPH = {"에", "에서", "으로/로", "도", "만", "은/는", "이/가", "을/를", "과/와", "까지", "부터"}


def focus_form(s):
    """grammar_focus에서 문법형만 떼어 낸다.

    뒤에 설명이 붙는 꼴이 많다 —
      '-을/ㄹ 뿐이다. 과거 보다→봤을'  '-어찌나 -는지. 형용사 어렵다→어려운지'
    설명까지 열쇠에 넣으면 같은 문법이 문항 수만큼 쪼개진다(실제로 366개가 됐다).
    ':' '.' '→' 앞에서 끊고, 그래도 문법형이 아니면 원문에서 '-…' 조각을 찾는다.
    """
    raw = unicodedata.normalize("NFC", str(s or "")).strip()
    t = re.split(r"[:.．]|→", raw)[0].strip()
    t = re.sub(r"\(.*?\)", "", t).strip()
    t = re.sub(r"\s+", " ", t)
    if is_grammar_form(t):
        return t
    # '동사 \'자다\'+는 편이다' 처럼 앞에 품사 설명이 붙은 꼴
    m = re.search(r"(-[가-힣ㄱ-ㅎ][가-힣ㄱ-ㅎ/ ]{0,20})", raw)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return t


def is_grammar_form(t):
    return bool(t) and (t.startswith("-") or "/" in t or "…" in t)


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def key_of(r):
    """급마다 믿을 만한 신호가 다르다.

    2~3급 grammar_focus는 '많다: 받침O(ㄶ) → -습니다'처럼 어간 설명이라
    거기서 문법형을 캐면 오히려 잘게 쪼개진다(42→61). 이 구간은 보기·정답이
    이미 깨끗한 형태소이므로 정답을 쓴다.
    4~8급은 focus 앞머리가 문법형 그대로라 그쪽이 정확하다.
    """
    if int(r["book_id"]) <= 3:
        return norm_answer(r.get("answer"))
    f = focus_form(r.get("grammar_focus"))
    return f if is_grammar_form(f) else norm_answer(r.get("answer"))


def distractor_type(r):
    """보기의 성격 — 이형태 / 대조 / 오형태.

    이형태: 보기가 같은 문법의 변이형(정답과 이형태 통일형이 같다)
    오형태: 보기 중에 문법적으로 틀린 꼴이 있다(규칙 미적용형)
    대조  : 보기가 서로 다른 문법이다
    """
    sels = [str(r.get(f"selection{i}") or "").strip() for i in range(1, 5)]
    sels = [s for s in sels if s]
    if len(sels) < 2:
        return ""
    base = {norm_answer(s) for s in sels}
    if len(base) == 1:
        return "이형태"
    # 품사로 갈리는 이형태 — 동사 -는데 / 형용사 -은데 처럼 앞머리만 다른 짝.
    # norm_answer는 받침 이형태만 통일하므로 이쪽은 뼈대로 한 번 더 본다.
    # 다만 뼈대가 비면(‘-는’·‘-은’처럼 표지 자체가 항목인 경우) 쓰지 않는다 —
    # 교재가 -는/-은/-을을 각각 다른 문법으로 가르치는 과가 있다(2급 11과).
    from n4_assign import skeleton
    sk = {skeleton(s) for s in sels}
    if len(sk) == 1 and all(sk):
        return "이형태"
    # 오형태: 같은 말을 잘못 활용한 꼴이 섞인 경우(덥어요/더워요, 않아요/안아요).
    # 두 보기가 문자열로 아주 비슷하면 서로 다른 문법일 수 없다 — 한쪽이 틀린 꼴이다.
    # 문자열 유사도는 쓰지 않는다. 1급 기록값을 보면 오형태가 0.33까지 내려가고
    # (덥어요/더워요) 대조가 0.75까지 올라간다(-지요?/-에요?). 잣대가 되지 못한다.
    # 저작자가 focus에 '…는 비문'이라 적어 둔 자리는 오답이 틀린 꼴이라는 뜻이다.
    # (‘쓰어’·‘갈 봤어요’처럼 유사도로는 안 잡히는 것들이 여기 걸린다.)
    g = str(r.get("grammar_focus") or "")
    # 저작자가 틀린 꼴을 지목해 둔 표현들 —
    #   '…는 비문', '비표준', 'ㅂ불규칙', "'안'이 아니라 '않'으로 씀"
    if ("비문" in g or "비표준" in g or "불규칙" in g
            or re.search(r"[이가] 아니라", g)):
        return "오형태"
    return "대조"


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n4_blank_question") if r.get("book_id")]
    un = [r for r in rows if not r.get("grammar_tag")]

    # (급, 형태) → 나오는 과들
    seen = collections.defaultdict(list)
    for r in un:
        seen[(int(r["book_id"]), key_of(r))].append(int(r["chapter"]))

    out = []
    for (b, k), chs in sorted(seen.items()):
        chs = sorted(set(chs))
        if k in HOMOGRAPH:
            for c in chs:                      # 동형이의 — 과별로 나눈다
                n = sum(1 for r in un if int(r["book_id"]) == b
                        and key_of(r) == k and int(r["chapter"]) == c)
                out.append(dict(book=b, form=k, 도입_과=c, 과들=str(c), 문항수=n))
        else:
            n = sum(1 for r in un if int(r["book_id"]) == b and key_of(r) == k)
            out.append(dict(book=b, form=k, 도입_과=chs[0],
                            과들="/".join(map(str, chs)), 문항수=n))

    with open(f"{HERE}/verify/n4_inventory.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["book", "form", "도입_과", "과들", "문항수"])
        wr.writeheader()
        wr.writerows(sorted(out, key=lambda x: (x["book"], x["도입_과"], x["form"])))
    print(f"문법 항목 {len(out)}개 (문항 {len(un)})")
    print("급별:", dict(collections.Counter(o["book"] for o in out)))
    print("\ndistractor_type 자동판정:",
          dict(collections.Counter(distractor_type(r) for r in un)))
    print(f"\n-> verify/n4_inventory.csv")


if __name__ == "__main__":
    main()
