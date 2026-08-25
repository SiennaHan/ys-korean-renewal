#!/usr/bin/env python3
"""엑셀 내용(n1 어휘·n2 대화)을 신판 본문에 실제로 있는지 대조한다.

쪽 단위 전체 diff는 레이아웃 재배치('어휘'/'문법' 사이드바 순서, 줄바꿈 위치)에
너무 민감해서 실제 변경과 구분이 안 됐다. 그래서 방향을 뒤집는다:
엑셀의 문자열을 신판·구판 본문에서 각각 찾아 4가지로 가른다.

  신O 구O → 유지        (양쪽에 다 있음)
  신O 구X → 신판 신설    (원장에 이미 반영돼 있던 것)
  신X 구O → 구판 잔재    ← 고쳐야 하는 것
  신X 구X → 교재 밖      (자체 창작·파생 데이터)

비교는 공백을 모두 제거한 문자열 포함 검사다. 신판은 공백 글리프 없이
좌표로 자간을 잡아 추출되므로 공백을 남기면 전부 불일치가 된다.

사용법: python verify_content.py [급 ...]
산출:  verify/n1_b{급}.csv, verify/n2_b{급}.csv, verify/summary.csv
"""
import os, sys, csv, re, json, collections, unicodedata
import fitz
import openpyxl
from global_text import GlobalPdf, _is_broken as _broken

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v12.xlsx"
OUT = f"{HERE}/verify"


def squash(text):
    """비교용: 공백·구두점 제거 + 원문자 통일."""
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"[➊-➓❶-❿①-⑳]", "", text)
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", text)


def page_text(doc, pno, gp=None):
    page = doc[pno - 1]
    parts = []
    for sp in page.get_texttrace():
        chars = sp["chars"]
        if not chars:
            continue
        if gp is not None:
            dec = gp._decoder(sp["font"])
            parts.append("".join(
                (dec(c[1]) or chr(c[0])) if (dec and _broken(chr(c[0]))) else chr(c[0])
                for c in chars))
        else:
            parts.append("".join(chr(c[0]) for c in chars))
    return " ".join(parts)


def blobs(b):
    """(급) -> {과: 신판 본문 덩어리}, {과: 구판 본문 덩어리}, 급 전체 덩어리"""
    gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
    old = fitz.open(f"{BASE}/{b}_yonsei3week_main.pdf")
    toc = json.load(open(f"{BASE}/work/book{b}/toc.json"))
    ranges = {int(k): v for k, v in toc["ranges"].items()}
    new_ch, old_ch = {}, {}
    for ch, (a, z) in ranges.items():
        z = min(z, len(gp), old.page_count)
        new_ch[ch] = squash(" ".join(page_text(gp.doc, p, gp) for p in range(a, z + 1)))
        old_ch[ch] = squash(" ".join(page_text(old, p) for p in range(a, z + 1)))
    return new_ch, old_ch


def read_sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def verdict(in_new, in_old):
    if in_new and in_old:
        return "유지"
    if in_new and not in_old:
        return "신판 신설"
    if in_old and not in_new:
        return "구판 잔재"
    return "교재 밖"


def where(blob_map, needle, exclude=None):
    """다른 과에 있는지 — 과 이동 검출."""
    return [ch for ch, t in blob_map.items() if ch != exclude and needle in t]


def run(b, wb):
    new_ch, old_ch = blobs(b)
    os.makedirs(OUT, exist_ok=True)
    summary = []

    # ── n1 어휘
    # 폐기 표시한 행은 이미 처리가 끝난 것이라 대조 대상이 아니다
    rows = [r for r in read_sheet(wb, "n1_word_list")
            if r.get("book_id") == b and str(r.get("review_status")) != "deleted"]
    out = []
    for r in rows:
        ch, w = int(r["chapter"]), str(r.get("word") or "")
        key = squash(w)
        if not key:
            continue
        inn, ino = key in new_ch.get(ch, ""), key in old_ch.get(ch, "")
        v = verdict(inn, ino)
        moved = "" if inn else " ".join(f"{c}과" for c in where(new_ch, key, ch))
        out.append(dict(item_id=r.get("item_id"), book=b, chapter=ch, word=w,
                        verdict=v, moved_to=moved,
                        note="신판 다른 과에 있음 → 과 이동" if moved else ""))
    with open(f"{OUT}/n1_b{b}.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)
    c1 = collections.Counter(o["verdict"] for o in out)
    mv = sum(1 for o in out if o["moved_to"])
    summary.append(dict(book=b, sheet="n1_word_list", total=len(out),
                        **{k: c1.get(k, 0) for k in
                           ("유지", "신판 신설", "구판 잔재", "교재 밖")},
                        moved=mv))

    # ── n2 대화 (턴 단위)
    rows = [r for r in read_sheet(wb, "n2_ai_role_play")
            if r.get("book_id") == b and str(r.get("review_status")) != "deleted"]
    out = []
    for r in rows:
        ch, ko = int(r["chapter"]), str(r.get("ko") or "")
        key = squash(ko)
        if len(key) < 4:
            continue
        inn, ino = key in new_ch.get(ch, ""), key in old_ch.get(ch, "")
        out.append(dict(item_id=r.get("item_id"), book=b, chapter=ch,
                        scenario=r.get("scenario_id"), turn=r.get("turn_seq"),
                        speaker=r.get("speaker"), ko=ko,
                        verdict=verdict(inn, ino)))
    if out:
        with open(f"{OUT}/n2_b{b}.csv", "w", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=list(out[0]))
            wr.writeheader()
            wr.writerows(out)
        c2 = collections.Counter(o["verdict"] for o in out)
        summary.append(dict(book=b, sheet="n2_ai_role_play", total=len(out),
                            **{k: c2.get(k, 0) for k in
                               ("유지", "신판 신설", "구판 잔재", "교재 밖")}, moved=0))

    for s in summary:
        print(f"  [{b}급 {s['sheet']}] {s['total']:>4}행 | 유지 {s['유지']:>4} | "
              f"구판 잔재 {s['구판 잔재']:>3} | 교재 밖 {s['교재 밖']:>3} | "
              f"신판 신설 {s['신판 신설']:>3}" + (f" | 과이동 {s['moved']}" if s['moved'] else ""))
    return summary


if __name__ == "__main__":
    books = [int(x) for x in sys.argv[1:]] or list(range(2, 9))
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    allsum = []
    for b in books:
        allsum += run(b, wb)
    with open(f"{OUT}/summary.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(allsum[0]))
        wr.writeheader()
        wr.writerows(allsum)
    print(f"\n-> {OUT}/")
