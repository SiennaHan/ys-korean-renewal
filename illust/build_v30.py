#!/usr/bin/env python3
"""v30 — 자모 조합 15행을 write3 로 고친다. 검수에서 나온 내 오류다.

무엇이 틀렸나
    build_v26.py 는 activity_sub 를 "묶음 안 모듈 위치" 로 정했다(1~5).
    그런데 자모 활동은 여섯이다 — BLOCKERS.md §2 가 sub6 = write3
    (/learn/jamo/combine3) 를 따로 적어 두었다. 위치만 보면 6번째가 없어서
    받침·겹받침의 3단 조합이 2단 조합(write)으로 들어갔다.

어떻게 갈렸나 — write 46행을 답의 모양으로 나누면 셋이다
    ① 답 2개 (초성+중성)              24행  YK0007·0012·0017·0022·0027  → write
    ② 답 3개, answer_3 = 조합 결과      7행  YK0002                      → write
    ③ 답 3개, answer_3 = 종성          15행  YK0032(7) · YK0037(8)       → write3
    설명 안 되는 행 0개.

    ③ 은 초성+중성+종성을 조합한다. Y3W479 (ㅂ,ㅏ,ㄲ) → 밖.
    한글 조합으로 실제로 맞춰 확인했다 — compose(a1,a2,a3) == target.
    BLOCKERS.md §2 가 YK0032 를 combine3 라 적은 것과 같다.

    ② 는 답이 셋이지만 종성이 아니라 결과다(ㅇ,ㅓ,어). 2단 조합이므로 write 로 둔다.

찾은 방법
    조합 활동의 answer_1+answer_2 가 answer_3 와 맞는지 기계로 셌다.
    22행이 안 맞았고, 그중 15행이 실제로는 3단이었다.
    (남은 7행은 ② 이고 내 검사가 결과를 종성으로 오해한 것이었다.)

돌리기
    python3 illust/build_v30.py

대상이 이미 있으면 아무것도 하지 않고 멈춘다.
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "글로벌_교재기반_콘텐츠_v29.xlsx"
OUT = ROOT / "글로벌_교재기반_콘텐츠_v30.xlsx"
SHEET = "n8_jamo"

# 3단 조합 모듈. 위치로는 2번(조합)이지만 종성까지 조합한다.
WRITE3_MODULES = {"YK0032", "YK0037"}

CHO = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
JUNG = "ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ"
JONG = ["", "ㄱ", "ㄲ", "ㄳ", "ㄴ", "ㄵ", "ㄶ", "ㄷ", "ㄹ", "ㄺ", "ㄻ", "ㄼ",
        "ㄽ", "ㄾ", "ㄿ", "ㅀ", "ㅁ", "ㅂ", "ㅄ", "ㅅ", "ㅆ", "ㅇ", "ㅈ",
        "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"]


def compose(c: str, v: str, j: str = "") -> str | None:
    """초성·중성·종성을 한 음절로. 호환 자모를 받는다."""
    if c not in CHO or v not in JUNG or j not in JONG:
        return None
    return chr(0xAC00 + CHO.index(c) * 588 + JUNG.index(v) * 28 + JONG.index(j))


def die(msg: str) -> None:
    print(f"멈춤 — {msg}", file=sys.stderr)
    raise SystemExit(1)


def main() -> int:
    if OUT.exists():
        die(f"{OUT.name} 이 이미 있다. 지우거나 다른 번호를 써라.")
    if not SRC.exists():
        die(f"{SRC.name} 이 없다")

    shutil.copy(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}
    for need in ("activity_sub", "module_code", "answer_1", "answer_2", "answer_3",
                 "target_jamo", "target_word", "legacy_id", "change_note"):
        if need not in col:
            die(f"{SHEET} 에 {need} 열이 없다")

    changed: list[str] = []
    for r in range(2, ws.max_row + 1):
        mod = str(ws.cell(row=r, column=col["module_code"]).value or "").strip()
        if mod not in WRITE3_MODULES:
            continue
        sub = str(ws.cell(row=r, column=col["activity_sub"]).value or "").strip()
        a1 = str(ws.cell(row=r, column=col["answer_1"]).value or "").strip()
        a2 = str(ws.cell(row=r, column=col["answer_2"]).value or "").strip()
        a3 = str(ws.cell(row=r, column=col["answer_3"]).value or "").strip()
        tgt = (str(ws.cell(row=r, column=col["target_jamo"]).value or "").strip()
               or str(ws.cell(row=r, column=col["target_word"]).value or "").strip())
        lid = str(ws.cell(row=r, column=col["legacy_id"]).value or "").strip()
        # 3단 조합이 실제로 대상 음절을 만드는지 확인한 뒤에만 바꾼다
        if compose(a1, a2, a3) != tgt:
            die(f"{lid} 이 3단 조합으로 대상을 못 만든다 ({a1},{a2},{a3}) → {tgt!r}")
        if sub == "write3":
            continue
        ws.cell(row=r, column=col["activity_sub"]).value = "write3"
        note = str(ws.cell(row=r, column=col["change_note"]).value or "").strip()
        ws.cell(row=r, column=col["change_note"]).value = (
            note + " / v29: 3단 조합이라 write → write3 (sub6)"
        ).strip(" /")
        changed.append(lid)

    if not changed:
        die("고칠 행이 없다 — 이미 write3 인가?")

    log = wb.create_sheet("99_변경내역_v30")
    log.append(["구분", "내용"])
    log.append(["활동 정정", f"n8_jamo {len(changed)}행을 write → write3 (sub6). "
                          f"모듈 {' · '.join(sorted(WRITE3_MODULES))}"])
    log.append(["왜 틀렸나", "build_v26.py 가 activity_sub 를 모듈 위치(1~5)로만 정했다. "
                          "자모 활동은 여섯이고 sub6=write3 이 위치로는 표현되지 않는다"])
    log.append(["어떻게 찾았나", "조합 활동의 answer_1+answer_2 가 answer_3 와 맞는지 세었다. "
                            "안 맞는 22행 중 15행이 answer_3 가 종성인 3단 조합이었다"])
    log.append(["확인", "15행 전부 compose(초성,중성,종성) == 대상 음절이다 (Y3W479 ㅂ+ㅏ+ㄲ=밖)"])
    log.append(["남긴 것", "YK0002 7행은 답이 셋이지만 answer_3 가 조합 결과라 write 로 둔다"])
    log.append(["작성 2026-08-24", ""])
    wb.save(OUT)

    print(f"{OUT.name} — {len(changed)}행을 write3 로 고쳤다")
    print(f"  {', '.join(changed)}")

    chk = openpyxl.load_workbook(OUT, read_only=True)[SHEET]
    rows = list(chk.iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(rows[0]) if h}
    subs = Counter(r[ix["activity_sub"]] for r in rows[1:]
                   if any(x is not None and str(x).strip() for x in r))
    print(f"\n활동 분포: {dict(subs)}  (합 {sum(subs.values())})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
