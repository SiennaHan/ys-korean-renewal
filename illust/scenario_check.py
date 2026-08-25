#!/usr/bin/env python3
"""대화를 '턴 단위 존재 여부'가 아니라 '시나리오 단위'로 다시 판정한다.

앞선 턴 단위 검사(verify_content)는 하한선만 준다. 4급 10과 과제1은 대화가
통째로 교체됐는데도, 몇 턴이 같은 과의 다른 쪽(연습 문항 등)에서 우연히
발견되어 '유지'로 분류됐다. 그래서 교체 규모가 실제보다 작게 나온다.

여기서는 엑셀 시나리오의 턴 묶음을 신판에서 추출한 대화 묶음(쪽 단위)과
통째로 비교해서, 가장 비슷한 묶음과의 유사도로 판정한다.

산출: verify/scenario_check.csv
"""
import os, csv, re, difflib, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v3.xlsx"
OUT = f"{HERE}/verify"
INSTRUCTION = re.compile(r"하십시오|완성|고르십시오|쓰십시오|빈칸|알맞은")


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n2_ai_role_play"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    n2 = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    # 신판 추출 대화를 쪽 단위 묶음으로 (한 쪽 = 대화 하나로 본다)
    blocks = collections.defaultdict(list)
    for r in csv.DictReader(open(f"{OUT}/new_dialogue.csv")):
        if r["kind"] == "대화" and not INSTRUCTION.search(r["ko"]):
            blocks[(int(r["book"]), int(r["chapter"]), int(r["pdf_page"]))].append(r["ko"])

    scen = collections.defaultdict(list)
    for r in n2:
        if r.get("book_id") and r.get("scenario_id"):
            scen[(int(r["book_id"]), int(r["chapter"]), str(r["scenario_id"]))].append(
                (r.get("turn_seq") or 0, r.get("ko")))

    out = []
    for (b, ch, sid), turns in sorted(scen.items()):
        cands = {k: v for k, v in blocks.items() if k[0] == b and k[1] == ch}
        if not cands:
            continue                       # 추출 대상 과가 아니면 건너뛴다
        mine = sq(" ".join(t for _, t in sorted(turns)))
        best, best_pg = 0.0, None
        for k, v in cands.items():
            r = difflib.SequenceMatcher(a=mine, b=sq(" ".join(v)),
                                        autojunk=False).ratio()
            if r > best:
                best, best_pg = r, k[2]
        verdict = ("유지" if best >= 0.75 else
                   "부분 수정" if best >= 0.45 else "전면 교체")
        out.append(dict(book=b, chapter=ch, scenario=sid, turns=len(turns),
                        best_match_page=best_pg, similarity=round(best, 3),
                        verdict=verdict))

    with open(f"{OUT}/scenario_check.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)

    c = collections.Counter(o["verdict"] for o in out)
    print(f"추출 대상 과의 시나리오 {len(out)}개: {dict(c)}")
    for o in out:
        if o["verdict"] != "유지":
            print(f"  {o['book']}급 {o['chapter']:>2}과 {o['scenario']:<12} "
                  f"{o['turns']:>2}턴 유사도 {o['similarity']:.2f} → {o['verdict']}")
    print(f"-> {OUT}/scenario_check.csv")


if __name__ == "__main__":
    main()
