#!/usr/bin/env python3
"""읽기 지문이 '아직 안 배운 말'을 얼마나 쓰는지 본다.

n5 지문은 교재 원문이 아니라 창작이다. 주제가 맞는지는 별개로,
그 급·과까지 배운 어휘로 쓰였는지가 앱에서 실제로 걸리는 문제다.

누적 어휘 = 그 급 이전 급 전체 + 같은 급의 그 과까지. 여기에 없는 낱말이
얼마나 되는지 센다. 다만 조사·어미가 붙은 활용형은 원장 표제어와 문자열이
다르므로, 표제어의 앞 두 글자로 느슨하게 맞춘다(엄격히 보면 전부 미학습이 된다).

산출: verify/read_gate.csv
"""
import os, re, csv, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v22.xlsx"

# 문법 기능어·아주 흔한 말은 어휘 원장에 없어도 미학습으로 보지 않는다
COMMON = set("""그리고 그러나 하지만 그래서 그런데 또는 또한 이렇게 그렇게 저렇게 우리 저희
사람 사람들 때문 경우 정도 자신 서로 모두 여러 가지 이런 그런 저런 어떤 무슨 지금 요즘
오늘 내일 어제 이번 다음 처음 마지막 시작 아주 정말 매우 너무 조금 많이 함께 같이 바로
있다 없다 하다 되다 같다 많다 크다 좋다 나쁘다 아니다 보다 주다 받다 알다 모르다 만들다
생각 이야기 방법 문제 내용 사실 위해 통해 대해 대한 위한 중요 필요 가능""".split())


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def stem2(w):
    w = unicodedata.normalize("NFC", str(w or ""))
    w = re.sub(r"[(（][^)）]*[)）]", "", w)
    w = re.sub(r"[^가-힣]", "", w)
    return w[:2]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    wl = [r for r in sheet(wb, "n1_word_list")
          if r.get("book_id") and r.get("word") and str(r.get("review_status")) != "deleted"]
    # (급, 과) → 그때까지 배운 낱말 앞 두 글자 모음
    known = collections.defaultdict(set)
    for r in wl:
        known[(int(r["book_id"]), int(r["chapter"]))].add(stem2(r["word"]))

    def cum(b, c):
        s = set(COMMON)
        for (bb, cc), v in known.items():
            if bb < b or (bb == b and cc <= c):
                s |= v
        return {x for x in s if x}

    n5 = [r for r in sheet(wb, "n5_read_answer_text") if r.get("book_id")]
    target = {r["item_id"] for r in csv.DictReader(open(f"{HERE}/verify/read_text.csv"))
              if r["verdict"].startswith("양쪽과 무관")}

    out = []
    for r in n5:
        b, c = int(r["book_id"]), int(r["chapter"])
        pool = cum(b, c)
        words = re.findall(r"[가-힣]{2,}", str(r.get("text") or ""))
        uniq = {w for w in words}
        unseen = sorted({w for w in uniq if w[:2] not in pool})
        out.append(dict(item_id=r.get("item_id"), book=b, chapter=c,
                        낱말=len(uniq), 미학습=len(unseen),
                        비율=round(len(unseen) / max(len(uniq), 1), 2),
                        창작="O" if r.get("item_id") in target else "",
                        예=" ".join(unseen[:10])))
    with open(f"{HERE}/verify/read_gate.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)

    a = [o for o in out if o["창작"] == "O"]
    b_ = [o for o in out if o["창작"] != "O"]
    f_ = lambda g: sum(o["비율"] for o in g) / max(len(g), 1)
    print(f"창작 11건 미학습 비율 평균 {f_(a):.0%}  |  나머지 {len(b_)}건 평균 {f_(b_):.0%}")
    print("\n창작 11건")
    for o in sorted(a, key=lambda x: -x["비율"]):
        print(f"  {o['book']}급 {o['chapter']:>2}과 {o['item_id']:12s} "
              f"낱말 {o['낱말']:3d} 중 미학습 {o['미학습']:3d} ({o['비율']:.0%})")
        print(f"       {o['예']}")
    print(f"\n-> verify/read_gate.csv")


if __name__ == "__main__":
    main()
