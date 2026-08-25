#!/usr/bin/env python3
"""v9 최종 점검 — 교재 대조 + 시트 간 정합성.

두 가지를 본다.

A. 교재 대조 — 엑셀 문자열이 신판 본교재에 실제로 있는가.
   n1 어휘 / n2 대화는 이미 돌렸고, 여기서는 그동안 손대지 않은
   n1_word_quiz(단어 퀴즈)·n4_blank_question(문법)·n3_listen_repeat(듣기 문항)까지
   포함해 전 시트를 한 번에 본다.

B. 시트 간 정합성 — 앞선 작업으로 폐기·교체한 것이 다른 시트에 남아 있지 않은가.
   어휘를 폐기했는데 그 단어가 단어 퀴즈 보기에 그대로 있으면 학습자에게 노출된다.
   이건 교재와 무관한 우리 데이터 내부 문제라 따로 봐야 한다.

산출: verify/final_audit.csv
"""
import os, csv, re, json, collections, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v12.xlsx"
OUT = f"{HERE}/verify"


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    blob = {}
    for b in range(1, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        toc = json.load(open(f"{BASE}/work/book{b}/toc.json"))
        ranges = {int(k): v for k, v in toc["ranges"].items()}
        per = {}
        for ch, (a, z) in ranges.items():
            per[ch] = sq(" ".join(gp.text(p) for p in range(a, min(z, len(gp)) + 1)))
        blob[b] = per

    rows = []

    # ── A. 교재 대조 (그동안 안 본 시트 중심)
    def check(name, field, label, min_len=6):
        stat = collections.Counter()
        for r in sheet(wb, name):
            b, ch = r.get("book_id"), r.get("chapter")
            if not b or ch is None:
                continue
            if str(r.get("review_status")) == "deleted":
                stat["폐기(제외)"] += 1
                continue
            t = sq(r.get(field))
            if len(t) < min_len:
                stat["짧음(제외)"] += 1
                continue
            hit = t[:22] in blob[int(b)].get(int(ch), "")
            stat["확인" if hit else "미확인"] += 1
            if not hit:
                rows.append(dict(kind="교재 미확인", sheet=name, item_id=r.get("item_id"),
                                 book=b, chapter=ch, detail=str(r.get(field))[:80]))
        print(f"  [{label}] {dict(stat)}")
        return stat

    print("A. 교재 대조")
    check("n1_word_list", "word", "어휘", 2)
    check("n2_ai_role_play", "ko", "대화")
    check("n1_word_quiz", "prompt", "단어 퀴즈(발문)")
    check("n4_blank_question", "completion", "문법(완성문)")
    check("n3_listen_repeat", "question", "듣기 문항")
    check("n5_read_answer_text", "text", "읽기 지문", 20)

    # ── B. 시트 간 정합성
    print("\nB. 시트 간 정합성")
    wl = sheet(wb, "n1_word_list")
    dead = {sq(r["word"]) for r in wl if str(r.get("review_status")) == "deleted" and r.get("word")}
    live = collections.defaultdict(set)
    for r in wl:
        if r.get("word") and str(r.get("review_status")) != "deleted":
            live[(int(r["book_id"]), int(r["chapter"]))].add(sq(r["word"]))
    print(f"  폐기 어휘 {len(dead)}개가 다른 시트에 남아 있는지 확인")

    n = 0
    for r in sheet(wb, "n1_word_quiz"):
        for i in range(1, 5):
            v = sq(r.get(f"selection{i}"))
            if v and v in dead:
                rows.append(dict(kind="폐기 어휘 잔존", sheet="n1_word_quiz",
                                 item_id=r.get("item_id"), book=r.get("book_id"),
                                 chapter=r.get("chapter"),
                                 detail=f"보기{i} = {r.get(f'selection{i}')}"))
                n += 1
    for r in sheet(wb, "n3_listen_repeat"):
        for i in range(1, 5):
            v = sq(r.get(f"selection{i}"))
            if v and v in dead:
                rows.append(dict(kind="폐기 어휘 잔존", sheet="n3_listen_repeat",
                                 item_id=r.get("item_id"), book=r.get("book_id"),
                                 chapter=r.get("chapter"),
                                 detail=f"보기{i} = {r.get(f'selection{i}')}"))
                n += 1
    print(f"  퀴즈·듣기 보기에 남은 폐기 어휘: {n}건")

    # 이미지가 붙은 어휘가 폐기됐는지
    m = 0
    linked = {r["item_id"] for r in csv.DictReader(open(f"{HERE}/image_map.csv"))}
    for r in wl:
        if str(r.get("review_status")) == "deleted" and r.get("item_id") in linked:
            rows.append(dict(kind="폐기 어휘에 삽화 연결", sheet="n1_word_list",
                             item_id=r.get("item_id"), book=r.get("book_id"),
                             chapter=r.get("chapter"), detail=str(r.get("word"))))
            m += 1
    print(f"  폐기 어휘 중 삽화 연결 남은 것: {m}건")

    # 단어 퀴즈 정답이 그 과 어휘 원장에 살아 있는지
    k = 0
    for r in sheet(wb, "n1_word_quiz"):
        b, ch, ai = r.get("book_id"), r.get("chapter"), r.get("answer_index")
        if not b or ai is None:
            continue
        ans = sq(r.get(f"selection{int(ai) + 1}"))
        if ans and ans not in live.get((int(b), int(ch)), set()):
            rows.append(dict(kind="퀴즈 정답이 원장에 없음", sheet="n1_word_quiz",
                             item_id=r.get("item_id"), book=b, chapter=ch,
                             detail=r.get(f"selection{int(ai) + 1}")))
            k += 1
    print(f"  단어 퀴즈 정답이 그 과 어휘 원장에 없는 것: {k}건")

    with open(f"{OUT}/final_audit.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["kind", "sheet", "item_id", "book",
                                           "chapter", "detail"])
        wr.writeheader()
        wr.writerows(rows)
    print(f"\n총 {len(rows)}건 -> {OUT}/final_audit.csv")
    print(dict(collections.Counter(r["kind"] for r in rows)))


if __name__ == "__main__":
    main()
