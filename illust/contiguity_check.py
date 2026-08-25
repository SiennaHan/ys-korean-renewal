#!/usr/bin/env python3
"""대화가 신판에 '한 덩어리로' 남아 있는지 본다 — 거짓 음성 잡이.

턴 단위 문자열 검사만 하면, 대화가 통째로 바뀌었는데도 몇 턴이 같은 과의
연습 문항 쪽에서 우연히 발견돼 '유지'로 나온다(4급 10과 과제1).
반대로 시나리오 묶음 유사도는 내 대화 추출 품질에 의존해서 거짓 양성이 많다
(2급은 잔재가 0턴인데 유사도로는 '전면 교체'가 잔뜩 나온다).

그래서 추출에 기대지 않는 지표를 쓴다. 신판 과 본문을 읽기 순서대로 이어
붙인 문자열에서 각 턴의 위치를 찾는다. 대화가 그대로면 턴들이 순서대로,
좁은 구간 안에 모여 있다. 흩어져 있거나 순서가 뒤집혀 있으면 원래 대화가
아니라 다른 곳에서 우연히 걸린 것이다.

산출: verify/contiguity.csv
"""
import os, csv, re, json, collections, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v12.xlsx"
OUT = f"{HERE}/verify"
KEY = 20          # 턴마다 앞 20자로 위치를 찾는다
SPREAD = 4000     # 대화 하나가 차지할 만한 최대 구간(문자)


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n2_ai_role_play"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    n2 = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    out = []
    for b in range(2, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        toc = json.load(open(f"{BASE}/work/book{b}/toc.json"))
        ranges = {int(k): v for k, v in toc["ranges"].items()}
        # 쪽 단위로 본다. 과 전체를 이어 붙이면, 같은 문장이 뒤쪽 연습 문항에
        # 다시 나올 때 find()가 그쪽을 먼저 잡아 순서가 꼬인다. 실제로 그렇게
        # 잡힌 '순서 뒤바뀜' 20건은 전부 쪽 안에서는 순서가 멀쩡했다.
        pages = {}
        for ch, (a, z) in ranges.items():
            pages[ch] = [(p, sq(gp.text(p))) for p in range(a, min(z, len(gp)) + 1)]

        scen = collections.defaultdict(list)
        for r in n2:
            if r.get("book_id") == b and r.get("scenario_id"):
                scen[(int(r["chapter"]), str(r["scenario_id"]))].append(
                    (r.get("turn_seq") or 0, r.get("ko")))

        for (ch, sid), turns in sorted(scen.items()):
            keys = [sq(ko)[:KEY] for _, ko in sorted(turns)]
            keys = [k for k in keys if len(k) >= 8]
            n = len(keys)
            if n == 0:
                continue
            # 이 대화가 실린 쪽 = 턴이 가장 많이 잡히는 쪽
            best, best_hit = None, -1
            for pno, text in pages.get(ch, []):
                hit = sum(1 for k in keys if k in text)
                if hit > best_hit:
                    best, best_hit = (pno, text), hit
            pno, text = best
            pos = [text.find(k) for k in keys]
            missing = sum(1 for i in pos if i < 0)
            got = [i for i in pos if i >= 0]
            ordered = all(got[i] < got[i + 1] for i in range(len(got) - 1))

            if missing >= n * 0.6:
                v = "전면 교체"
            elif missing or not ordered:
                v = "일부 교체·재배치"
            else:
                v = "유지"
            out.append(dict(book=b, chapter=ch, scenario=sid, turns=n,
                            missing=missing, ordered=ordered, page=pno,
                            verdict=v))

    with open(f"{OUT}/contiguity.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=list(out[0]))
        wr.writeheader()
        wr.writerows(out)

    c = collections.Counter(o["verdict"] for o in out)
    print(f"시나리오 {len(out)}개: {dict(c)}")
    for o in out:
        if o["verdict"] != "유지":
            why = []
            if o["missing"]:
                why.append(f"미발견 {o['missing']}턴")
            if not o["ordered"]:
                why.append("순서 뒤바뀜")
            print(f"  {o['book']}급 {o['chapter']:>2}과 {o['scenario']:<13} "
                  f"{o['turns']:>2}턴 p{o['page']} → {o['verdict']}  ({', '.join(why)})")
    print(f"-> {OUT}/contiguity.csv")


if __name__ == "__main__":
    main()
