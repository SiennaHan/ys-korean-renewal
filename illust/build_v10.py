#!/usr/bin/env python3
"""어휘 정리(v5)를 단어 퀴즈에 반영해 v10을 만든다.

v5에서 어휘를 교체·폐기·이동했는데 n1_word_quiz에는 손을 안 댔다. 그 결과
폐기한 단어가 퀴즈 정답으로 남아 있고(9문항), 보기에도 남아 있다(17건).
학습자에게 더 이상 가르치지 않는 단어가 정답으로 나오는 상태다.

처리 방침
  교체된 단어    → 퀴즈 보기·정답을 새 표제어로 갱신. 같은 단어의 표기만 바뀐 것이라 안전하다.
  폐기 단어가 정답 → 문항 자체를 폐기한다. 정답이 학습 어휘가 아니면 문항이 성립하지 않는다.
  폐기 단어가 오답 → 자동 교체하지 않는다. 오답을 갈아 끼우면 복수정답이 생길 수 있어
                     hold_reason으로 표기하고 사람이 고른다.
  과 이동한 단어  → 퀴즈의 chapter도 같이 옮긴다.

산출: 글로벌_교재기반_콘텐츠_v10.xlsx
"""
import os, re, csv, shutil, collections, unicodedata
import openpyxl

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v9.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v10.xlsx"

REPLACE = {"귀고리": "귀걸이", "체력단련실": "헬스장", "심리 추리 영화": "스릴러 영화",
           "극단적이다": "극단적", "이중적이다": "이중적", "구세대": "기성세대",
           "중성적이다": "중성적", "현저하다": "현저히"}
MOVED = {"야영": 11, "핵심": 12, "대명사": 12}


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    t = re.sub(r"\(.*?\)", "", t)
    return re.sub(r"[\s.,·\[\]/]", "", t).strip()


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    ws = wb["n1_word_list"]
    Cw = {n: col(ws, n) for n in ("item_id", "book_id", "chapter", "word", "review_status")}
    dead = collections.defaultdict(set)
    live = collections.defaultdict(set)
    for row in range(2, ws.max_row + 1):
        b = ws.cell(row, Cw["book_id"]).value
        w = ws.cell(row, Cw["word"]).value
        if not b or not w:
            continue
        k = (int(b), int(ws.cell(row, Cw["chapter"]).value))
        if str(ws.cell(row, Cw["review_status"]).value) == "deleted":
            dead[k].add(sq(w))
        else:
            live[k].add(sq(w))

    ws = wb["n1_word_quiz"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "chapter", "answer_index",
                                 "selection1", "selection2", "selection3", "selection4",
                                 "review_status", "change_note", "hold_reason")}
    log = []
    n_rep = n_del = n_hold = n_mov = 0
    rev = {sq(k): v for k, v in REPLACE.items()}

    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        b = ws.cell(row, C["book_id"]).value
        if not iid or not b:
            continue
        ch = int(ws.cell(row, C["chapter"]).value)
        k = (int(b), ch)

        # 1) 교체된 표제어 반영
        changed = []
        for i in range(1, 5):
            v = ws.cell(row, C[f"selection{i}"]).value
            if v and sq(v) in rev:
                ws.cell(row, C[f"selection{i}"]).value = rev[sq(v)]
                changed.append(f"{v}→{rev[sq(v)]}")
        if changed:
            ws.cell(row, C["change_note"]).value = "어휘 표제어 교체 반영: " + ", ".join(changed)
            n_rep += 1
            log.append(["보기 표제어 갱신", iid, ", ".join(changed), "", ""])

        # 2) 정답이 폐기 어휘면 문항 폐기
        ai = ws.cell(row, C["answer_index"]).value
        ans = ws.cell(row, C[f"selection{int(ai) + 1}"]).value if ai is not None else None
        if ans and sq(ans) in dead[k]:
            ws.cell(row, C["review_status"]).value = "deleted"
            ws.cell(row, C["change_note"]).value = (
                f"정답 '{ans}'이 폐기 어휘라 문항 폐기")
            n_del += 1
            log.append(["문항 폐기", iid, str(ans), "", "정답이 폐기 어휘"])
            continue

        # 3) 과 이동한 단어가 정답이면 퀴즈도 같이 이동
        if ans and sq(ans) in {sq(x) for x in MOVED}:
            to = MOVED[next(x for x in MOVED if sq(x) == sq(ans))]
            if ch != to:
                ws.cell(row, C["chapter"]).value = to
                ws.cell(row, C["change_note"]).value = f"정답 '{ans}' 과 이동에 따라 {ch}과 → {to}과"
                n_mov += 1
                log.append(["문항 과이동", iid, f"{ch}과", f"{to}과", str(ans)])
                continue

        # 4) 오답 보기에 폐기 어휘가 남았으면 사람 판단
        left = [ws.cell(row, C[f"selection{i}"]).value for i in range(1, 5)
                if ws.cell(row, C[f"selection{i}"]).value
                and sq(ws.cell(row, C[f"selection{i}"]).value) in dead[k]]
        if left:
            ws.cell(row, C["hold_reason"]).value = (
                f"오답 보기에 폐기 어휘 있음({', '.join(map(str, left))}) — "
                "교체 시 복수정답 위험이라 사람이 고를 것")
            ws.cell(row, C["review_status"]).value = "draft"
            n_hold += 1
            log.append(["오답 보기 보류", iid, ", ".join(map(str, left)), "", ""])

    name = "99_변경내역_v10"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "v5 어휘 정리를 단어 퀴즈에 반영", "", "",
               "어휘만 고치고 퀴즈를 안 고쳐 폐기 어휘가 정답으로 남아 있었음"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((18, 16, 40, 24, 40), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"보기 표제어 갱신 {n_rep} / 문항 폐기 {n_del} / 문항 과이동 {n_mov} / "
          f"오답 보류 {n_hold}")
    for r in log:
        print(f"  [{r[0]}] {r[1]}  {r[2]}{' → ' + r[3] if r[3] else ''}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
