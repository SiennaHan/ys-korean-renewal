#!/usr/bin/env python3
"""대사는 그대로인데 화자만 뒤바뀐 대화를 찾는다.

7급 10과 본문이 그렇다. 구판은 t1이 '민철 엄마'인데 신판은 '옆집 아줌마'이고,
이후 턴도 계속 반대다. 문장 자체는 안 바뀌었으므로 문자열 검색으로는
'유지'로 나오지만(잔재 0/8), 앱에서는 화자·성별이 어긋나 TTS 목소리가
반대로 나온다.

방법: 신판 쪽에서 '화자 이름' 조각과 그 뒤 발화를 뽑아, 엑셀 각 턴의 문장이
어느 화자 밑에 있는지 확인한다. 발화를 문자열로 찾으므로 지면 구조가
달라도(어절이 span마다 쪼개진 7급 같은 경우) 견딘다.

산출: verify/speaker_check.csv
"""
import os, csv, re, json, difflib, collections, unicodedata
import openpyxl
from global_text import GlobalPdf

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "/Users/soohyeon/Documents/2606-yonsei3week_parse"
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v3.xlsx"
OUT = f"{HERE}/verify"


def sq(t):
    t = unicodedata.normalize("NFC", str(t or ""))
    return re.sub(r"[\s.,·’‘“”\"'()\[\]?!~\-—:;/]", "", t)


def page_speaker_map(gp, pno, names):
    """[(화자, 그 뒤 발화 덩어리)] — 화자 조각 사이의 텍스트를 발화로 본다."""
    items = []
    for txt, r, _ in gp.spans(pno):
        t = txt.strip().rstrip(":").strip()
        items.append((round(r.y0 / 5), r.x0, t))
    items.sort()
    # '민철 엄마'처럼 이름이 두 조각으로 쪼개져 있으면 붙여서 본다
    # (7급 10과 화자 전면 스왑을 이것 때문에 놓쳤다)
    merged, i = [], 0
    flat = [t for _, _, t in items if t]
    while i < len(flat):
        if i + 1 < len(flat) and sq(flat[i] + flat[i + 1]) in names:
            merged.append(flat[i] + " " + flat[i + 1]); i += 2
        else:
            merged.append(flat[i]); i += 1

    out, cur, buf = [], None, []
    for t in merged:
        if not t:
            continue
        if sq(t) in names:
            if cur:
                out.append((cur, sq("".join(buf))))
            cur, buf = t, []
        elif cur:
            buf.append(t)
    if cur:
        out.append((cur, sq("".join(buf))))
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n2_ai_role_play"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    n2 = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    out = []
    for b in range(2, 9):
        gp = GlobalPdf(f"{BASE}/(최종본)연세글로벌한국어_{b}급_본교재-최종(26.8.10).pdf")
        toc = json.load(open(f"{BASE}/work/book{b}/toc.json"))
        ranges = {int(k): v for k, v in toc["ranges"].items()}
        rows = [r for r in n2 if r.get("book_id") == b]
        by_sc = collections.defaultdict(list)
        for r in rows:
            by_sc[(int(r["chapter"]), str(r.get("scenario_id")))].append(r)

        for (ch, sid), turns in sorted(by_sc.items()):
            if ch not in ranges:
                continue
            names = {sq(str(r.get("speaker"))) for r in turns if r.get("speaker")}
            a, z = ranges[ch]
            smaps = []
            for pno in range(a, min(z, len(gp)) + 1):
                smaps += page_speaker_map(gp, pno, names)
            if not smaps:
                continue
            turns.sort(key=lambda r: r.get("turn_seq") or 0)
            mism = []
            for r in turns:
                key = sq(r.get("ko"))[:24]
                if len(key) < 8:
                    continue
                hit = next((sp for sp, body in smaps if key in body), None)
                if hit and sq(hit) != sq(str(r.get("speaker"))):
                    mism.append((r.get("item_id"), r.get("speaker"), hit))
            # 한 턴만 어긋난 것은 대개 발화 덩어리 경계를 잘못 잡은 것이다
            # (3급 3과가 그랬다 — 신판도 유리/슈테판 교대가 맞다).
            # 같은 치환(구 화자 → 신 화자)이 두 번 이상 반복될 때만 진짜로 본다.
            pairs = collections.Counter(
                (sq(str(o)), sq(n)) for _, o, n in mism)
            solid = {p for p, c in pairs.items() if c >= 2}
            real = [m for m in mism if (sq(str(m[1])), sq(m[2])) in solid]
            if real:
                out.append(dict(book=b, chapter=ch, scenario=sid,
                                turns=len(turns), mismatched=len(real),
                                swap=" · ".join(f"{o}→{n}" for o, n in
                                                sorted({(m[1], m[2]) for m in real})),
                                sample=" / ".join(m[0] for m in real[:4])))

    if out:
        with open(f"{OUT}/speaker_check.csv", "w", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=list(out[0]))
            wr.writeheader()
            wr.writerows(out)
    print(f"화자 불일치 시나리오 {len(out)}개")
    for o in out:
        flag = " ← 전면 스왑" if o["mismatched"] >= o["turns"] * 0.7 else ""
        print(f"  {o['book']}급 {o['chapter']:>2}과 {o['scenario']:<13} "
              f"{o['mismatched']}/{o['turns']}턴{flag}  {o['swap']}")
    print(f"-> {OUT}/speaker_check.csv")


if __name__ == "__main__":
    main()
