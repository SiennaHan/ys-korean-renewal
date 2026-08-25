#!/usr/bin/env python3
"""8급 3과 읽기 지문을 신판 기준으로 다시 쓴다 (v7).

구판 8급 3과 과제3은 영화 '아름다운 비행' 감상문이었고, 신판은 영화 '기생충'
감상문으로 통째로 바뀌었다. 지문에 딸린 어휘 13건이 잔재로 잡힌 것도 이 때문이다
(늪·부화시키다·서먹하다는 '아름다운 비행' 지문의 어휘였다).

n5는 교재 원문을 그대로 옮기지 않고 교재를 바탕으로 새로 쓴 지문이다. 그래서
여기서도 신판 감상문을 같은 방식으로 다시 썼다 — 다른 8급 지문과 비슷한 분량
(470~600자), 같은 문체.

문항 3개도 신판 지문에 맞춰 다시 썼다. 교재가 싣고 있는 문항을 따랐다.
지시문(instruction_*)은 유형별 표준 문구라 기존 행에서 그대로 가져온다.

산출: 글로벌_교재기반_콘텐츠_v7.xlsx
"""
import os, shutil
import openpyxl

ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v6.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v7.xlsx"

NEW_TEXT = (
    "영화 <기생충>은 2019년 한국에서 제작된 작품으로, 봉준호 감독이 각본과 연출을 "
    "맡았다. 이 영화는 세계적인 흥행을 거두었고 2020년 아카데미 영화제에서 작품상, "
    "감독상, 각본상, 국제 영화상까지 네 개 부문을 수상했다.\n"
    "이야기는 반지하 집에 사는 기택 가족에서 시작된다. 정규직 일자리가 없어 "
    "아르바이트로 살아가던 이들은 아들 기우가 학력을 속여 부유한 박 사장 집에 과외 "
    "선생으로 들어가면서 하나둘 그 집에 발을 들인다. 딸과 부부까지 가족 관계를 숨기고 "
    "이력을 꾸며 가사도우미와 운전기사로 일하게 된다. 그러나 이전 가정부 문광이 다시 "
    "찾아오면서 두 가족은 서로의 비밀을 알게 되고, 그 갈등은 걷잡을 수 없는 비극으로 "
    "이어진다.\n"
    "이 영화가 인상적인 것은 가난한 사람과 부유한 사람이 같은 사회에 살면서도 전혀 "
    "다른 세계에 있음을 보여 준다는 점이다. 기택 가족과 문광 부부는 거짓으로 차지한 "
    "자리를 지키려고 서로 다투지만, 정작 위층에 있는 사람들은 그 다툼에 거의 영향을 "
    "받지 않는다. 감독은 불평등한 구조 속에서 약자들끼리 부딪히게 되고 그것이 더 큰 "
    "비극으로 이어진다는 것을 풍자적으로 그려 낸다. 계층 문제를 개인의 노력이나 "
    "도덕이 아니라 사회가 함께 풀어야 할 문제로 제시한 작품이다."
)

# (item_id, question, type, s1..s4, answer_index)  — ox는 보기가 항상 X/O, 0=X 1=O
NEW_Q = {
    "RC-8-3-001": dict(
        question="이 사람의 영화에 대한 감상으로 가장 알맞은 것은 무엇입니까?",
        type="choice",
        s=("계층 간 갈등은 개인의 성격 차이에서 생긴다.",
           "부유한 가족이 의도적으로 가난한 가족을 괴롭혔다.",
           "가난한 사람들의 노력이 부족해서 비극이 일어났다.",
           "불평등한 사회 구조가 약자들 사이의 경쟁을 비극으로 만들었다."),
        ans=3),
    "RC-8-3-002": dict(
        question="글의 내용과 같으면 O, 다르면 X를 고르십시오. "
                 "'기택 가족은 일자리를 얻으려고 가족 관계를 숨기고 거짓말을 했다.'",
        type="ox", s=("X", "O", None, None), ans=1),
    "RC-8-3-003": dict(
        question="글의 내용과 같으면 O, 다르면 X를 고르십시오. "
                 "'이 영화는 계층 간의 갈등이 각자의 노력으로 쉽게 해결될 수 있다고 본다.'",
        type="ox", s=("X", "O", None, None), ans=0),
}


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    ws = wb["n5_read_answer_text"]
    C = {n: col(ws, n) for n in ("item_id", "text", "review_status",
                                 "change_note", "source_page")}
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, C["item_id"]).value == "RT-8-3-001":
            old = str(ws.cell(row, C["text"]).value or "")[:40]
            ws.cell(row, C["text"]).value = NEW_TEXT
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["source_page"]).value = 39
            ws.cell(row, C["change_note"]).value = (
                "신판 과제3 지문 교체 — 영화 '아름다운 비행' → '기생충' 감상문. 재저작")
            log.append(["읽기 지문 재저작", "RT-8-3-001", old,
                        NEW_TEXT[:40], "신판 p39 기준"])
            break

    ws = wb["n5_read_answer_questions"]
    Cq = {n: col(ws, n) for n in ("item_id", "question", "type", "selection1",
                                  "selection2", "selection3", "selection4",
                                  "answer_index", "review_status", "change_note")}
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, Cq["item_id"]).value
        if iid not in NEW_Q:
            continue
        q = NEW_Q[iid]
        old = str(ws.cell(row, Cq["question"]).value or "")[:40]
        ws.cell(row, Cq["question"]).value = q["question"]
        ws.cell(row, Cq["type"]).value = q["type"]
        for i, v in enumerate(q["s"], 1):
            ws.cell(row, Cq[f"selection{i}"]).value = v
        ws.cell(row, Cq["answer_index"]).value = q["ans"]
        ws.cell(row, Cq["review_status"]).value = "reviewed"
        ws.cell(row, Cq["change_note"]).value = "지문 교체에 따라 문항 재저작 (신판 p40 기준)"
        log.append(["읽기 문항 재저작", iid, old, q["question"][:40], ""])

    name = "99_변경내역_v7"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "8급 3과 읽기 지문 + 문항 3개", "", "",
               "전 급 117개 지문 중 원천이 갈린 것은 이 하나뿐"])
    sh.append(["판별", "지문 주제어가 신판에 남아 있는지", "구판 18/28", "신판 2/28",
               "n5는 창작 지문이라 문자열 대조가 안 된다. 주제 어휘 겹침으로 판정"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    for i, w in enumerate((18, 16, 42, 42, 40), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"지문 1건 + 문항 {len(log)-1}건 재저작 (본문 {len(NEW_TEXT)}자)")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
