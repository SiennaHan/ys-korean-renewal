#!/usr/bin/env python3
"""신판 교재에 그림+단어로 실려 있는데 엑셀 어휘 원장엔 없는 단어를 찾는다.

8급 11과처럼 신판에서 어휘가 통째로 갈린 자리가 있다.
(구판 '의료용 로봇/정사각형 코드/휘는 디스플레이' → 신판 '인공지능/자율주행/전기차/
증강현실/항공택시/지능형 공장'). 이런 데는 이미지가 틀린 게 아니라 엑셀이 구판이다.

그림 아래 단어가 찍힌 삽화만 본다 — 교재가 스스로 라벨을 붙인 것이라
'이 급 이 과의 학습 어휘'라는 근거가 확실하다.

산출: changed_vocab.csv
"""
import os, csv, re, collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v3.xlsx"


def norm(s):
    s = re.sub(r"\(.*?\)", "", str(s or ""))
    s = re.split(r"[(\[]", s)[0]
    return re.sub(r"[\s.,·]+", "", s).strip()


NOT_VOCAB = re.compile(r"보기|여러분|다음|아래|위와|같이|이야기|쓰세요|고르세요|문화|"
                       r"하십시오|하세요|합니까|입니까|예요|이에요|해요")
# 문장 종결(구두점 떼고 봄) — 어휘는 '먹다'처럼 기본형이고 '먹어요'가 아니다
SENTENCE_END = re.compile(r"(까|요|죠|네|자|오|시오)$")


def is_vocab_label(m):
    """어휘 박스 그림의 단어 라벨인가.

    relabel.vocab_label이 위치·간격으로 이미 걸렀으므로 여기서는 '어휘가 아닌
    문자열'만 떨군다: 문장·문법 형태·번호·깨진 글자.
    """
    lab = str(m["vocab_label"]).strip()
    if not lab or not (2 <= len(lab) <= 14):
        return False
    if "�" in lab:
        return False
    if any(c in lab for c in ".?!:;～~"):          # 문장·표 캡션
        return False
    if lab.startswith("-") or "/" in lab:          # '-을/ㄹ' 같은 문법 형태
        return False
    if NOT_VOCAB.search(lab):
        return False
    if SENTENCE_END.search(re.sub(r"[^가-힣]+$", "", lab)):
        return False
    if lab.count(" ") > 2:
        return False
    return float(m["text_cover"]) < 0.2


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["n1_word_list"]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    # 엑셀 어휘: (급, 과) 단위와 급 단위 둘 다 본다.
    # 과가 옮겨간 것뿐이면 '신규'가 아니므로 구분해서 표시한다.
    per_ch = collections.defaultdict(set)
    per_book = collections.defaultdict(set)
    for r in rows:
        if not r.get("book_id") or not r.get("word"):
            continue
        b, ch = int(r["book_id"]), int(r["chapter"])
        per_ch[(b, ch)].add(norm(r["word"]))
        per_book[b].add(norm(r["word"]))

    out = []
    for b in range(1, 9):
        for m in csv.DictReader(open(f"{HERE}/manifest_b{b}.csv")):
            if not is_vocab_label(m):
                continue
            lab = norm(m["vocab_label"])
            ch = int(m["chapter"])
            if lab in per_ch[(b, ch)]:
                continue
            out.append(dict(book=b, chapter=ch, ch_title=m["ch_title"],
                            book_word=m["vocab_label"], asset=m["filename"],
                            pdf_page=m["pdf_page"],
                            status="다른 과에 있음" if lab in per_book[b] else "엑셀에 없음"))

    # 같은 단어가 여러 그림에 걸리면 첫 장만
    seen, uniq = set(), []
    for o in out:
        k = (o["book"], o["chapter"], norm(o["book_word"]))
        if k in seen:
            continue
        seen.add(k)
        uniq.append(o)

    with open(f"{HERE}/changed_vocab.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["book", "chapter", "ch_title", "book_word",
                                           "status", "pdf_page", "asset"])
        wr.writeheader()
        wr.writerows(uniq)

    miss = [o for o in uniq if o["status"] == "엑셀에 없음"]
    print(f"교재엔 그림+단어로 있는데 해당 과 어휘 원장엔 없는 것: {len(uniq)}건")
    print(f"  그 중 그 급 어디에도 없음(신규/교체 의심): {len(miss)}건")
    per = collections.Counter(o["book"] for o in miss)
    print("  급별:", dict(sorted(per.items())))
    print(f"-> {HERE}/changed_vocab.csv")


if __name__ == "__main__":
    main()
