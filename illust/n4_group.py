#!/usr/bin/env python3
"""n4 문항을 '문법 항목' 단위로 묶는다.

태그 원칙(문법태그_1급_확정_v3.xlsx 안내):
  1) 이형태는 한 태그로 묶는다 — 은/는, 을/를, -을까요/-ㄹ까요
  2) 표면형이 같아도 교재가 다른 과에서 다른 기능으로 가르치면 나눈다
     (에: 장소 7과 / 시간 12과 / 방향 14과)

그래서 묶는 열쇠는 (급, 과, 이형태를 벗긴 정답)이다. 과를 열쇠에 넣는 것이
원칙 2를 자동으로 지켜 준다.

1급은 이미 태그가 붙어 있으므로, 이 묶기가 1급에서 기존 22개 태그를
그대로 재현하는지로 규칙을 검증한다.
"""
import re, unicodedata, collections
import openpyxl

XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v19.xlsx"

# 이형태 짝 — 앞의 것으로 통일한다. 긴 것부터 본다.
ALLO = [
    ("습니다", ["ㅂ니다"]), ("습니까", ["ㅂ니까"]),
    ("읍시다", ["ㅂ시다"]),
    ("어요", ["아요", "여요"]), ("어서", ["아서", "여서"]),
    ("어", ["아", "여"]), ("어야", ["아야", "여야"]),
    ("어도", ["아도", "여도"]), ("어서요", ["아서요"]),
    ("었", ["았", "였"]),
]
# 조사 이형태 짝 — 받침 유무로만 갈리는 것들. 하나로 묶는다.
PARTICLE = {
    "은": "은/는", "는": "은/는",
    "이": "이/가", "가": "이/가",
    "을": "을/를", "를": "을/를",
    "과": "과/와", "와": "과/와",
    "으로": "으로/로", "로": "으로/로",
    "이에요": "이에요/예요", "예요": "이에요/예요",
    "이야": "이야/야", "야": "이야/야",
    "이나": "이나/나", "나": "이나/나",
    "이랑": "이랑/랑", "랑": "이랑/랑",
    "이라서": "이라서/라서", "라서": "이라서/라서",
    "이라고": "이라고/라고", "라고": "이라고/라고",
    "이라는": "이라는/라는", "라는": "이라는/라는",
    "이란": "이란/란", "란": "이란/란",
    "이든지": "이든지/든지", "든지": "이든지/든지",
    "이며": "이며/며", "며": "이며/며",
    "이고": "이고/고", "고": "이고/고",
    "이 아니에요": "이/가 아니에요", "가 아니에요": "이/가 아니에요",
}


def norm_answer(s):
    t = unicodedata.normalize("NFC", str(s or "")).strip()
    t = re.sub(r"[?.!]", "", t).strip()
    t = re.sub(r"^[-–]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    for keep, alts in ALLO:
        for a in alts:
            if t == a or t.startswith(a):
                t = keep + t[len(a):]
                break
    # 앞머리 자음 이형태 통일: -ㄹ까요→-을까요, -ㄴ데→-은데, -ㅂ니다→-습니다
    if t.startswith("으") and len(t) > 1:      # -으세요/-세요, -으러/-러
        t = t[1:]
    if t.startswith("ㄹ"):
        t = "을" + t[1:]
    elif t.startswith("ㄴ"):
        t = "은" + t[1:]
    elif t.startswith("ㅂ"):
        t = "습" + t[1:]
    if t in PARTICLE:
        return PARTICLE[t]
    return t


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n4_blank_question") if r.get("book_id")]

    # ── 검증: 1급에서 기존 태그를 재현하는가
    g1 = [r for r in rows if int(r["book_id"]) == 1 and r.get("grammar_tag")]
    key2tag = collections.defaultdict(set)
    for r in g1:
        key2tag[(int(r["chapter"]), norm_answer(r["answer"]))].add(r["grammar_tag"])
    split = {k: v for k, v in key2tag.items() if len(v) > 1}
    tag2key = collections.defaultdict(set)
    for k, v in key2tag.items():
        for t in v:
            tag2key[t].add(k)
    merged = {t: v for t, v in tag2key.items() if len(v) > 1}
    print(f"1급 검증 — 묶음 {len(key2tag)}개 / 기존 태그 {len(tag2key)}개")
    print(f"  한 묶음이 두 태그로 갈림(과분류): {len(split)}건 {list(split.items())[:4]}")
    print(f"  한 태그가 여러 묶음에 걸침(정상일 수 있음): {len(merged)}건")
    for t, v in list(merged.items())[:6]:
        print(f"     {t}: {sorted(v)}")

    # ── 2~8급 문법 항목 목록
    print("\n2~8급 묶음")
    un = [r for r in rows if not r.get("grammar_tag")]
    inv = collections.defaultdict(list)
    for r in un:
        inv[(int(r["book_id"]), int(r["chapter"]), norm_answer(r["answer"]))].append(r)
    print(f"  문항 {len(un)} → 묶음 {len(inv)}개 "
          f"(과당 평균 {len(inv)/len({(k[0],k[1]) for k in inv}):.1f})")
    return inv


if __name__ == "__main__":
    main()
