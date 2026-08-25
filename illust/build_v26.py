#!/usr/bin/env python3
"""v26 — 자모 529행 복원. v25 의 129행에 사라진 400행을 잇는다.

무슨 일이 있었나
    v24(8/20)의 n8_jamo 529행이 build_v24.py 의 shutil.copy(v23, v24) 에
    덮여 사라졌다. v23 계열은 자모가 1행(예시)뿐이다. v25 는 그 v24 에서
    만들었으므로 v25 에도 없고, v25 의 129행은 그 뒤에 따로 넣은 것이다.

왜 손실이 복구 가능한가
    포팅 원본 app/src/shared/data/problem.ts 가 529항목 그대로 살아 있고
    (git HEAD 와 바이트 동일) legacy_id 로 v25 와 맞춰진다. 대조해 보니
    v25 의 129행은 problem.ts 와 필드까지 같다 — choice_2 의 모음 격자,
    answer_3 의 조합 결과, 다른 항목을 가리키는 content_sound 까지.
    즉 v25 의 129행도 기계 변환이고, 남은 400행을 같은 규칙으로 만들면
    품질이 같다. 사라진 529행은 전부 review_status=draft 였으므로
    (BLOCKERS.md §2) 사람이 검수한 판단은 하나도 잃지 않았다.

파생 규칙 — 전부 실측으로 확인했다
    activity_sub  problem.ts 의 instructions 가 1:1 로 결정한다.
                  v25 가 덮은 모듈 11개에서 이 대응이 예외 없이 성립한다.
    jamo_group    module_code 가 5개씩 묶음을 이룬다(YK0001~0005 = 첫 묶음).
                  빈 번호 0033·0038 을 포함해 정확히 여덟 묶음이고,
                  app/src/shared/data/unit.ts 의 자모 묶음 여덟과 맞는다.
    chapter       unit.ts 의 chapter_id 를 그대로 쓴다.

unit.ts 를 묶음 이름의 정본으로 쓴다
    v25 의 129행은 모음1 을 8자로 적었다("ㅏ,ㅓ,ㅗ,ㅜ,ㅡ,ㅣ,ㅚ,ㅟ").
    unit.ts 는 10자다 — BLOCKERS.md §2 가 2026-08-21 에 ㅐ·ㅔ 를 더한
    기록이 있다. 새 책 기준이 unit.ts 이므로 529행 전부 그쪽으로 맞춘다.
    바뀐 129행은 실행 끝에 찍는다.

돌리기
    python3 illust/build_v26.py

대상이 이미 있으면 아무것도 하지 않고 멈춘다. 이번 사고의 원인이
바로 그 확인이 없었던 것이다.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "글로벌_교재기반_콘텐츠_v25.xlsx"
OUT = ROOT / "글로벌_교재기반_콘텐츠_v26.xlsx"
PROBLEM_TS = ROOT / "app" / "src" / "shared" / "data" / "problem.ts"
UNIT_TS = ROOT / "app" / "src" / "shared" / "data" / "unit.ts"

SHEET = "n8_jamo"

# 묶음 안 모듈 위치가 활동을 정한다. 묶음마다 모듈이 정확히 다섯이고
# (BLOCKERS.md §2 "묶음마다 5개") 라우트가 sub 로 주소를 잡는다.
#
# 지시문으로 가르면 안 된다 — 원본의 지시문이 틀린 곳이 있다. YK0023 은
# 62행 전부 낱말(꼬리·토끼·꾸미다)이고 위치가 3번(단어 듣고 따라하기)인데
# 50행의 지시문이 "듣고 따라서 발음하세요"(낱자용)로 적혀 있다.
# 그래서 위치를 쓰고, 지시문이 어긋나면 경고만 찍는다.
SUB_BY_POSITION = {
    1: "listen-repeat",
    2: "write",
    3: "listen-repeat2",
    4: "read-write",
    5: "listen",
}

# 교차 확인용. 위치와 어긋나면 원본 지시문이 틀린 것으로 본다.
SUB_BY_INSTRUCTION = {
    "듣고 따라서 발음하세요.": "listen-repeat",
    "자음과 모음을 조합하고 써 보세요": "write",
    "단어를 듣고 따라서 발음하세요.": "listen-repeat2",
    "단어를 한 글자씩 써 보세요.": "read-write",
    "듣고 맞는 것을 고르세요.": "listen",
}

# 낱자 칸과 낱말 칸을 가르는 기준은 활동이 아니라 글자 수다 — v25 를
# 전수로 재 보니 target_jamo 는 97행 전부 1글자, target_word 는 22행이
# 2~3글자이고 둘 다 찬 행이 없다. 그 규칙을 그대로 쓴다.
JAMO_MAX_LEN = 1

# 이 스크립트가 스스로 정하는 칸. v25 값으로 덮지 않는다.
OWNED = {
    "item_id", "book_id", "chapter", "jamo_group", "activity_sub",
    "target_jamo", "target_word", "instruction", "problem_type",
    "review_status", "change_note",
}


def die(msg: str) -> None:
    print(f"멈춤 — {msg}", file=sys.stderr)
    raise SystemExit(1)


def load_problems() -> list[dict[str, str]]:
    """problem.ts 의 배열을 그대로 읽는다. JSON 이라 파싱이 안전하다."""
    text = PROBLEM_TS.read_text(encoding="utf-8")
    m = re.search(r"=\s*(\[.*?\])\s*;?\s*$", text, re.S)
    if not m:
        die(f"{PROBLEM_TS.name} 에서 배열을 못 찾았다")
    items = json.loads(m.group(1))
    if len(items) != 529:
        die(f"{PROBLEM_TS.name} 항목이 {len(items)}개다 — 529 를 기대했다")
    return items


def load_units() -> list[dict]:
    """unit.ts 에서 자모 묶음 여덟을 순서대로 읽는다."""
    text = UNIT_TS.read_text(encoding="utf-8")
    m = re.search(r"=\s*(\[.*?\])\s*;?\s*$", text, re.S)
    if not m:
        die(f"{UNIT_TS.name} 에서 배열을 못 찾았다")
    units = json.loads(m.group(1))
    jamo = [u for u in units if re.match(r"(모음|자음|받침|겹받침)", u["title"])]
    if len(jamo) != 8:
        die(f"unit.ts 의 자모 묶음이 {len(jamo)}개다 — 여덟을 기대했다")
    return jamo


def module_block(module_code: str) -> int:
    """YK0001~0005 → 0, YK0006~0010 → 1 …  다섯씩 묶는다."""
    n = int(module_code[2:])
    return (n - 1) // 5


def module_position(module_code: str) -> int:
    """묶음 안에서 몇 번째 모듈인가 (1~5). 이것이 활동을 정한다."""
    return (int(module_code[2:]) - 1) % 5 + 1


def main() -> int:
    if OUT.exists():
        die(
            f"{OUT.name} 이 이미 있다. 지우거나 다른 번호를 써라.\n"
            f"         v24 를 잃은 원인이 이 확인이 없었던 것이다."
        )
    if not SRC.exists():
        die(f"{SRC.name} 이 없다")

    problems = load_problems()
    units = load_units()

    # ── 원본을 뜨고 그 사본만 손댄다
    shutil.copy(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    if SHEET not in wb.sheetnames:
        die(f"{SRC.name} 에 {SHEET} 시트가 없다")
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header) if h}
    for need in ("legacy_id", "module_code", "activity_sub", "jamo_group", "chapter"):
        if need not in col:
            die(f"{SHEET} 에 {need} 열이 없다")

    # ── 기존 129행을 legacy_id 로 잡아 둔다 (대조용)
    before: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        lid = row[col["legacy_id"]]
        if lid:
            before[str(lid).strip()] = {h: row[col[h]] for h in col}
    print(f"v25 {SHEET}: {len(before)}행")

    # ── 529행을 새로 짓는다
    by_block: dict[int, list[dict]] = defaultdict(list)
    for p in problems:
        by_block[module_block(p["module_code"])].append(p)
    if sorted(by_block) != list(range(8)):
        die(f"모듈 묶음이 {sorted(by_block)} 다 — 0~7 여덟을 기대했다")

    rows: list[list] = []
    seq_in_group: Counter[int] = Counter()
    carried: Counter[str] = Counter()
    instr_conflict: Counter[tuple] = Counter()
    for block in range(8):
        unit = units[block]
        chapter = unit["chapter_id"]
        group = unit["title"]
        for p in by_block[block]:
            instr = p.get("instructions", "").strip()
            sub = SUB_BY_POSITION.get(module_position(p["module_code"]))
            if sub is None:
                die(f"{p['id']} 의 모듈 위치를 활동에 못 맺는다: {p['module_code']}")
            by_instr = SUB_BY_INSTRUCTION.get(instr)
            if by_instr is not None and by_instr != sub:
                instr_conflict[(p["module_code"], by_instr, sub)] += 1
            content = (p.get("content") or "").strip()
            seq_in_group[block] += 1
            rec = {h: None for h in col}
            rec["item_id"] = f"JM-{chapter}-{block + 1:02d}-{seq_in_group[block]:03d}"
            rec["book_id"] = 1
            rec["chapter"] = chapter
            rec["jamo_group"] = group
            rec["activity_sub"] = sub
            if content:
                if len(content) <= JAMO_MAX_LEN:
                    rec["target_jamo"] = content
                else:
                    rec["target_word"] = content
            rec["instruction"] = instr
            rec["problem_type"] = p.get("type") or None
            for n in (1, 2, 3):
                rec[f"choice_{n}"] = p.get(f"choice_{n}") or None
                rec[f"answer_{n}"] = p.get(f"answer_{n}") or None
            # 객관식은 낱자를 고르는 것이라 발음 칸을 비운다 — v25 와 같다
            rec["pronunciation"] = content or None if p.get("type") != "객관식" else None
            rec["content_img"] = p.get("content_img") or None
            rec["content_vid"] = p.get("content_vid") or None
            rec["content_sound"] = p.get("content_sound") or None
            rec["legacy_id"] = p["id"]
            rec["module_code"] = p["module_code"]
            rec["scene_num"] = str(p.get("scene_num", "")) or None
            rec["review_status"] = "draft"
            rec["change_note"] = (
                "problem.ts 에서 재포팅(v24 유실 복원)"
                if p["id"] not in before
                else "problem.ts 에서 재포팅 · v25 에도 있던 행"
            )
            # v25 에 있던 행은, problem.ts 로 못 만드는 칸을 지우지 않는다.
            # word_refs 가 그렇다 — 값은 content 와 80/80 같은데 어느 행에
            # 채우는지가 구글 시트 조인의 흔적이라 파생되지 않는다.
            # 내가 일부러 정한 칸(아래 OWNED)은 덮지 않는다.
            old = before.get(p["id"])
            if old:
                for h in col:
                    if h in OWNED:
                        continue
                    if rec[h] in (None, "") and old[h] not in (None, ""):
                        rec[h] = old[h]
                        carried[h] += 1
            rows.append([rec[h] for h in header])

    if len(rows) != 529:
        die(f"만든 행이 {len(rows)}개다 — 529 를 기대했다")

    # ── 시트를 비우고 새로 쓴다
    ws.delete_rows(2, ws.max_row)
    for r in rows:
        ws.append(r)
    wb.save(OUT)

    # ── 보고
    print(f"{OUT.name} 에 {SHEET} {len(rows)}행 썼다 (새로 복원 {529 - len(before)}행)\n")

    lid_i = col["legacy_id"]
    made = {r[lid_i]: r for r in rows}
    print("묶음별 · 활동별")
    for block in range(8):
        u = units[block]
        got = [r for r in rows if r[col["jamo_group"]] == u["title"]]
        subs = Counter(r[col["activity_sub"]] for r in got)
        print(f"  {u['chapter_id']}과 {u['title'][:22]:24} {len(got):3}행  {dict(subs)}")

    # v25 에 있던 129행이 그대로 재현되는지
    diffs: list[str] = []
    watch = [h for h in col if h not in
             ("item_id", "review_status", "change_note", "jamo_group")]
    for lid, old in before.items():
        if lid not in made:
            diffs.append(f"  {lid} 이 새 529행에 없다")
            continue
        new = made[lid]
        for h in watch:
            a, b = old[h], new[col[h]]
            sa = "" if a is None else str(a).strip()
            sb = "" if b is None else str(b).strip()
            if sa != sb:
                diffs.append(f"  {lid} {h}: v25={sa!r} → v26={sb!r}")
    print(f"\nv25 의 {len(before)}행 재현 대조 — 어긋난 칸 {len(diffs)}개")
    for d in diffs[:12]:
        print(d)
    if len(diffs) > 12:
        print(f"  … 그 밖에 {len(diffs) - 12}개")

    # 묶음 이름은 일부러 바꿨다
    renamed = sum(
        1 for lid, old in before.items()
        if lid in made and str(old["jamo_group"]) != str(made[lid][col["jamo_group"]])
    )
    if instr_conflict:
        print("\n원본 지시문이 모듈 위치와 어긋나는 곳 — 위치를 따랐다")
        for (mc, bi, sub), c in sorted(instr_conflict.items()):
            print(f"  {mc} {c:3}행  지시문→{bi} · 위치→{sub} (위치 채택)")
        print("  YK0023 은 62행 전부 낱말이라 위치(단어 듣고 따라하기)가 맞다.")

    if carried:
        print(f"\nv25 에서 이어받은 칸(problem.ts 로 못 만드는 것): {dict(carried)}")
        print("  word_refs 는 값이 content 와 같지만 v25 가 129행 중 80행만 채웠다.")
        print("  새 400행은 비워 두었다 — 전부 채울지 전부 비울지는 사람이 정할 일이다.")
    print(f"\n묶음 이름을 unit.ts 기준으로 맞춘 행: {renamed} (v25 의 모음1 은 8자, unit.ts 는 10자)")
    print("전부 review_status=draft 다 — 검수는 아직이다. BLOCKERS.md §2")
    return 0


if __name__ == "__main__":
    sys.exit(main())
