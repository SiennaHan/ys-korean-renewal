#!/usr/bin/env python3
"""v22 — completion이 문제와 어긋난 4건 정정 + 보기유형 재판정.

네 건 모두 같은 결함이다 — completion이 문제 문장을 채운 것이 아니라
다른(대개 짧은) 문장이었다. 앱은 completion을 정답 확인 화면에 보여 주므로
문제와 다르면 학습자가 혼란스럽다.

고치는 방향은 둘로 갈렸다.
  · 문제가 멀쩡하고 completion만 잘린 것 → completion을 문제에 맞춘다(2건)
  · 문제 쪽이 어색한 것 → 문제를 다듬고 completion을 맞춘다(2건)
    GF-2-11-006은 '비빔밥'이 두 번 나오는 순환문이었고,
    GF-2-15-008은 '매워서 매운 음식을'로 겹쳤다.

아울러 보기유형 판정에서 문자열 유사도를 걷어냈다. 1급 기록값을 보면
오형태가 0.33까지 내려가고(덥어요/더워요) 대조가 0.75까지 올라가(-지요?/-에요?)
잣대가 되지 못한다. 저작자가 focus에 남긴 표시만으로 1급 70/70을 맞춘다.

산출: 글로벌_교재기반_콘텐츠_v22.xlsx
"""
import shutil, datetime, collections
import openpyxl
from n4_inventory import distractor_type

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v21.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v22.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")


FIX = {
    # 문제가 순환문이었다 — '비빔밥인데 … 비빔밥이에요'. 문제를 completion 쪽으로 맞춘다.
    "GF-2-11-006": dict(
        question="지금 먹(     ) 음식이 비빔밥이에요.",
        completion="지금 <b>먹는</b> 음식이 비빔밥이에요.",
        change_note="문제가 '비빔밥'을 두 번 쓰는 순환문이라 정리하고 completion과 맞춤"),
    # 문제는 멀쩡하고 completion만 앞 절이 잘렸다.
    "GF-2-12-009": dict(
        completion="제 친구는 다른 외국어는 못하고 <b>한국어만</b> 잘해요.",
        change_note="completion에 빠져 있던 앞 절을 되살려 문제와 맞춤"),
    # 문제에 조사가 빠져 있었다.
    "GF-2-14-002": dict(
        question="제니 씨는 영국에서 왔잖아요. 그럼 영국 사람이(     )?",
        completion="제니 씨는 영국에서 왔잖아요. 그럼 영국 <b>사람이지요</b>?",
        change_note="문제의 '제니 씨'에 조사 '는'을 넣고 completion을 문제와 맞춤"),
    # '너무 매워서 매운 음식을'로 겹쳤다.
    "GF-2-15-008": dict(
        question="저는 매운 음식을 먹고 싶지만 너무 매워서( ) 먹어요.",
        completion="저는 매운 음식을 먹고 싶지만 너무 매워서 <b>못</b> 먹어요.",
        change_note="'매워서 매운 음식을' 겹침을 풀고 completion을 문제와 맞춤"),
}


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n4_blank_question"]
    hdr = [c.value for c in ws[1]]
    C = {k: hdr.index(k) + 1 for k in hdr if k}

    n = 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        if iid not in FIX:
            continue
        for k, v in FIX[iid].items():
            ws.cell(r, C[k]).value = v
        ws.cell(r, C["hold_reason"]).value = None
        ws.cell(r, C["review_status"]).value = "fixed_v22"
        n += 1
    print(f"completion 정정 {n}건")

    stat, changed = collections.Counter(), 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, C["book_id"]).value:
            continue
        row = {k: ws.cell(r, C[k]).value for k in C}
        d = distractor_type(row)
        if d and d != ws.cell(r, C["distractor_type"]).value:
            changed += 1
        if d:
            ws.cell(r, C["distractor_type"]).value = d
        stat[d] += 1
    print(f"보기유형 재판정: {dict(stat)} (바뀐 행 {changed})")

    lg = wb.create_sheet("99_변경내역_v22")
    lg.append(["구분", "item_id", "내용"])
    for iid, v in FIX.items():
        lg.append(["completion 정정", iid, v["change_note"]])
    lg.append(["보기유형 재판정", "", f"{changed}행 — 문자열 유사도 잣대 폐기(1급 70/70 유지)"])
    lg.append([f"작성 {datetime.date.today()}", "", ""])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
