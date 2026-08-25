#!/usr/bin/env python3
"""v21 — GF-3-14-002 교체 + 보기유형 재판정.

교체 사유: 3급 14과 문법은 '-는데2'와 '-어 보다2'인데 이 문항만 '-어 보이다'
(4급 9과)를 묻고 있었다. 3급 학습자가 아직 안 배운 문법이다.

무엇으로 채웠나: 같은 과의 '-는데2' 문항 세 개가 **전부 '-는데'**뿐이고
'-은데' 사례가 하나도 없었다. 이형태 문항은 양쪽 조건을 함께 놓아야 한다
(저작지침_복수정답 5번). 그래서 형용사(받침 O) + '-은데' 자리를 채운다.

  이 티셔츠는 디자인이 좋(     ) 치수가 좀 작아요.   → 좋은데
  오답 '-는데'는 형용사 '좋다'에 붙으면 비문('좋는데')이라 정답이 유일하다.

쓴 어휘는 모두 3급 14과 이전에 나온다 — 티셔츠·디자인(2급 10과),
치수(3급 14과), 좋다(1급 9과), 작다(2급 10과).

아울러 보기유형을 다시 판정한다. 품사로 갈리는 이형태(-는데/-은데)를
대조로 잘못 잡던 것을 고쳤다(1급 기록값 70/70 일치 유지).

산출: 글로벌_교재기반_콘텐츠_v21.xlsx
"""
import shutil, datetime, collections
import openpyxl
from n4_inventory import distractor_type

SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v20.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v21.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")


NEW = {
    "question": "이 티셔츠는 디자인이 좋(     ) 치수가 좀 작아요.",
    "selections": "-은데, -는데",
    "selection1": "-은데",
    "selection2": "-는데",
    "answer": "-은데",
    "answer_index": 0,
    "answer_text": "-은데",
    "completion": "이 티셔츠는 디자인이 <b>좋은데</b> 치수가 좀 작아요.",
    "grammar_focus": "좋다(형용사, 받침O ㅎ) → 좋 + -은데 → 좋은데. 형용사에 '-는데'는 비문.",
    "grammar_focus_수정": "형용사는 '-은데/-ㄴ데'를 써요. 좋다+은데=좋은데",
    "grammar_tag": "-는데2",
    "hold_reason": None,
    "review_status": "authored_v21",
    "change_note": ("-어 보이다(4급 9과)를 묻던 문항을 교체. 같은 과 -는데2 문항이 "
                    "전부 '-는데'뿐이라 형용사+'-은데' 사례로 채움"),
}


def col(ws, name):
    return [c.value for c in ws[1]].index(name) + 1


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n4_blank_question"]
    hdr = [c.value for c in ws[1]]
    C = {k: hdr.index(k) + 1 for k in hdr if k}

    done = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, C["item_id"]).value == "GF-3-14-002":
            for k, v in NEW.items():
                ws.cell(r, C[k]).value = v
            done = True
            break
    print("교체:", "완료" if done else "행 없음")

    # ── 보기유형 재판정
    stat = collections.Counter()
    changed = 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, C["book_id"]).value:
            continue
        row = {k: ws.cell(r, C[k]).value for k in C}
        d = distractor_type(row)
        if d and d != ws.cell(r, C["distractor_type"]).value:
            changed += 1
        if d:
            ws.cell(r, C["distractor_type"]).value = d
        stat[d] += 1
    print(f"보기유형 재판정: {dict(stat)} (바뀐 행 {changed})")

    lg = wb.create_sheet("99_변경내역_v21")
    lg.append(["구분", "item_id", "내용", "근거"])
    lg.append(["문항 교체", "GF-3-14-002", NEW["question"],
               "3급 14과 문법(-는데2·-어 보다2) 밖의 -어 보이다를 묻고 있었음"])
    lg.append(["보기유형 재판정", "", f"{changed}행 변경",
               "품사로 갈리는 이형태(-는데/-은데)를 대조로 잘못 잡던 것 교정"])
    lg.append([f"작성 {datetime.date.today()}", "", "", ""])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
