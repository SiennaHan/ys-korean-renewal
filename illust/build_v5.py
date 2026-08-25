#!/usr/bin/env python3
"""어휘 구판 잔재 48건을 정리해 v5를 만든다.

세 갈래로 나눈다.
  교체 8건 — 신판 본문에 대응어가 있고, 구어는 신판 어디에도 없음을 확인했다.
             (귀고리→귀걸이, 체력단련실→헬스장, 심리 추리 영화→스릴러 영화,
              극단적이다→극단적, 이중적이다→이중적, 중성적이다→중성적,
              현저하다→현저히, 구세대→기성세대)
  이동 3건 — 단어는 살아 있고 과만 옮겼다. v4에서 이미 표기했다.
  폐기 37건 — 신판 본문에 대응어가 없다.

폐기는 행을 지우지 않고 review_status='deleted'로 닫는다. item_id가 삽화
매핑에서 참조되고 있어서 행을 없애면 연결이 끊긴다. 되돌리기도 쉬워야 한다.

산출: 글로벌_교재기반_콘텐츠_v5.xlsx
"""
import os, csv, shutil, collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{ROOT}/글로벌_교재기반_콘텐츠_v4.xlsx"
DST = f"{ROOT}/글로벌_교재기반_콘텐츠_v5.xlsx"
V = f"{HERE}/verify"

REPLACE = {
    "VL-3-14-003": ("귀고리", "귀걸이"),
    "VL-5-3-017": ("체력단련실", "헬스장"),
    "VL-6-6-015": ("심리 추리 영화", "스릴러 영화"),
    "VL-7-7-003": ("극단적이다", "극단적"),
    "VL-7-7-017": ("이중적이다", "이중적"),
    "VL-7-11-003": ("구세대", "기성세대"),
    "VL-8-1-027": ("중성적이다", "중성적"),
    "VL-8-15-027": ("현저하다", "현저히"),
}
# 과만 옮긴 것 — 단어는 신판에 살아 있고 실린 과가 달라졌다.
# v4에서는 '구판 잔재' 판정이 먼저 걸려 이동 표기가 묻혔으므로 여기서 바로잡는다.
# 여러 과에서 발견되면 첫 등장 과를 따른다(1급 changelog의 규칙).
MOVED = {
    "VL-3-15-016": ("야영", 15, 11),
    "VL-7-9-026": ("핵심", 9, 12),
    "VL-8-1-012": ("대명사", 1, 12),
}


def col(ws, name):
    for i, c in enumerate(ws[1], 1):
        if c.value == name:
            return i
    return None


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    stale = {}
    for b in range(2, 9):
        for r in csv.DictReader(open(f"{V}/n1_b{b}.csv")):
            if r["verdict"] == "구판 잔재":
                stale[r["item_id"]] = (int(r["book"]), int(r["chapter"]), r["word"])

    # 삽화가 붙은 어휘인지 — 폐기 시 끊기는 연결을 세어 둔다
    linked = {r["item_id"] for r in csv.DictReader(open(f"{HERE}/image_map.csv"))}

    ws = wb["n1_word_list"]
    C = {n: col(ws, n) for n in ("item_id", "book_id", "chapter", "word", "image",
                                 "review_status", "change_note", "hold_reason")}
    log, drop_with_image = [], []
    n_rep = n_del = n_mov = 0
    for row in range(2, ws.max_row + 1):
        iid = ws.cell(row, C["item_id"]).value
        if iid in REPLACE:
            old, new = REPLACE[iid]
            ws.cell(row, C["word"]).value = new
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = f"신판 표제어 교체 {old}→{new}"
            n_rep += 1
            log.append(["어휘 교체", iid, old, new, "구어는 신판 전권에 없음을 확인"])
        elif iid in MOVED:
            w, frm, to = MOVED[iid]
            ws.cell(row, C["chapter"]).value = to
            ws.cell(row, C["review_status"]).value = "reviewed"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = f"신판에서 {frm}과 → {to}과 이동"
            n_mov += 1
            log.append(["어휘 과이동", iid, f"{frm}과", f"{to}과", w])
        elif iid in stale:
            b, ch, w = stale[iid]
            ws.cell(row, C["review_status"]).value = "deleted"
            ws.cell(row, C["hold_reason"]).value = None
            ws.cell(row, C["change_note"]).value = (
                "신판 본교재에 없어 폐기 — 행은 이력·삽화 연결 때문에 남겨 둠")
            n_del += 1
            note = ""
            if iid in linked:
                note = "삽화 연결 있음 — 이미지도 함께 정리 필요"
                drop_with_image.append((iid, w))
            log.append(["어휘 폐기", iid, w, "", note])

    # 변경내역
    name = "99_변경내역_v5"
    if name in wb.sheetnames:
        del wb[name]
    sh = wb.create_sheet(name)
    sh.append(["구분", "item_id", "구값", "신값", "비고"])
    sh.append(["", "", "", "", ""])
    sh.append(["범위", "n1 어휘 구판 잔재 48건 정리", "", "",
               "교체 8 / 과 이동 3 / 폐기 37"])
    sh.append(["폐기 방식", "행 삭제 아님", "review_status='deleted'", "",
               "item_id가 삽화 매핑에서 참조됨. 행을 지우면 연결이 끊긴다"])
    sh.append(["", "", "", "", ""])
    for r in log:
        sh.append(r)
    sh.append(["", "", "", "", ""])
    sh.append(["추가 발견", "8급 3과 본문 대화", "구판 대화", "유리·혜정 아이돌 컴백 대화",
               "어휘 13건이 통째로 갈린 것이 신호였다. 재저작 대상 — 아직 미반영"])
    for i, w in enumerate((14, 16, 30, 30, 60), 1):
        sh.column_dimensions[chr(64 + i)].width = w

    wb.save(DST)
    print(f"교체 {n_rep}건 / 과이동 {n_mov}건 / 폐기 {n_del}건")
    if drop_with_image:
        print(f"폐기 중 삽화 연결 있는 것 {len(drop_with_image)}건:")
        for iid, w in drop_with_image:
            print(f"   {iid} {w}")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
