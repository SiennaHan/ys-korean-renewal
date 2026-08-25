#!/usr/bin/env python3
"""부록 색인과 어휘 원장(n1_word_list)의 과 배정을 맞대 본다.

과 첫 쪽 '● 어휘' 목록은 본문을 고치고 갱신을 빠뜨려 못 믿을 물건이었다.
색인은 책 전체를 훑어 만든 것이라 그보다 낫고, 무엇보다 **교재가 스스로 밝힌
과 배정**이라 우리 원장을 검산할 정본이 된다.

색인은 한 낱말을 여러 과에 올리기도 한다(가게 14과 새어휘 / 가게 14과 주제어휘,
화장품 14과 과제어휘 / 화장품 15과 주제어휘). 그래서 '과가 다르다'는 판정은
그 낱말의 **색인 과 어디에도 없을 때만** 내린다.

산출: verify/index_check.csv
"""
import os, re, csv, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v18.xlsx"


def norm(w):
    """괄호 주석과 공백을 턴 대표형. '계시다(계세요)'와 '계시다'를 같게 본다."""
    w = unicodedata.normalize("NFC", str(w or ""))
    w = re.sub(r"[(（][^)）]*[)）]", "", w)
    return re.sub(r"[\s·]+", "", w).strip("-").strip()


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    wl = [r for r in sheet(wb, "n1_word_list")
          if r.get("book_id") and r.get("word")
          and str(r.get("review_status")) != "deleted"]

    idx = collections.defaultdict(set)          # (급, 낱말) -> {과}
    idx_cat = {}
    for r in csv.DictReader(open(f"{HERE}/verify/index_vocab.csv")):
        k = (int(r["book"]), norm(r["word"]))
        idx[k].add(int(r["chapter"]))
        idx_cat.setdefault(k, r["cat"])

    ours = collections.defaultdict(set)
    for r in wl:
        ours[(int(r["book_id"]), norm(r["word"]))].add(int(r["chapter"]))

    rows = []
    stat = collections.defaultdict(collections.Counter)
    for (b, w), chs in sorted(ours.items()):
        if not w:
            continue
        s = stat[b]
        if (b, w) not in idx:
            s["색인에 없음"] += 1
            rows.append(dict(kind="색인에 없음", book=b, word=w,
                             원장과="/".join(map(str, sorted(chs))), 색인과="", cat=""))
            continue
        ic = idx[(b, w)]
        if chs & ic:
            s["일치"] += 1
        else:
            s["과 불일치"] += 1
            rows.append(dict(kind="과 불일치", book=b, word=w,
                             원장과="/".join(map(str, sorted(chs))),
                             색인과="/".join(map(str, sorted(ic))),
                             cat=idx_cat[(b, w)]))
    for (b, w), ic in sorted(idx.items()):
        if (b, w) not in ours:
            stat[b]["원장에 없음"] += 1
            rows.append(dict(kind="원장에 없음", book=b, word=w, 원장과="",
                             색인과="/".join(map(str, sorted(ic))), cat=idx_cat[(b, w)]))

    with open(f"{HERE}/verify/index_check.csv", "w", newline="") as f:
        wr = csv.DictWriter(f, fieldnames=["kind", "book", "word", "원장과", "색인과", "cat"])
        wr.writeheader()
        wr.writerows(rows)

    print(f"{'':4s}{'일치':>7s}{'과 불일치':>10s}{'색인에 없음':>12s}{'원장에 없음':>12s}")
    for b in range(1, 9):
        s = stat[b]
        print(f"{b}급 {s['일치']:7d}{s['과 불일치']:10d}{s['색인에 없음']:12d}{s['원장에 없음']:12d}")
    print(f"\n-> verify/index_check.csv ({len(rows)}건)")
    print("\n■ 1급 과 불일치 (검증 구간 — 0에 가까워야 한다)")
    for r in rows:
        if r["book"] == 1 and r["kind"] == "과 불일치":
            print(f"   {r['word']}  원장 {r['원장과']}과 vs 색인 {r['색인과']}과 ({r['cat']})")


if __name__ == "__main__":
    main()
