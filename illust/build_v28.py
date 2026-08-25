#!/usr/bin/env python3
"""v28 — 신판과 어긋난 미션대화 상황 4건 교체.

미션대화는 문법을 겨누는 설계라 대부분 개정에 견딘다(117개 전건 확인). 그러나
아래 넷은 **상황 자체가 구판 본문을 전제**로 하고 있었다.

  8권 1과 — '남녀 임원 비율 기사'는 구판 소재다. 신판은 남성 육아 휴직과 성역할이다.
            첫 발화가 '여성 임원 비율 6.5%, 2019년 대비'라는 통계를 말하는데,
            판본 문제와 별개로 지금 쓰기에도 낡았다. 숫자를 빼고 다시 쓴다.
  8권 3과 — '바빠서 문화생활을 하기가 어려워요'는 대중문화를 이야기하지 않는 쪽으로 흐른다.
            신판은 한류 30년사와 영화 감상문이다. 바쁨(-으랴 -으랴)은 살리고
            대중문화를 즐기는 쪽으로 돌린다.
  8권 6과 — '공휴일 확대'는 학습목표 범위 안이지만 신판 본문·과제는 기념일과 기념일 마케팅이다.
  8권 11과 — 신판 본문은 뇌 과학·동물 복제다. '최근 기술 변화'는 과의 초점과 겹치지 않는다.

target_grammar 는 건드리지 않는다. 문법은 신구판이 같고 이 활동의 뼈대다.
문법이 자연스럽게 나올 상황을 고르는 것이 이 교체의 조건이다.

산출: 글로벌_교재기반_콘텐츠_v28.xlsx
"""
import os, sys, shutil, datetime
import openpyxl

BASE = "/Users/soohyeon/Documents/2608-yonsei_renewal"
SRC = f"{BASE}/글로벌_교재기반_콘텐츠_v27.xlsx"
DST = f"{BASE}/글로벌_교재기반_콘텐츠_v28.xlsx"

FIX = {
 # -긴 -나 봐요 / 아무리 -기로서니 — 추측과 양보가 나올 상황
 "MC-8-01-001": dict(
   situation_ko="남성 육아 휴직에 대한 뉴스를 보고 성역할이 어떻게 바뀌는지 이야기해요.",
   situation_en="You talk about changing gender roles after seeing news about men taking parental leave.",
   ai_first_line="방금 뉴스를 봤는데, 육아 휴직을 쓰는 남성이 늘고 있대요. 주변에도 그런 경우가 있어요?",
   mission_detail=("인식 변화: Infer how attitudes toward gender roles are changing / "
                   "망설이는 이유: Explain why men still hesitate to take parental leave / "
                   "한계선: State what should not be crossed / "
                   "기준: Explain what standard should still be kept"),
   change_note="구판 소재('남녀 임원 비율 기사')와 낡은 통계(2019년 대비 6.5%) 교체 — 신판 8급 1과는 남성 육아 휴직·성역할"),
 # -으랴 -으랴 / -으면 몰라도 — 바쁨과 조건이 나올 상황
 "MC-8-03-001": dict(
   situation_ko="바쁜 일상 속에서도 대중문화를 어떻게 즐기는지 이야기해요.",
   situation_en="You talk about how you enjoy popular culture despite a busy daily life.",
   ai_first_line="요즘 케이팝이나 한국 영화가 세계에서 인기가 많다던데요. 그런 거 보실 시간은 있으세요?",
   mission_detail=("바쁜 일상: Describe how busy your days are / "
                   "대중문화 소비: Talk about the movies, shows, or music you follow / "
                   "가능한 조건: Explain under what condition you could enjoy more"),
   change_note="'바빠서 문화생활을 하기가 어려워요'는 대중문화를 이야기하지 않는 쪽으로 흘렀음 — 바쁨은 살리고 대중문화로 돌림"),
 # -는다뿐이지 / -는답시고 — 인용과 의도가 나올 상황
 "MC-8-06-001": dict(
   situation_ko="젊은 사람들이 만든 기념일과 기념일 마케팅에 대해 의견을 나눠요.",
   situation_en="You exchange opinions about newly created anniversaries and anniversary marketing.",
   ai_first_line="요즘 빼빼로데이니 삼겹살데이니 기념일이 참 많아졌지요? 그런 날 챙기시는 편이에요?",
   mission_detail=("주장 인용: Refer to what others are saying about these days / "
                   "의도 설명: Explain why businesses create them / "
                   "결과 평가: Point out an unexpected or negative result / "
                   "개인 의견: State your stance on the issue"),
   change_note="'공휴일 확대'는 학습목표 범위 안이나 신판 본문·과제는 기념일과 기념일 마케팅"),
 # -는다는 듯이 / -기에 망정이지 — 태도 묘사와 아찔함이 나올 상황
 "MC-8-11-001": dict(
   situation_ko="동물 복제나 뇌 과학 같은 과학 연구에 대해 찬반 의견을 나눠요.",
   situation_en="You exchange opinions for and against scientific research such as animal cloning and brain science.",
   ai_first_line="동물 복제 연구가 많이 진행됐다는 기사를 봤어요. 이런 연구는 어떻게 생각하세요?",
   mission_detail=("연구 소개: Describe what the research does / "
                   "긍정·부정 영향: Explain both benefits and drawbacks / "
                   "미래 전망: Predict how it may develop"),
   change_note="'최근 기술 변화'는 과의 초점과 겹치지 않음 — 신판 8급 11과 본문은 뇌 과학·동물 복제"),
}


def main():
    if os.path.exists(DST):
        sys.exit(f"중단: {DST} 가 이미 있다. 덮어쓰지 않는다.")
    if not os.path.exists(SRC):
        sys.exit(f"중단: {SRC} 가 없다.")
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["n7_mission_chat"]
    hdr = [c.value for c in ws[1]]
    C = {k: hdr.index(k) + 1 for k in hdr if k}
    n = 0
    for r in range(2, ws.max_row + 1):
        iid = ws.cell(r, C["item_id"]).value
        if iid not in FIX:
            continue
        for k, v in FIX[iid].items():
            ws.cell(r, C[k]).value = v
        ws.cell(r, C["review_status"]).value = "authored_v28"
        n += 1
        print(f"  {iid} 교체")
    print(f"상황 교체 {n}건")

    lg = wb.create_sheet("99_변경내역_v28")
    lg.append(["item_id", "새 상황", "사유"])
    for iid, v in FIX.items():
        lg.append([iid, v["situation_ko"], v["change_note"]])
    lg.append([])
    lg.append(["원칙", "target_grammar 는 건드리지 않음 — 문법은 신구판 동일하며 이 활동의 뼈대. "
                     "문법이 자연스럽게 나올 상황을 고르는 것이 교체 조건"])
    lg.append([f"작성 {datetime.date.today()}", "", ""])
    wb.save(DST)
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
