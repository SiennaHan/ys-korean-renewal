#!/usr/bin/env python3
"""v25 — n7 미션대화·n8 자모를 글로벌 엑셀로 합친다.

두 활동은 '미착수'가 아니었다. 우리 엑셀 시트만 비어 있었고 콘텐츠는 별도
구글 시트 '3주완성 연세 한국어'(doc 1uS8pdPK…)에 앱 DB 스키마로 들어 있었다.
그 시트를 CSV로 뽑아(illust/gsheet/) 우리 활동 스키마로 옮긴다.

옮길 때 두 가지를 지킨다.
  · **원본 필드를 버리지 않는다.** 우리 스키마에 자리가 없는 것(ai_role,
    target_grammar, 문제 보기·음성 경로 등)은 컬럼을 늘려 담는다. n6 플래시카드가
    legacy_id로 구 시트를 이어 둔 선례를 따른다.
  · **없는 것을 만들지 않는다.** 일본어·중국어·베트남어 상황 설명은 원본에 없으므로
    빈칸으로 두고 hold_reason에 적는다.

커버리지 (원본 그대로):
  미션대화 117개 — 1권 4~15과 12개, 2~8권 각 15과. 전 8권. 프롬프트·미션 전건 보유.
  자모     129문제 — **1권 1과(한글1)에만 있다.** 자모 활동 38개 중 11개(YK0001~11)뿐이고
                    그 11개가 모두 1과 소속이다. 1과도 15개 활동 중 11개만 채워져 있고,
                    2과(한글2, YK0016~30)·3과(한글3, YK0031~40)는 문제가 하나도 없다.
                    빈 활동은 탭·낱말만 있거나 아무 것도 없다 → n8_자모_미보유 시트.

산출: 글로벌_교재기반_콘텐츠_v25.xlsx
"""
import csv, shutil, datetime, collections, os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v24.xlsx"
DST = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v25.xlsx"
# ── 덮어쓰기 방지 (2026-08-24 추가)
# build_v24.py 가 이미 있던 v24 를 덮어써서 자모 529행을 잃었다.
# shutil.copy 는 대상이 있어도 묻지 않고 휴지통도 거치지 않는다.
# 다시 돌릴 일이 있으면 대상을 먼저 치워라.
import os as _os
if _os.path.exists(DST):
    raise SystemExit(f"멈춤 — {_os.path.basename(DST)} 이 이미 있다. 지우거나 다른 번호를 써라.")

G = f"{HERE}/gsheet"

N7 = ["item_id", "book_id", "chapter", "scenario_title", "situation_ko", "situation_en",
      "situation_jp", "situation_cn", "situation_vi", "mission_keywords", "mission_detail",
      "ai_persona_prompt", "ai_first_line", "ai_role", "user_role", "ai_gender",
      "target_grammar", "level", "content_img", "video_refs", "legacy_id", "module_code",
      "review_status", "source_page", "change_note", "hold_reason"]

N8 = ["item_id", "book_id", "chapter", "jamo_group", "activity_sub", "target_jamo",
      "target_word", "word_refs", "instruction", "problem_type", "choice_1", "answer_1",
      "choice_2", "answer_2", "choice_3", "answer_3", "pronunciation",
      "content_img", "content_vid", "content_sound", "legacy_id", "module_code",
      "scene_num", "review_status", "source_page", "change_note", "hold_reason"]


def load(n):
    return list(csv.DictReader(open(f"{G}/{n}.csv")))


def write(wb, name, cols, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    return ws


def build_n7():
    cd = load("chat_dialog")
    ch = {(r["book_id"], r["seq"]): r for r in load("chapter")}
    miss = collections.defaultdict(list)
    for r in load("chat_mission"):
        miss[r["dialog_id"]].append(r)
    vid = collections.defaultdict(list)
    for r in load("chat_video"):
        vid[r["dialog_id"]].append(r)

    out = []
    for r in sorted(cd, key=lambda x: (int(x["book_id"]), int(x["chapter"]))):
        b, c = int(r["book_id"]), int(r["chapter"])
        ms = sorted(miss[r["id"]], key=lambda x: int(x["seq"]))
        out.append(dict(
            item_id=f"MC-{b}-{c:02d}-001", book_id=b, chapter=c,
            scenario_title=ch.get((str(b), str(c)), {}).get("title", ""),
            situation_ko=r["scenario"], situation_en=r["scenario_eng"],
            situation_jp=None, situation_cn=None, situation_vi=None,
            mission_keywords=";".join(m["keyword"] for m in ms),
            mission_detail=" / ".join(f"{m['keyword']}: {m['content']}" for m in ms),
            ai_persona_prompt=r["prompt"], ai_first_line=r["first_msg"],
            ai_role=r["ai_role"], user_role=r["user_role"], ai_gender=r["ai_gender"],
            target_grammar=r["target_grammar"], level=r["level"],
            content_img=r["content_img"],
            video_refs=";".join(v["youtube_id"] for v in vid[r["id"]]) or None,
            legacy_id=r["id"], module_code=r["module_code"],
            review_status="imported_v25",
            change_note="구글 시트 '3주완성 연세 한국어' chat_dialog/chat_mission/chat_video에서 이관",
            hold_reason=("상황 설명 일·중·베 미입력 — 원본에 없음(영어만 있음)")))
    return out


def build_n8():
    jp = load("jamo_problem")
    mod = {r["code"]: r for r in load("module")}
    un = {r["id"]: r for r in load("unit")}
    words = collections.defaultdict(list)
    for r in load("jamo_word"):
        words[r["problem_id"]].append(r["word"])

    seq = collections.Counter()
    out = []
    for r in jp:
        m = mod.get(r["module_code"], {})
        u = un.get(m.get("unit_id", ""), {})
        c = int(u.get("chapter_id") or 0)
        seq[c] += 1
        content = r["content"]
        out.append(dict(
            item_id=f"JM-1-{c:02d}-{seq[c]:03d}", book_id=1, chapter=c,
            jamo_group=u.get("title", ""), activity_sub=m.get("scene_type", ""),
            target_jamo=content if len(content) == 1 else None,
            target_word=content if len(content) > 1 else None,
            word_refs=";".join(words.get(r["id"], [])) or None,
            instruction=r["instructions"], problem_type=r["type"],
            choice_1=r["choice_1"], answer_1=r["answer_1"],
            choice_2=r["choice_2"], answer_2=r["answer_2"],
            choice_3=r["choice_3"], answer_3=r["answer_3"],
            pronunciation=r.get("정답 발음"),
            content_img=r["content_img"], content_vid=r["content_vid"],
            content_sound=r["content_sound"],
            legacy_id=r["id"], module_code=r["module_code"], scene_num=r["scene_num"],
            review_status="imported_v25",
            change_note="구글 시트 '3주완성 연세 한국어' jamo_problem/jamo_word에서 이관",
            hold_reason=None))
    return out


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    n7 = build_n7()
    n8 = build_n8()
    write(wb, "n7_mission_chat", N7, n7)
    write(wb, "n8_jamo", N8, n8)

    # 자모 활동 중 문제가 없는 것 — 빈 곳을 표로 남긴다
    mod = load("module")
    un = {r["id"]: r for r in load("unit")}
    ch = {r["id"] for r in load("chapter") if r.get("type") == "jamo"}
    have = {r["module_code"] for r in load("jamo_problem")}
    tabs = collections.Counter(r["module_code"] for r in load("jamo_tab"))
    gaps = [dict(module_code=m["code"], 과=un[m["unit_id"]]["chapter_id"],
                 자모묶음=un[m["unit_id"]]["title"], 활동=m["scene_type"],
                 제목=m["title"], 탭수=tabs.get(m["code"], 0))
            for m in mod if m["unit_id"] in un
            and un[m["unit_id"]]["chapter_id"] in ch and m["code"] not in have]
    write(wb, "n8_자모_미보유", ["module_code", "과", "자모묶음", "활동", "제목", "탭수"], gaps)

    lg = wb.create_sheet("99_변경내역_v25")
    lg.append(["구분", "내용"])
    lg.append(["n7 미션대화 이관", f"{len(n7)}건 — 1권 4~15과 12개, 2~8권 각 15개(전 8권)"])
    lg.append(["n8 자모 이관", f"{len(n8)}문제 — 1권 1과(한글1)에만. 활동 YK0001~YK0011"])
    lg.append(["n8 미보유", f"자모 활동 38개 중 {len(gaps)}개 문제 없음 — 1과 4개·2과 15개·3과 8개. n8_자모_미보유 시트"])
    lg.append(["출처", "구글 시트 '3주완성 연세 한국어' doc 1uS8pdPK313-9jnjJEho0VcyTxFNADoe6TkB55IWK4Bc"])
    lg.append(["원본 보존", "우리 스키마에 자리 없던 필드는 컬럼을 늘려 담음(ai_role·target_grammar·보기·음성 경로 등). legacy_id로 원본 id 연결"])
    lg.append(["미입력", "미션대화 상황 설명 일·중·베 117건 — 원본에 없어 저작 대상"])
    lg.append([f"작성 {datetime.date.today()}", ""])
    wb.save(DST)
    print(f"n7 미션대화 {len(n7)}건 · n8 자모 {len(n8)}문제 · 자모 미보유 활동 {len(gaps)}개")
    print(f"-> {DST}")


if __name__ == "__main__":
    main()
