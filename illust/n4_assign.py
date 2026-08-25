#!/usr/bin/env python3
"""각 n4 문항을 그 과의 '교재 문법 항목'에 배정한다.

교재 과 첫 쪽의 ● 문법 목록이 정답지다. 문항의 grammar_focus·정답·보기를
그 목록과 맞대어 가장 가까운 항목에 붙인다.

문자열을 그대로 대면 안 된다 — 정답은 활용형이고(‘봤을’, ‘어려운지’),
목록은 대표형이다(‘-을 뿐이다’, ‘어찌나 -는지’). 그래서 이형태 머리를
벗기고 남는 뼈대끼리 견준다.

1급은 태그가 이미 붙어 있으므로, 배정 결과가 태그와 일대일로 맞는지로 검증한다.
"""
import re, csv, os, collections, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/soohyeon/Documents/2608-yonsei_renewal/글로벌_교재기반_콘텐츠_v19.xlsx"

TASKY = re.compile(r"(읽고|듣고|쓰기|말하기|하기|합니까|십니까|하세요|드세요|%|\d{4})")


def gram_items(v):
    """과 첫 쪽 ● 문법 값에서 항목만 골라낸다. 과제 설명이 섞인 과가 있다."""
    out = []
    for tok in re.split(r"\s*/\s*|\s*,\s*", str(v or "")):
        t = tok.strip()
        if not t or len(t) > 24:
            continue
        if TASKY.search(t) and not t.startswith("-"):
            continue
        if not t.startswith("-") and t.endswith((".", "!", "?")) and len(t) > 6:
            continue
        out.append(t)
    return out


def skeleton(s):
    """이형태 머리를 벗기고 한글 뼈대만 남긴다."""
    t = unicodedata.normalize("NFC", str(s or ""))
    t = re.sub(r"[^가-힣ㄱ-ㅎ]", " ", t)
    MARK = ("으", "을", "ㄹ", "은", "ㄴ", "는", "습", "ㅂ", "았", "었", "였",
            "아", "어", "여", "이", "가", "기")
    parts = []
    for w in t.split():
        # '-을/ㄹ 뻔하다'는 '을 ㄹ 뻔하다'로 갈린다. 홀로 선 이형태 표지는 버린다.
        # 안 버리면 '을ㄹ뻔하다'가 되어 '-을 뻔하다'(을뻔하다)와 안 맞는다.
        if w in MARK:
            continue
        for head in MARK:
            if w.startswith(head) and len(w) > len(head):
                w = w[len(head):]
                break
        parts.append(w)
    return "".join(parts)


def score(q, item):
    """문항 쪽 문자열과 문법 항목의 뼈대가 얼마나 겹치는가."""
    a, b = skeleton(q), skeleton(item)
    if not a or not b:
        return 0
    if b in a or a in b:
        return len(b) + 10
    # 가장 긴 공통 조각
    best = 0
    for i in range(len(b)):
        for j in range(i + 2, len(b) + 1):
            if b[i:j] in a:
                best = max(best, j - i)
    return best


def sheet(wb, name):
    ws = wb[name]
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


# 파싱 때 빠진 과 — 교재 첫 쪽에서 직접 읽어 보강한다.
# 교재가 같은 형태를 번호로 갈라 가르치는 곳이 있다(-고1 4과 / -고2 8과).
SUPPLEMENT = {
    (2, 4): ["-고1", "보다"],
    (2, 5): ["-어서1", "-지만"],
    (2, 8): ["-고2", "에 ~쯤"],
    (2, 12): ["-어서2", "만"],
}


def load_syllabus():
    g = {}
    for r in csv.DictReader(open(f"{HERE}/syllabus.csv")):
        if r["field"] == "문법":
            g[(int(r["book"]), int(r["chapter"]))] = gram_items(r["value"])
    g.update(SUPPLEMENT)
    return g


# 불규칙 활용 항목은 대표형이 'ㄷ 동사'처럼 자모로만 적혀 있어 뼈대 대조가 안 된다.
# 대신 grammar_focus가 '(ㄷ불규칙)', 'ㄹ탈락'처럼 규칙 이름을 적어 준다.
IRREGULAR = [
    (re.compile(r"ㄷ\s*불규칙"), "ㄷ"), (re.compile(r"ㄹ\s*(동사|탈락)"), "ㄹ"),
    (re.compile(r"ㅎ\s*불규칙"), "ㅎ"), (re.compile(r"르\s*불규칙"), "르"),
    (re.compile(r"ㅅ\s*불규칙"), "ㅅ"), (re.compile(r"ㅂ\s*불규칙"), "ㅂ"),
    (re.compile(r"으\s*탈락"), "으"),
]


def by_irregular(r, items):
    g = str(r.get("grammar_focus") or "")
    for pat, jamo in IRREGULAR:
        if not pat.search(g):
            continue
        for it in items:
            if it.startswith(jamo) and ("동사" in it or "불규칙" in it):
                return it
    if "반말" in g:
        for it in items:
            if "반말" in it:
                return it
    return None


def assign(r, items):
    """문항을 항목 하나에 붙인다. focus를 먼저 보고, 없으면 정답을 본다."""
    hit = by_irregular(r, items)
    if hit:
        return hit, 10
    cands = [str(r.get("grammar_focus") or ""), str(r.get("answer") or ""),
             str(r.get("completion") or "")]
    best, bs = None, 0
    for it in items:
        s = max(score(c, it) for c in cands)
        if s > bs:
            best, bs = it, s
    return best, bs


def report(rows, syl):
    """2~8급 배정 품질 — 과별 항목 수와 배정 점수 분포."""
    un = [r for r in rows if not r.get("grammar_tag")]
    conf = collections.Counter()
    per = collections.defaultdict(collections.Counter)
    noitem = 0
    for r in un:
        b, c = int(r["book_id"]), int(r["chapter"])
        items = syl.get((b, c), [])
        if not items:
            noitem += 1
            continue
        it, s = assign(r, items)
        conf["강함(≥10)" if s >= 10 else ("보통(4~9)" if s >= 4 else "약함(<4)")] += 1
        per[(b, c)][it or "—"] += 1
    print(f"2~8급 문항 {len(un)} / 교재 문법 목록 없는 과의 문항 {noitem}")
    print("  배정 신뢰도:", dict(conf))
    sizes = [len(v) for v in per.values()]
    print(f"  과당 항목 수: 평균 {sum(sizes)/len(sizes):.1f} 최대 {max(sizes)}")
    print("\n  표본:")
    for k in sorted(per)[:8]:
        print(f"   {k[0]}급 {k[1]:>2}과: " + " | ".join(f"{a}({n})" for a, n in per[k].most_common()))


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    rows = [r for r in sheet(wb, "n4_blank_question") if r.get("book_id")]
    syl = load_syllabus()
    report(rows, syl)
    print()

    # ── 검증: 1급
    g1 = [r for r in rows if int(r["book_id"]) == 1 and r.get("grammar_tag")]
    m = collections.defaultdict(collections.Counter)
    unres = 0
    for r in g1:
        items = syl.get((1, int(r["chapter"])), [])
        it, s = assign(r, items)
        if not it or s < 2:
            unres += 1
            continue
        m[(int(r["chapter"]), it)][r["grammar_tag"]] += 1
    bad = {k: v for k, v in m.items() if len(v) > 1}
    print(f"1급 검증 — 문항 {len(g1)} / 배정 못 함 {unres}")
    print(f"  한 항목이 두 태그로 갈림: {len(bad)}건")
    for k, v in list(bad.items())[:8]:
        print(f"     {k[0]}과 '{k[1]}' → {dict(v)}")
    return syl, rows


if __name__ == "__main__":
    main()
